#!/usr/bin/python
# -*- coding: utf-8 -*-

from qgis.core import Qgis, QgsProject, QgsFeatureRequest, QgsExpression
import os
import openpyxl


class CapFt:
    """ Second pour contenir quelques fonctions qui seront utilisées dans le fichier principale CheckOptyce.py """

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def verificationsDonneesCapft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):

        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        request_cap_ft = QgsFeatureRequest()
        clause_cap_ft = QgsFeatureRequest.OrderByClause(colonne_cap_ft, ascending=True)
        orderby_cap_ft = QgsFeatureRequest.OrderBy([clause_cap_ft])
        request_cap_ft.setOrderBy(orderby_cap_ft)

        ###################### POT-BT n'ont situé dans une étude cap_ft #################
        pot_bt_hors_etude = []
        for feat_pot in infra_pt_pot.getFeatures(request):
            pot_dans_une_etude_cap_ft = False
            inf_num = feat_pot["inf_num"]

            for feat_cap_ft in etude_cap_ft.getFeatures(request_cap_ft):
                if feat_cap_ft.geometry().contains(feat_pot.geometry()):
                    pot_dans_une_etude_cap_ft = True
                    break

            if not pot_dans_une_etude_cap_ft:
                pot_bt_hors_etude.append(inf_num)

        ###################### 2 ou plusieurs études cap_ft ayant le même nom #################
        doublonsNomEtudesCapFt = []
        for feat_cap_ft_1 in etude_cap_ft.getFeatures(request_cap_ft):
            nomEtude = feat_cap_ft_1[colonne_cap_ft]
            for feat_cap_ft_2 in etude_cap_ft.getFeatures(request_cap_ft):
                if feat_cap_ft_1.id() != feat_cap_ft_2.id():
                    if nomEtude == feat_cap_ft_2[colonne_cap_ft]:
                        doublonsNomEtudesCapFt.append(nomEtude)
                        break

        return doublonsNomEtudesCapFt, pot_bt_hors_etude

    def liste_poteau_cap_ft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):
        dicoEtudeCapFtPoteauQgis = {}

        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        request_cap_ft = QgsFeatureRequest()
        clause_cap_ft = QgsFeatureRequest.OrderByClause(colonne_cap_ft, ascending=True)
        orderby_cap_ft = QgsFeatureRequest.OrderBy([clause_cap_ft])
        request_cap_ft.setOrderBy(orderby_cap_ft)

        for feat_cap_ft in etude_cap_ft.getFeatures(request_cap_ft):
            listePoteau = []
            etudes = feat_cap_ft[colonne_cap_ft]
            for feat_pot in infra_pt_pot.getFeatures(request):
                if feat_cap_ft.geometry().contains(feat_pot.geometry()):
                    inf_num = feat_pot["inf_num"]

                    listePoteau.append(inf_num)

            if listePoteau:
                dicoEtudeCapFtPoteauQgis[etudes] = listePoteau

        return dicoEtudeCapFtPoteauQgis

    def LectureFichiersExcelsCap_ft(self, repertoire):
        """Fonction pour parcourir les fichiers Excel pour renseigner la référence des appuis."""
        dicoPoteauFt_SousTraitant = {}

        for subdir, dirs, files in os.walk(repertoire):
            dossier_parent = ""
            listePoteauFt = []
            for name in files:

                if "FicheAppui_" in name and name.endswith('.xlsx'):
                    cheminComplet = subdir + os.sep + name
                    listePoteauFt.append(name)

                    ## NOM DOSSIER ?! Barbara
                    dossier_parent = os.path.split(os.path.split(os.path.split(cheminComplet)[0])[0])[1]
                    # print(dossier_parent)

            if listePoteauFt:
                dicoPoteauFt_SousTraitant[dossier_parent]=listePoteauFt

        return dicoPoteauFt_SousTraitant

    def traitementResultatFinauxCapFt(self, dicoEtudeCapFtPoteauQgis, dicoPoteauFt_SousTraitant):
        """Fonction pour traiter les résultats finaux des deux dictionnaires """
        dicoPotFt_Excel_Introuvable = {}
        dicoPotFtExistants = {}

        comptabilite = 0
        #print(dicoEtudeCapFtPoteauQgis)
        #(dicoPoteauFt_SousTraitant)
        #################### EXCEL SOUS-TRAITANT ###############################
        for excel, listePoteau in dicoPoteauFt_SousTraitant.items():
            PotFtintrouvableSt = []

            for poteauSt in listePoteau:
                potSt = str(poteauSt.replace("FicheAppui_", "").replace(".xlsx", ""))
                introuvable = True

                #################### ETUDE QGIS ###############################
                for etude_cap_ft, listePoteauxQgis in dicoEtudeCapFtPoteauQgis.items():
                    inf_num_trouve = ""
                    nouveauListePoteauxQgis = []
                    for inf_num_ft in listePoteauxQgis:
                        nouveauListePoteauxQgis = listePoteauxQgis.copy()
                        #position = inf_num_ft.find('FT-')
                        #inf_num = inf_num_ft[position+3:]
                        inf_num = inf_num_ft.split("/")[0]
                        #print(inf_num)
                        #print(excel, " , ", etude_cap_ft)

                        try:
                            # On vérifie les entiers d'abord
                            if int(inf_num) == int(potSt) and excel.upper() == etude_cap_ft.upper():

                                introuvable = False
                                comptabilite += 1

                                #################### VALEURS TROUVEES ########################################
                                dicoPotFtExistants[comptabilite] = [inf_num_ft, etude_cap_ft, poteauSt, excel]
                                inf_num_trouve = inf_num_ft
                                break

                        except Exception as e:
                            if str(inf_num) == str(potSt) and excel.upper() == etude_cap_ft.upper():
                                introuvable = False
                                comptabilite += 1

                                #################### VALEURS TROUVEES ########################################
                                dicoPotFtExistants[comptabilite] = [inf_num_ft, etude_cap_ft, poteauSt, excel]
                                inf_num_trouve = inf_num_ft
                                print(f"Une erreur\n{e}")
                                break

                            # print("Ce ne ne sont pas un entier. potSt : ", potSt,  " inf_num : ",  "inf_num")
                    ## Sortir du boucle si poteau trouve ##
                    if not introuvable:
                        nouveauListePoteauxQgis.remove(inf_num_trouve)

                        # S'il ne reste plus rien dans le dictionnaire, on supprime ce dictionnaire
                        if not nouveauListePoteauxQgis:
                            del dicoEtudeCapFtPoteauQgis[etude_cap_ft]

                        # Sinon, on supprime juste un élément de la liste du dictionnaire
                        else:
                            dicoEtudeCapFtPoteauQgis[etude_cap_ft]=nouveauListePoteauxQgis
                        break

                if introuvable:
                    PotFtintrouvableSt.append(poteauSt)

            if PotFtintrouvableSt:
                dicoPotFt_Excel_Introuvable[excel]=PotFtintrouvableSt

        return dicoPotFt_Excel_Introuvable, dicoEtudeCapFtPoteauQgis, dicoPotFtExistants   #

    def ecrireResultatsAnalyseExcelsCapFt(self, resultatsFinaux, nom):
        """Fonction qui permet d'écrire dans le fichier Excel à partir d'un fichier modèle'"""

        dicoPotFt_Excel_Introuvable = resultatsFinaux[0]
        dicoEtudeCapFtPotQgisIntrouvable = resultatsFinaux[1]
        dicoPotFtExistants = resultatsFinaux[2]

        fichierXlsx = openpyxl.workbook.Workbook()

        ############################## FEUILLE 1 : VALIDATION_COMPLET ################################################
        my_ligne = 1  # Ligne à partir de laquelle on commencera à écrire dans le fichier Annexe
        # Parcourir fichier C3A, pour comparer ses contenus aux fichiers C6
        feuille = fichierXlsx.create_sheet("ANALYSE CAP_FT", 0)

        alignement = openpyxl.styles.Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False,
                                               shrink_to_fit=True, indent=0)

        # Pour mettre du GRAS à partir de la colonne 18 et 23 pour les câbles.
        feuille.cell(row=1, column=1, value="INF_NUM QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=2, value="ETUDE QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=3, value="INF_NUM EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=4, value="NOM FICHIER EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=5, value="REMARQUES").font = openpyxl.styles.Font(bold=True)

        feuille.cell(row=1, column=1).alignment = alignement
        feuille.cell(row=1, column=2).alignment = alignement
        feuille.cell(row=1, column=3).alignment = alignement
        feuille.cell(row=1, column=4).alignment = alignement
        feuille.cell(row=1, column=5).alignment = alignement

        feuille.column_dimensions["A"].width = 25
        feuille.column_dimensions["B"].width = 30
        feuille.column_dimensions["C"].width = 25
        feuille.column_dimensions["D"].width = 50
        feuille.column_dimensions["E"].width = 40

        red_color = openpyxl.styles.colors.Color(rgb='fc4e2a')
        orange_color = openpyxl.styles.colors.Color(rgb='fd8d3c')

        ####################### SOUS-TRAITS INTROUVABLE #######################################"
        for fichierExcels, listesDesAppuisSous_traitant in dicoPotFt_Excel_Introuvable.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for inf_num_excel in listesDesAppuisSous_traitant:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=3, value=inf_num_excel).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=4, value=fichierExcels).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans QGIS").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)

        ####################### QGIS INTROUVABLE #######################################
        for etudeQgis, listesDesAppuisQGis in dicoEtudeCapFtPotQgisIntrouvable.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for inf_num in listesDesAppuisQGis:
                # Si tirage câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=1, value=inf_num).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=2, value=etudeQgis).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans les Fiches Appuis").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)

        my_colonne = 1
        my_ligne += 1

        ####################### POTEAUX TROUVEES #######################################")
        for etiquetteDuCable, listesDesAppuisBT in dicoPotFtExistants.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for valeur in listesDesAppuisBT:
                feuille.cell(row=my_ligne, column=my_colonne, value=valeur)
                my_colonne += 1

            feuille.cell(row=my_ligne, column=5, value="correspondance trouvée")

            # On passe à la ligne suivante
            # Si la première liste est terminé, on passe à la ligne suivante en récommençant par au toujours
            # au niveau de la colonne A
            my_ligne += 1
            my_colonne = 1

        std = fichierXlsx.get_sheet_by_name('Sheet')
        fichierXlsx.remove_sheet(std)

        fichierXlsx.save(filename=nom)   # Enregistrement du fichier
