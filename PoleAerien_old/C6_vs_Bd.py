#!/usr/bin/python
# -*- coding: utf-8 -*-

from qgis.core import Qgis, QgsProject, QgsFeatureRequest, QgsExpression, NULL
import os
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
# from tabulate import tabulate


class C6_vs_Bd:
    """ Comparaison des données C6 Par rapport à la base de données"""

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def LectureFichiersExcelsC6(self, df, repertoire):
        """Fonction pour parcourir les fichiers Excel pour renseigner la référence des appuis."""

        # headers = ['N° appui', 'Nature des travaux', 'Études']
        for subdir, dirs, files in os.walk(repertoire):
            for name in files:

                if name.endswith('.xlsx') and str("~$") not in str(name):
                    cheminComplet = subdir + os.sep + name

                    try:

                        with pd.ExcelFile(cheminComplet) as xls:
                            # On prend les données à partir de la ligne 7 (la colonne).
                            df1 = pd.read_excel(xls, "Export 1", header=7, index_col=None)  # , index_col=0,  skiprows=[15], na_values=["NA"],)

                        # Creation d'un Dataframe à partir des données collectées dans Excel
                        # En supprimant les lignes qui ont une plus de un np.NaN
                        # Les colonnes qui nous intéresse sont dans la
                        resultat = df1.iloc[:, 0:32:31].dropna(thresh=1)
                        # print("AV  :\n")
                        # print(tabulate(resultat.values, headers, tablefmt='psql'))
                        resultat = resultat[pd.to_numeric(resultat['N° appui'], errors='coerce').notnull()]

                        # On convertir les numéros des appuis en appuis.
                        resultat['N° appui'] = resultat['N° appui'].astype(int)

                        # On supprime les numéros des appuis qui sont pas des entiers
                        # resultat = resultat.loc[resultat['N° appui'].str.isnumeric()]
                        # print("\nAPPR  :\n")
                        # print(tabulate(resultat.values, headers, tablefmt='psql'))

                        # print("SECOND : ", resultat.loc[resultat['N° appui'].str.isnumeric()])
                        # print("resultat  :\n", resultat)

                        # On ajoute le nom du fichier dans une nouvelle colonne
                        resultat.insert(2, "Excel", name, True)
                        resultat.insert(3, "Études", name.replace(".xlsx", ""), True)
                        df = df.append(resultat)

                    except AttributeError as lettre:
                        print(f"FICHIER : {cheminComplet}\nLettre : {lettre}")
                        resultat = df1.iloc[:, 0:32:31].dropna(thresh=1)

                        resultat.insert(2, "Excel", name, True)
                        resultat.insert(3, "Études", name.replace(".xlsx", ""), True)
                        df = df.append(resultat)

                    except Exception as e:
                        print(f"Erreur avec ce fichier {cheminComplet}\nPrécision : {e}")

        # headers = ['N° appui', 'Nature des travaux', 'Études']
        # print(tabulate(df.values, headers, tablefmt='psql'))
        #
        # repertoire = f"C:/Users/asoumare/Documents/Teletravail/Developpement/CAP_COMAC/EXCEL.xlsx"
        # with pd.ExcelWriter(repertoire, engine="openpyxl") as writer:
        #     sheet_name = "EXCEL"
        #     # Export DataFrame content
        #     df.to_excel(writer, sheet_name=sheet_name, index=False)

        return df

    def liste_poteau_cap_ft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):
        """Récupérer les données liées à la base de données"""
        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        listePoteaux = []
        listePoteauxComplet = []
        fichiers = []
        listeEtat = []
        nature_travaux = []
        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        request_cap_ft = QgsFeatureRequest()
        clause_cap_ft = QgsFeatureRequest.OrderByClause(colonne_cap_ft, ascending=True)
        orderby_cap_ft = QgsFeatureRequest.OrderBy([clause_cap_ft])
        request_cap_ft.setOrderBy(orderby_cap_ft)

        for feat_cap_ft in etude_cap_ft.getFeatures(request_cap_ft):
            etudes = feat_cap_ft[colonne_cap_ft]
            for feat_pot in infra_pt_pot.getFeatures(request):
                if feat_cap_ft.geometry().contains(feat_pot.geometry()):
                    inf_num = feat_pot["inf_num"]

                    etat = np.NaN if feat_pot["etat"] == NULL else feat_pot["etat"]
                    nature = "Recalage" if etat == "A RECALER" else ("Remplacement" if etat == "A REMPLACER" else etat)

                    nature_travaux.append(nature)
                    listePoteauxComplet.append(inf_num)
                    position = inf_num.find('FT-')
                    inf_num = int(inf_num[position + 3:])

                    listePoteaux.append(str(inf_num))
                    fichiers.append(etudes)
                    listeEtat.append(etat)

        df = pd.DataFrame({'N° appui': listePoteaux, "Études": fichiers, "Nature des travaux": nature_travaux,
                           'inf_num (QGIS)': listePoteauxComplet, "Etat": listeEtat}, dtype="float64")  # 'inf_num': listePoteauxComplet, "etat": listeEtat,

        return df

    def ecrictureExcel(self, final, fichier):
        """Ecriture du résultat final dans un fichier Excel"""

        # headers = ['N° appui', 'Nature des travaux', 'Études', "inf_num (QGIS)", "Excel"]
        # headers = ['N° appui', 'Nature des travaux', 'Études', 'Excel', 'inf_num (QGIS)', 'Etat']
        # print(tabulate(final.values, headers, tablefmt='psql'))

        nb_erreurs = 0
        with pd.ExcelWriter(fichier, engine="openpyxl") as writer:
            sheet_name = "ANALYSE C6 BD"
            # Export DataFrame content
            final.to_excel(writer, sheet_name=sheet_name, index=False)
            # Set backgrund colors depending on cell values
            sheet = writer.sheets[sheet_name]

            alphabet = ["A", "B", "C", "D", "E", "F"]
            index = 1

            # Parcourir les lettres
            for alpha in alphabet:
                # Largeur des colonnes
                sheet.column_dimensions[alpha].width = 25
                index += 1

            for compte, [colA, colB, colC, colD, colE, colF] in enumerate(sheet[f'A2:F{len(final) + 1}']):
                value_num = final["inf_num (QGIS)"].iloc[compte]  # value is "True" or "False"
                value_excel = final["Excel"].iloc[compte]  # value is "True" or "False"

                # On met toute la ligne en couleur rouge
                if "ABSENT" in value_num or "ABSENT" in value_excel:
                    colA.fill = PatternFill("solid", start_color="feb24c")
                    colB.fill = PatternFill("solid", start_color="feb24c")
                    colD.fill = PatternFill("solid", start_color="feb24c")
                    colC.fill = PatternFill("solid", start_color="feb24c")
                    colE.fill = PatternFill("solid", start_color="feb24c")
                    colF.fill = PatternFill("solid", start_color="feb24c")
                    nb_erreurs += 1

        return nb_erreurs
    