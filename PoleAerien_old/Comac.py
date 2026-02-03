#!/usr/bin/python
# -*- coding: utf-8 -*-

from qgis.core import Qgis, QgsProject, QgsFeatureRequest, QgsExpression
import os
import xlrd
import openpyxl


class Comac:
    """ Second pour contenir quelques fonctions qui seront utilisées dans le fichier principale CheckOptyce.py """

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def verificationsDonneesComac(self, table_poteau, table_etude_comac, colonne_comac):

        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_comac = QgsProject.instance().mapLayersByName(table_etude_comac)[0]

        requete = QgsExpression("inf_type LIKE 'POT-BT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        request_comac = QgsFeatureRequest()
        clause_comac = QgsFeatureRequest.OrderByClause(colonne_comac, ascending=True)
        orderby_comac = QgsFeatureRequest.OrderBy([clause_comac])
        request_comac.setOrderBy(orderby_comac)

        ###################### POT-BT n'ont situé dans une étude comac #################
        pot_bt_hors_etude = []
        for feat_pot in infra_pt_pot.getFeatures(request):
            pot_dans_une_etude_comac = False
            inf_num = feat_pot["inf_num"]

            for feat_comac in etude_comac.getFeatures(request_comac):
                if feat_comac.geometry().contains(feat_pot.geometry()):
                    pot_dans_une_etude_comac = True
                    break

            if not pot_dans_une_etude_comac:
                pot_bt_hors_etude.append(inf_num)

        ###################### 2 ou plusieurs études comac ayant le même nom #################
        doublonsNomEtudesComac = []
        for feat_comac_1 in etude_comac.getFeatures(request_comac):
            nomEtude = feat_comac_1[colonne_comac]
            for feat_comac_2 in etude_comac.getFeatures(request_comac):
                if feat_comac_1.id() != feat_comac_2.id():
                    if nomEtude == feat_comac_2[colonne_comac]:
                        doublonsNomEtudesComac.append(nomEtude)
                        break

        return doublonsNomEtudesComac, pot_bt_hors_etude

    def liste_poteau_comac(self, table_poteau, table_etude_comac, colonne_comac):
        dicoEtudeComacPoteauQgis = {}

        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_comac = QgsProject.instance().mapLayersByName(table_etude_comac)[0]

        requete = QgsExpression("inf_type LIKE 'POT-BT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        request_comac = QgsFeatureRequest()
        clause_comac = QgsFeatureRequest.OrderByClause(colonne_comac, ascending=True)
        orderby_comac = QgsFeatureRequest.OrderBy([clause_comac])
        request_comac.setOrderBy(orderby_comac)

        for feat_comac in etude_comac.getFeatures(request_comac):
            listePoteau = []
            etudes = feat_comac[colonne_comac]
            for feat_pot in infra_pt_pot.getFeatures(request):
                if feat_comac.geometry().contains(feat_pot.geometry()):
                    listePoteau.append(feat_pot["inf_num"])

            if listePoteau:
                dicoEtudeComacPoteauQgis[etudes] = listePoteau

        return dicoEtudeComacPoteauQgis

    def LectureFichiersExcelsComac(self, repertoire):
        """Fonction pour parcourir les fichiers Excel pour renseigner la référence des appuis."""
        dicoPoteauBt_SousTraitant = {}
        fichiersComacExistants = []
        fichiersComacEnDoublons = []
        impossibiliteDelireFichier = {}

        for subdir, dirs, files in os.walk(repertoire):
            doss = subdir.split(os.sep)[-1]
            if "ETUDE FINALE" in doss.upper():
                doss = subdir.split(os.sep)[-2]
            print("dossier :", doss)
            for name in files:
                if name.endswith('.xlsx'):
                    if "EXPORTCOMAC" in name.upper() and str("~$") not in str(name):
                        filepath = subdir + os.sep + name
                        try:
                            document = xlrd.open_workbook(filepath)
                        except Exception as e:
                            # Si erreur lors de la lecture du fichier
                            impossibiliteDelireFichier[filepath] = str(e)
                            print(f"Une autre erreur : {e}")
                            continue

                        # idée : indiquez les fichiers pour les quels les numéros n'ont pas été trouvés dans C3A
                        feuille_1 = document.sheet_by_index(0)
                        rows = feuille_1.nrows
                        listePoteauBt = []

                        # Les valeurs retenues commencent à partir de la colonne A et ligne A4.
                        # On créé une liste qui contient toutes valeurs des appuis à remplacer
                        for r in range(3, rows):
                            numPotBt = feuille_1.cell_value(rowx=r, colx=0)
                            # Si le numéro n'existe pas, on le supprime
                            ## nompot = str(numPotBt).replace("BT ", "BT-")
                            nompot = numPotBt

                            if nompot:
                                listePoteauBt.append(nompot)

                        if listePoteauBt:
                            #dicoPoteauBt_SousTraitant[name]=listePoteauBt
                            dicoPoteauBt_SousTraitant[doss] = [name, listePoteauBt]

                            # On vérifie si le poteau n'existait pas déjà
                            #if name in fichiersComacExistants:
                                #fichiersComacEnDoublons.append(name)
                            if doss in fichiersComacExistants:
                                fichiersComacEnDoublons.append(doss)

                            #fichiersComacExistants.append(name)
                            fichiersComacExistants.append(doss)

        return fichiersComacEnDoublons, impossibiliteDelireFichier, dicoPoteauBt_SousTraitant

    def traitementResultatFinaux(self, dicoEtudeComacPoteauQgis, dicoPoteauBt_SousTraitant):

        """Fonction pour traiter les résultats finaux des deux dictionnaires """
        dicoPotBt_Excel_Introuvable = {}
        dicoPotBtExistants = {}

        comptabilite = 0

        #################### EXCEL SOUS-TRAITANT ###############################
        #for excel, listePoteau in dicoPoteauBt_SousTraitant.items():
        for doss, val in dicoPoteauBt_SousTraitant.items():
            excel, listePoteau = val[0], val[1]
            PotBtintrouvableSt = []
            for poteauSt in listePoteau:
                print(excel)
                print('poteauSt: ', poteauSt)
                introuvable = True

                #################### ETUDE QGIS ###############################
                for etude_comac, listePoteauxQgis in dicoEtudeComacPoteauQgis.items():
                    print(etude_comac)
                    inf_num_trouve = ""
                    nouveauListePoteauxQgis = []
                    for inf_num_bt in listePoteauxQgis:
                        nouveauListePoteauxQgis = listePoteauxQgis.copy()
                        #position = inf_num_bt.find('BT')
                        #inf_num = inf_num_bt[position:]
                        inf_num = inf_num_bt.split("/")[0]
                        print(inf_num)
                        #if str(inf_num) == str(poteauSt):  # str(poteauSt.replace("BT", 'BT-')
                        if str(inf_num) == str(poteauSt) and doss.upper() == etude_comac.upper():
                            introuvable = False
                            comptabilite += 1
                            
                            #################### VALEURS TROUVEES ########################################
                            dicoPotBtExistants[comptabilite] = [inf_num_bt, etude_comac, poteauSt, excel]
                            inf_num_trouve = inf_num_bt
                            break

                    ## Sortir du boucle si poteau trouve ##
                    if not introuvable:
                        nouveauListePoteauxQgis.remove(inf_num_trouve)

                        # S'il ne reste plus rien dans le dictionnaire, on supprime ce dictionnaire
                        if not nouveauListePoteauxQgis:
                            del dicoEtudeComacPoteauQgis[etude_comac]

                        # Sinon, on supprime juste un élément de la liste du dictionnaire
                        else:
                            dicoEtudeComacPoteauQgis[etude_comac]=nouveauListePoteauxQgis
                        break

                if introuvable:
                    PotBtintrouvableSt.append(poteauSt)

            if PotBtintrouvableSt:
                dicoPotBt_Excel_Introuvable[excel]=PotBtintrouvableSt

        return dicoPotBt_Excel_Introuvable, dicoEtudeComacPoteauQgis, dicoPotBtExistants   #

    def ecrireResultatsAnalyseExcels(self, resultatsFinaux, nom):
        """Fonction qui permet d'écrire dans le fichier Excel à partir d'un fichier modèle'"""

        dicoPotBt_Excel_Introuvable = resultatsFinaux[0]
        dicoEtudeComacPotQgisIntrouvable = resultatsFinaux[1]
        dicoPotBtExistants = resultatsFinaux[2]

        fichierXlsx = openpyxl.workbook.Workbook()

        ############################## FEUILLE 1 : VALIDATION_COMPLET ################################################

        my_ligne = 1  # Ligne à partir de laquelle on commencera à écrire dans le fichier Annexe
        # Parcourir fichier C3A, pour comparer ses contenus aux fichiers C6
        feuille = fichierXlsx.create_sheet("ANALYSE COMAC", 0)

        alignement = openpyxl.styles.Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False,
                                               shrink_to_fit=True, indent=0)

        # Pour mettre du GRAS à partir de la colonne 18 et 23 pour les câbles.
        feuille.cell(row=1, column=1, value="INF_NUM QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=2, value="ETUDE QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=3, value="INF_NUM EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=4, value="NOM FICHIER EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=5, value="REMARQUES").font = openpyxl.styles.Font(bold=True)

        feuille.cell(row=1, column=1).alignment = alignement
        feuille.cell(row=1, column=2).alignment = alignement
        feuille.cell(row=1, column=3).alignment = alignement
        feuille.cell(row=1, column=4).alignment = alignement
        feuille.cell(row=1, column=5).alignment = alignement

        feuille.column_dimensions["A"].width = 25
        feuille.column_dimensions["B"].width = 30
        feuille.column_dimensions["C"].width = 25
        feuille.column_dimensions["D"].width = 50
        feuille.column_dimensions["E"].width = 40

        red_color = openpyxl.styles.colors.Color(rgb='fc4e2a')
        orange_color = openpyxl.styles.colors.Color(rgb='fd8d3c')

        ####################### SOUS-TRAITS INTROUVABLE #######################################")
        for fichierExcels, listesDesAppuisSous_traitant in dicoPotBt_Excel_Introuvable.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for inf_num_excel in listesDesAppuisSous_traitant:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=3, value=inf_num_excel).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=4, value=fichierExcels).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans QGIS").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)

        ####################### QGIS INTROUVABLE #######################################
        for etudeQgis, listesDesAppuisQGis in dicoEtudeComacPotQgisIntrouvable.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for inf_num in listesDesAppuisQGis:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=1, value=inf_num).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=2, value=etudeQgis).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans les Excels").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)

        my_colonne = 1
        my_ligne += 1

        ####################### POTEAUX TROUVEES #######################################")
        for etiquetteDuCable, listesDesAppuisBT in dicoPotBtExistants.items():
            """Feuille Saisies Terrain pour écrire les données associés"""
            for valeur in listesDesAppuisBT:
                # Si tirage câble complètement validé
                feuille.cell(row=my_ligne, column=my_colonne, value=valeur)
                my_colonne += 1

            feuille.cell(row=my_ligne, column=5, value="correspondance trouvée")

            # On passe à la ligne suivante
            # Si la première liste est terminé, on passe à la ligne suivante en récommençant par au toujours au niveau de la colonne A
            my_ligne += 1
            my_colonne = 1

        std = fichierXlsx.get_sheet_by_name('Sheet')
        fichierXlsx.remove_sheet(std)

        fichierXlsx.save(filename=nom)   # Enregistrement du fichier
