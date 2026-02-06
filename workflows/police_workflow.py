# -*- coding: utf-8 -*-
"""
Workflow pour l'analyse Police C6 v2.0

NOUVELLE LOGIQUE (fddcpi2):
1. Récupérer SRO depuis couche QGIS
2. Charger câbles découpés via fddcpi2 (PostgreSQL)
3. Intersectionner câbles avec appuis
4. Comparer avec Annexe C6 (nombre + capacités)
5. Générer rapport anomalies

Plus de GraceTHD - tout vient de la fonction PostgreSQL fddcpi2.
"""

from qgis.PyQt.QtCore import QObject, pyqtSignal
from qgis.core import Qgis, QgsMessageLog, QgsProject, QgsVectorLayer, NULL
from ..PoliceC6 import PoliceC6, PoliceC6Cancelled
from ..db_connection import extract_sro_from_layer
from ..cable_analyzer import extraire_appuis_from_layer
from ..qgis_utils import get_layer_safe
from ..async_tasks import PoliceC6Task, run_async_task
import os
import glob
from datetime import datetime

class PoliceWorkflow(QObject):
    """
    Contrôleur pour le flux d'analyse Police C6.
    """
    
    # Signaux
    progress_changed = pyqtSignal(int)
    message_received = pyqtSignal(str, str)  # message, couleur
    analysis_finished = pyqtSignal(dict)  # resultats
    error_occurred = pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.police_logic = PoliceC6()
        self.current_task = None
        self._cancelled = False

    def reset_logic(self):
        """Réinitialise la logique métier"""
        self.police_logic._reset_state()
    
    def cancel(self):
        """Annule le traitement en cours"""
        self._cancelled = True
        self.police_logic.request_cancel()
        QgsMessageLog.logMessage("Annulation demandée...", "POLICE_C6", Qgis.Info)

    def _emit_report_to_ui(self, report):
        """Diffuse le rapport détaillé vers l'UI avec code couleur."""
        if not report:
            return
        color_map = {
            "section": "grey",
            "info": "black",
            "ok": "green",
            "warn": "orange",
            "error": "red",
        }
        for item in report:
            color = color_map.get(item.get("level"), "black")
            self.message_received.emit(item.get("message", ""), color)

    def check_layers_exist(self):
        """Vérifie la présence des couches nécessaires.
        
        Returns:
            list: Liste des couches manquantes (vide si toutes présentes)
        """
        missing = []
        # Seule couche obligatoire pour la nouvelle logique fddcpi2
        if not QgsProject.instance().mapLayersByName('infra_pt_pot'):
            missing.append('infra_pt_pot')
        return missing

    def apply_style(self, style_name):
        """Applique un style QGIS via la logique métier"""
        self.police_logic.appliquerstyle(style_name)
    
    def extract_sro_from_project(self) -> str:
        """Extrait le SRO depuis les couches du projet QGIS."""
        return self.police_logic.extraire_sro()

    def detect_etude_field(self, layer):
        """
        Auto-détecte le champ étude dans une couche CAP FT.
        
        Args:
            layer: QgsVectorLayer
            
        Returns:
            str: Nom du champ détecté ou None
        """
        if not layer or not layer.isValid():
            return None
        
        candidates = ['nom_etudes', 'etudes', 'nom_etude', 'nom', 'decoupage', 'zone', 'ref_fci']
        for field in layer.fields():
            if field.name().lower() in [c.lower() for c in candidates]:
                return field.name()
        return None

    def find_c6_file(self, repertoire_c6, nom_etude):
        """
        Trouve le fichier C6 correspondant à une étude.
        Cherche récursivement dans les sous-dossiers (CMD 1, CMD 2, etc.)
        
        Args:
            repertoire_c6: Répertoire racine des C6
            nom_etude: Nom de l'étude
            
        Returns:
            str: Chemin du fichier C6 ou None
        """
        if not repertoire_c6 or not nom_etude:
            return None
        
        # Patterns de recherche (priorité décroissante)
        patterns = [
            # Pattern principal: **/nom_etude/nom_etude.xlsx (récursif)
            os.path.join(repertoire_c6, "**", nom_etude, f"{nom_etude}.xlsx"),
            # Sous-dossier direct avec nom étude
            os.path.join(repertoire_c6, nom_etude, f"{nom_etude}.xlsx"),
            os.path.join(repertoire_c6, nom_etude, "*Annexe*C6*.xlsx"),
            os.path.join(repertoire_c6, nom_etude, "*C6*.xlsx"),
            # Récursif: chercher dans tous les sous-dossiers
            os.path.join(repertoire_c6, "**", nom_etude, "*Annexe*C6*.xlsx"),
            os.path.join(repertoire_c6, "**", nom_etude, "*C6*.xlsx"),
            os.path.join(repertoire_c6, "**", f"{nom_etude}.xlsx"),
            # Fichier direct avec nom étude
            os.path.join(repertoire_c6, f"*{nom_etude}*.xlsx"),
        ]
        
        for pattern in patterns:
            # recursive=True pour supporter **
            matches = glob.glob(pattern, recursive=True)
            # Filtrer les fichiers non-C6 (FicheAppui, C7, GESPOT)
            for match in matches:
                fname = os.path.basename(match).lower()
                if 'ficheappui' in fname or 'c7' in fname or 'gespot' in fname:
                    continue
                return match
        
        return None

    def get_etudes_from_layer(self, table_etude, colonne_etude):
        """
        Récupère la liste des études depuis la couche etude_cap_ft.
        
        Args:
            table_etude: Nom de la couche
            colonne_etude: Nom de la colonne contenant les noms d'études
            
        Returns:
            list: Liste des noms d'études uniques
        """
        try:
            layer = get_layer_safe(table_etude, "Police_C6")
        except ValueError as e:
            QgsMessageLog.logMessage(f"get_etudes_from_layer: {e}", "PoleAerien", Qgis.Warning)
            return []
        
        etudes = set()
        idx = layer.fields().indexFromName(colonne_etude)
        if idx < 0:
            return []
        
        for feat in layer.getFeatures():
            val = feat[colonne_etude]
            if val and val != NULL:
                etudes.add(str(val).strip())
        
        return sorted(list(etudes))

    def run_analysis_auto_browse(self, params):
        """
        Lance l'analyse Police C6 en mode auto-browse ASYNCHRONE.
        
        ARCHITECTURE ASYNC:
        1. Main thread: Extract QGIS data (appuis, SRO, liste fichiers C6)
        2. Worker thread: PostgreSQL + comparaison + Excel
        
        Args:
            params (dict): Paramètres d'analyse
        """
        self._cancelled = False
        repertoire_c6 = params.get('repertoire_c6', '')
        table_etude = params.get('table_etude', '')
        colonne_etude = params.get('colonne_etude', '')
        export_path = params.get('export_path', repertoire_c6)
        
        self.message_received.emit("Mode auto-browse: préparation...", "blue")
        self.progress_changed.emit(2)
        
        # === MAIN THREAD: Extraction données QGIS ===
        
        # 1. Récupérer liste des études
        etudes = []
        if table_etude and colonne_etude:
            etudes = self.get_etudes_from_layer(table_etude, colonne_etude)
        
        if not etudes:
            # Fallback: scanner les fichiers C6 directement
            self.message_received.emit("Scan des fichiers C6...", "grey")
            c6_files = glob.glob(os.path.join(repertoire_c6, "**", "*C6*.xlsx"), recursive=True)
            c6_files = [f for f in c6_files if 'ficheappui' not in f.lower() and 'c7' not in f.lower()]
            etudes = [os.path.splitext(os.path.basename(f))[0] for f in c6_files]
        
        if not etudes:
            self.error_occurred.emit("Aucune étude trouvée")
            return
        
        # 2. Construire liste (etude_name, c6_file)
        c6_files_list = []
        etudes_sans_c6 = []
        
        for etude in etudes:
            c6_file = self.find_c6_file(repertoire_c6, etude)
            if c6_file:
                c6_files_list.append((etude, c6_file))
            else:
                etudes_sans_c6.append(etude)
        
        if not c6_files_list:
            self.error_occurred.emit("Aucun fichier C6 trouvé")
            return
        
        self.message_received.emit(f"Mode auto-browse: {len(c6_files_list)} études trouvées", "blue")
        
        # 3. Extraire appuis et SRO depuis couche infra_pt_pot (MAIN THREAD)
        layer_appuis = QgsProject.instance().mapLayersByName('infra_pt_pot')
        
        sro = None
        if layer_appuis:
            sro = extract_sro_from_layer(layer_appuis[0])
        if not sro:
            first_file = c6_files_list[0][1]
            sro = self._extract_sro_from_filename(first_file)
        
        if not sro:
            self.error_occurred.emit("SRO non trouvé dans les données QGIS")
            return
        
        self.message_received.emit(f"SRO: {sro}", "grey")
        appuis_data = []
        if layer_appuis:
            raw_appuis = extraire_appuis_from_layer(layer_appuis[0])
            # Sérialiser les géométries en WKB (QgsGeometry non picklable)
            for appui in raw_appuis:
                geom = appui.get('geom')
                appuis_data.append({
                    'num_appui': appui.get('num_appui', ''),
                    'feature_id': appui.get('feature_id'),
                    'geom_wkb': geom.asWkb().data() if geom and not geom.isNull() else None
                })
            self.message_received.emit(f"Appuis QGIS: {len(appuis_data)}", "grey")
        
        # === WORKER THREAD: Lancer tâche async ===
        
        task_params = {
            'c6_files': c6_files_list,
            'sro': sro,
            'export_path': export_path
        }
        
        qgis_data = {
            'appuis': appuis_data
        }
        
        self.current_task = PoliceC6Task(task_params, qgis_data)
        
        # Connecter signaux
        self.current_task.signals.progress.connect(self.progress_changed)
        self.current_task.signals.message.connect(self.message_received)
        self.current_task.signals.finished.connect(self._on_task_finished)
        self.current_task.signals.error.connect(self.error_occurred)
        
        run_async_task(self.current_task)
    
    def _extract_sro_from_filename(self, filepath):
        """Extrait le SRO depuis le nom du fichier C6."""
        basename = os.path.basename(filepath)
        # Pattern: 63041-B1I-PMZ-00003 ou 63041/B1I/PMZ/00003
        import re
        match = re.search(r'(\d{5})[/-]([A-Z]\d[A-Z])[/-](PMZ)[/-](\d{5})', basename)
        if match:
            return f"{match.group(1)}/{match.group(2)}/{match.group(3)}/{match.group(4)}"
        return None
    
    def _on_task_finished(self, result):
        """Callback quand la tâche async est terminée (main thread)."""
        self.current_task = None
        
        if result.get('error'):
            self.error_occurred.emit(result['error'])
            return
        
        # Charger les câbles fddcpi2 comme couche temporaire QGIS
        cables = result.get('cables', [])
        sro = result.get('sro', '')
        if cables:
            try:
                self._load_cables_layer(cables, sro)
            except Exception as e:
                self.message_received.emit(f"[!] Erreur chargement couche câbles: {e}", "orange")
        
        # Émettre résultat
        self.analysis_finished.emit(result)
        
        # Message récapitulatif
        stats = result.get('stats', [])
        self.message_received.emit(f"✓ Analyse terminée: {len(stats)} études traitées", "green")
    
    def _load_cables_layer(self, cables, sro):
        """Charge les câbles fddcpi2 comme couche temporaire dans QGIS."""
        from qgis.core import (
            QgsVectorLayer, QgsFeature, QgsGeometry, QgsField, QgsProject
        )
        from qgis.PyQt.QtCore import QVariant
        
        # Créer couche mémoire LineString en Lambert 93
        layer = QgsVectorLayer(
            "LineString?crs=EPSG:2154",
            f"fddcpi2_{sro.replace('/', '_')}",
            "memory"
        )
        
        provider = layer.dataProvider()
        
        # Ajouter les champs
        provider.addAttributes([
            QgsField("gid_dc2", QVariant.LongLong),
            QgsField("gid_dc", QVariant.LongLong),
            QgsField("cab_capa", QVariant.Int),
            QgsField("cab_type", QVariant.String),
            QgsField("cb_etiquet", QVariant.String),
            QgsField("posemode", QVariant.Int),
            QgsField("length", QVariant.Double),
        ])
        layer.updateFields()
        
        # Ajouter les features
        features = []
        for cable in cables:
            wkt = cable.get('geom_wkt')
            if not wkt:
                continue
            
            geom = QgsGeometry.fromWkt(wkt)
            if geom.isNull() or geom.isEmpty():
                continue
            
            feat = QgsFeature(layer.fields())
            feat.setGeometry(geom)
            feat.setAttribute("gid_dc2", cable.get('gid_dc2'))
            feat.setAttribute("gid_dc", cable.get('gid_dc'))
            feat.setAttribute("cab_capa", cable.get('cab_capa'))
            feat.setAttribute("cab_type", cable.get('cab_type', ''))
            feat.setAttribute("cb_etiquet", cable.get('cb_etiquet', ''))
            feat.setAttribute("posemode", cable.get('posemode'))
            feat.setAttribute("length", cable.get('length'))
            features.append(feat)
        
        provider.addFeatures(features)
        layer.updateExtents()
        
        # Ajouter au projet QGIS
        QgsProject.instance().addMapLayer(layer)
        
        # Appliquer le style cables_dc.qml
        style_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'styles', 'cables_dc.qml'
        )
        if os.path.exists(style_path):
            layer.loadNamedStyle(style_path)
            layer.triggerRepaint()
        
        self.message_received.emit(
            f"✓ Couche '{layer.name()}' chargée: {len(features)} câbles découpés",
            "green"
        )
    
    def run_analysis_auto_browse_sync(self, params):
        """
        ANCIENNE VERSION SYNCHRONE - Conservée pour compatibilité.
        Utiliser run_analysis_auto_browse pour la version async.
        """
        self._cancelled = False
        repertoire_c6 = params.get('repertoire_c6', '')
        table_etude = params['table_etude']
        colonne_etude = params['colonne_etude']
        
        etudes = self.get_etudes_from_layer(table_etude, colonne_etude)
        if not etudes:
            self.error_occurred.emit("Aucune étude trouvée dans la couche")
            return
        
        self.message_received.emit(f"Parcours de {len(etudes)} études...", "blue")
        self.progress_changed.emit(5)
        
        etudes_sans_c6 = []
        etudes_traitees = 0
        stats_globales = []
        
        from qgis.PyQt.QtWidgets import QApplication
        
        for i, etude in enumerate(etudes):
            # Vérifier annulation
            QApplication.processEvents()
            if self._cancelled:
                self.message_received.emit("Traitement annulé par l'utilisateur", "orange")
                break
            
            progress = 5 + int((i / len(etudes)) * 90)
            self.progress_changed.emit(progress)
            
            # Trouver le fichier C6
            c6_file = self.find_c6_file(repertoire_c6, etude)
            
            if not c6_file:
                etudes_sans_c6.append(etude)
                self.message_received.emit(f"[!] {etude}: Fichier C6 introuvable", "orange")
                continue
            
            self.message_received.emit(f"[>] {etude}: {os.path.basename(c6_file)}", "blue")
            
            # Préparer les params pour cette étude
            etude_params = params.copy()
            etude_params['fname'] = c6_file
            etude_params['filterValeur'] = etude
            
            # Lancer l'analyse pour cette étude
            try:
                self.police_logic._reset_state()
                self._run_single_analysis(etude_params)
                etudes_traitees += 1
                # Collecter stats pour tableau récap + données détaillées pour export
                # SENS CORRECT: C6 → QGIS (ce qui est dans C6 mais pas dans QGIS)
                cables_c6_absents = getattr(self.police_logic, 'cables_c6_absents_qgis', [])
                stats_globales.append({
                    'etude': etude,
                    'appuis_ok': self.police_logic.nb_appui_corresp,
                    'appuis_c6_manq': self.police_logic.nb_appui_absentPot,  # Appuis C6 absents de QGIS
                    'appuis_qgis_manq': self.police_logic.nb_appui_absent,   # Pour info seulement
                    'bpe_ok': self.police_logic.nb_pbo_corresp,
                    'bpe_anom': len(self.police_logic.ebp_non_appui) + len(self.police_logic.ebp_appui_inconnu),
                    'cables_ok': getattr(self.police_logic, 'cable_corresp', 0),
                    'cables_c6_manq': len(cables_c6_absents),  # Câbles C6 absents de QGIS (sens correct)
                    'rac_exclus': getattr(self.police_logic, 'cables_rac_exclus', 0),
                    # Données détaillées pour export Excel - SENS C6 → QGIS
                    'detail_appuis_c6_manq': list(self.police_logic.absence),  # Appuis C6 absents de QGIS
                    'detail_appuis_qgis_manq': list(getattr(self.police_logic, 'idPotAbsent', [])),
                    'detail_bpe_non_appui': list(self.police_logic.ebp_non_appui),
                    'detail_bpe_appui_inconnu': list(self.police_logic.ebp_appui_inconnu),
                    'detail_cables_c6_absents': list(cables_c6_absents)  # Câbles C6 absents de QGIS
                })
            except PoliceC6Cancelled:
                self._cancelled = True
                self.message_received.emit("Traitement annulé par l'utilisateur", "orange")
                break
            except Exception as e:
                self.message_received.emit(f"[X] {etude}: Erreur - {e}", "red")
                QgsMessageLog.logMessage(f"Erreur analyse {etude}: {e}", "PoleAerien", Qgis.Warning)
        
        # 3. Rapport final avec tableau récapitulatif compact
        self.progress_changed.emit(100)
        
        # Construire tableau récapitulatif bien formaté
        # SENS: C6 → QGIS (anomalies = ce qui est dans C6 mais pas dans QGIS)
        totaux = {'appuis_ok': 0, 'c6_manq': 0, 'qg_manq': 0, 'bpe_ok': 0, 'bpe_anom': 0, 'cables_ok': 0, 'cables_c6_manq': 0, 'rac_exclus': 0}
        lignes = []
        
        for s in stats_globales:
            # Extraire numéro court de l'étude (ex: "FTTH-NGE-ETUDE-B1L-63041-14" -> "14")
            etude_short = s['etude'].split('-')[-1] if '-' in s['etude'] else s['etude'][:8]
            lignes.append({
                'etude': etude_short,
                'appuis_ok': s['appuis_ok'],
                'c6_manq': s['appuis_c6_manq'],  # Appuis C6 absents de QGIS
                'qg_manq': s['appuis_qgis_manq'],
                'bpe_ok': s['bpe_ok'],
                'bpe_anom': s['bpe_anom'],
                'cables_ok': s['cables_ok'],
                'cables_c6_manq': s.get('cables_c6_manq', 0)  # Câbles C6 absents de QGIS
            })
            totaux['appuis_ok'] += s['appuis_ok']
            totaux['c6_manq'] += s['appuis_c6_manq']
            totaux['qg_manq'] += s['appuis_qgis_manq']
            totaux['bpe_ok'] += s['bpe_ok']
            totaux['bpe_anom'] += s['bpe_anom']
            totaux['cables_ok'] += s['cables_ok']
            totaux['cables_c6_manq'] += s.get('cables_c6_manq', 0)
            totaux['rac_exclus'] += s.get('rac_exclus', 0)
        
        # Afficher récapitulatif en tableau HTML (comme MAJ BD)
        # Termes clairs:
        # - Appuis C6∩QGIS = présents dans les deux
        # - Absents QGIS = dans C6 mais pas dans QGIS  
        # - Absents C6 = dans QGIS mais pas dans C6
        # - Boîtes OK = correspondance PBO trouvée
        # - Boîtes Ano = EBP sans appui ou appui inconnu
        # - Câbles OK = triplet appui-câble-appui trouvé
        # - Câbles Ano = câbles QGIS non trouvés dans C6
        html = '<p><b style="color:#6366f1">RECAP POLICE C6</b></p>'
        html += '<table cellspacing="0" cellpadding="3" style="font-size:9pt;border-collapse:collapse">'
        
        # En-têtes avec termes clairs
        html += '<tr style="background:#6366f1;color:#fff">'
        html += '<th style="padding:4px 8px">Etude</th>'
        html += '<th style="padding:4px 8px" title="Appuis présents dans C6 ET QGIS">Appuis OK</th>'
        html += '<th style="padding:4px 8px" title="Appuis dans C6 mais absents de QGIS">Abs. QGIS</th>'
        html += '<th style="padding:4px 8px" title="Appuis dans QGIS mais absents de C6">Abs. C6</th>'
        html += '<th style="padding:4px 8px" title="Boîtes (PBO) avec correspondance">Boîtes OK</th>'
        html += '<th style="padding:4px 8px" title="EBP sans appui ou appui inconnu">Boîtes Ano</th>'
        html += '<th style="padding:4px 8px" title="Triplets Appui-Câble-Appui trouvés">Câbles OK</th>'
        html += '<th style="padding:4px 8px" title="Câbles C6 absents de QGIS/GraceTHD">Câbles C6 Manq</th>'
        html += '</tr>'
        
        # Lignes de données
        for row in lignes:
            has_anomaly = (row['c6_manq'] > 0 or row['bpe_anom'] > 0 or row['cables_c6_manq'] > 0)
            bg = '#fee2e2' if has_anomaly else '#dcfce7'  # Rouge clair ou vert clair
            html += f'<tr style="background:{bg}">'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0"><b>{row["etude"]}</b></td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["appuis_ok"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["c6_manq"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["qg_manq"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["bpe_ok"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["bpe_anom"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["cables_ok"]}</td>'
            html += f'<td style="padding:3px 6px;border:1px solid #e2e8f0;text-align:center">{row["cables_c6_manq"]}</td>'
            html += '</tr>'
        
        # Ligne totaux
        html += '<tr style="background:#6366f1;color:#fff;font-weight:bold">'
        html += '<td style="padding:3px 6px">TOTAL</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["appuis_ok"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["c6_manq"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["qg_manq"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["bpe_ok"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["bpe_anom"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["cables_ok"]}</td>'
        html += f'<td style="padding:3px 6px;text-align:center">{totaux["cables_c6_manq"]}</td>'
        html += '</tr>'
        html += '</table>'
        
        # Émettre le tableau HTML
        self.message_received.emit(html, "html")
        
        # Info sur les câbles RAC exclus
        if totaux['rac_exclus'] > 0:
            self.message_received.emit(f"<p style='color:#64748b;font-size:9pt'>Note: {totaux['rac_exclus']} câbles RAC exclus (non présents dans Annexe C6)</p>", "html")
        
        self.message_received.emit(f"Etudes traitees: {etudes_traitees}/{len(etudes)}", "green" if etudes_traitees > 0 else "orange")
        
        if etudes_sans_c6:
            self.message_received.emit(f"Etudes sans C6: {', '.join(etudes_sans_c6)}", "orange")
        
        # Export Excel si chemin fourni
        export_path = params.get('export_path')
        if export_path and stats_globales:
            excel_file = self._export_to_excel(export_path, stats_globales, totaux)
            if excel_file:
                self.message_received.emit(f"<p style='color:#22c55e;font-weight:bold'>✓ Export Excel: {os.path.basename(excel_file)}</p>", "html")
        
        # Émettre le résultat final
        result = {
            'success': True,
            'mode': 'auto_browse',
            'etudes_traitees': etudes_traitees,
            'etudes_sans_c6': etudes_sans_c6
        }
        self.analysis_finished.emit(result)

    def _run_single_analysis(self, params):
        """
        Exécute l'analyse pour une seule étude.
        
        NOUVELLE LOGIQUE:
        1. Lire Annexe C6 (Excel)
        2. Récupérer SRO depuis couche QGIS
        3. Charger câbles découpés via fddcpi2
        4. Compter câbles par appui + capacités
        5. Comparer avec C6
        """
        fname = params['fname']  # Fichier C6 Excel
        table = params['table_etude']
        colonne = params['colonne_etude']
        filterValeur = params['filterValeur']  # Nom de l'étude
        
        # 1. Lire l'Annexe C6
        donnees_c6, liste_brute = self.police_logic.lire_annexe_c6(fname)
        
        if not donnees_c6:
            self.message_received.emit(f"  [!] Aucune donnée dans C6: {fname}", "orange")
            return
        
        self.police_logic.donnees_c6 = donnees_c6
        self.message_received.emit(f"  C6: {len(donnees_c6)} appuis lus", "blue")
        
        # 2. Récupérer la couche appuis
        layer_appuis = None
        layers = QgsProject.instance().mapLayersByName('infra_pt_pot')
        if layers:
            layer_appuis = layers[0]
        
        if not layer_appuis:
            self.message_received.emit("  [X] Couche infra_pt_pot introuvable", "red")
            return
        
        # 3. Extraire le SRO
        sro = self.police_logic.extraire_sro(layer_appuis)
        
        if not sro:
            self.message_received.emit("  [!] SRO non trouvé, tentative depuis nom étude", "orange")
            # Essayer de construire le SRO depuis le nom de l'étude
            # Format: FTTH-NGE-ETUDE-B1L-63041-14 -> 63041/B1L/PMZ/xxxxx
            sro = self._extract_sro_from_etude_name(filterValeur)
        
        if not sro:
            self.message_received.emit("  [X] Impossible de déterminer le SRO", "red")
            return
        
        self.message_received.emit(f"  SRO: {sro}", "blue")
        
        # 4. Analyser charge câbles via fddcpi2
        result = self.police_logic.analyser_charge_cables(
            sro=sro,
            layer_appuis=layer_appuis,
            donnees_c6=donnees_c6,
            only_aerien=True
        )
        
        if result.erreur:
            self.message_received.emit(f"  [X] Erreur: {result.erreur}", "red")
            return
        
        # 5. Afficher les résultats
        if result.appuis_ok > 0:
            self.message_received.emit(f"  [OK] {result.appuis_ok} appui(s) OK", "green")
        
        if result.appuis_anomalie > 0:
            self.message_received.emit(f"  [!] {result.appuis_anomalie} appui(s) avec anomalie", "orange")
            
            # Afficher quelques anomalies
            for anom in result.anomalies[:5]:
                self.message_received.emit(
                    f"      - {anom.num_appui}: BDD={anom.nb_cables_bdd} câbles ({anom.capacite_totale_bdd}FO) vs C6={anom.nb_cables_c6} ({anom.capacite_totale_c6}FO)",
                    "orange"
                )
            
            if len(result.anomalies) > 5:
                self.message_received.emit(f"      ... et {len(result.anomalies) - 5} autres", "orange")
        
        # Stocker pour export Excel
        self.police_logic.nb_appui_corresp = result.appuis_ok
        self.police_logic.nb_appui_absentPot = result.appuis_anomalie
    
    def _extract_sro_from_etude_name(self, etude_name: str) -> str:
        """Tente d'extraire le SRO depuis le nom de l'étude."""
        import re
        # Pattern: FTTH-NGE-ETUDE-B1I-63041-14 ou similaire
        match = re.search(r'(\d{5})', etude_name)
        if match:
            code_insee = match.group(1)
            # Chercher le type (B1I, B1L, etc.)
            type_match = re.search(r'(B\d[A-Z])', etude_name, re.IGNORECASE)
            if type_match:
                type_code = type_match.group(1).upper()
                # Construire le SRO (format approximatif)
                return f"{code_insee}/{type_code}/PMZ/00001"
        return None

    def run_analysis(self, params):
        """
        Exécute l'analyse complète via fddcpi2.
        
        NOUVELLE LOGIQUE SIMPLIFIÉE:
        1. Lire Annexe C6
        2. Charger câbles découpés via fddcpi2
        3. Compter câbles par appui
        4. Comparer avec C6
        
        Args:
            params (dict): Paramètres d'analyse
                - fname: Chemin fichier Excel C6
                - filterValeur: Nom de l'étude
                - sro: Code SRO (optionnel, sinon extrait de couche)
        """
        try:
            self.progress_changed.emit(5)
            self.message_received.emit("============ POLICE C6 v2.0 (fddcpi2) ============", "grey")
            
            # Exécuter l'analyse
            self._run_single_analysis(params)
            
            self.progress_changed.emit(100)
            self.message_received.emit("============ ANALYSE TERMINÉE ============", "grey")
            
            # Résultat
            result = {
                'success': True,
                'nb_appuis_ok': self.police_logic.nb_appui_corresp,
                'nb_anomalies': self.police_logic.nb_appui_absentPot,
                'anomalies': self.police_logic.anomalies_cables
            }
            self.analysis_finished.emit(result)

        except PoliceC6Cancelled:
            self.message_received.emit("Traitement annulé par l'utilisateur", "orange")
            self.analysis_finished.emit({'success': False, 'cancelled': True})
        except Exception as e:
            QgsMessageLog.logMessage(f"Erreur PoliceWorkflow: {e}", "PoleAerien", Qgis.Critical)
            self.error_occurred.emit(str(e))

    def _export_to_excel(self, export_path, stats_globales, totaux):
        """
        Exporte les résultats détaillés vers un fichier Excel avec plusieurs feuilles.
        
        Args:
            export_path: Répertoire d'export
            stats_globales: Liste des statistiques par étude
            totaux: Dictionnaire des totaux
            
        Returns:
            Chemin du fichier créé ou None si erreur
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.message_received.emit("Module openpyxl non disponible - export Excel impossible", "orange")
            return None
        
        try:
            # Nom fichier avec timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"PoliceC6_Export_{timestamp}.xlsx"
            filepath = os.path.join(export_path, filename)
            
            wb = openpyxl.Workbook()
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            ok_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            warn_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # ===== FEUILLE 1: RECAPITULATIF =====
            ws_recap = wb.active
            ws_recap.title = "Récapitulatif"
            
            recap_headers = ["Etude", "Appuis OK", "Abs. QGIS", "Abs. C6", "Boîtes OK", "Boîtes Ano", "Câbles OK", "Câbles Ano", "RAC Exclus"]
            for col, header in enumerate(recap_headers, 1):
                cell = ws_recap.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            for row_idx, s in enumerate(stats_globales, 2):
                ws_recap.cell(row=row_idx, column=1, value=s['etude']).border = thin_border
                ws_recap.cell(row=row_idx, column=2, value=s['appuis_ok']).border = thin_border
                ws_recap.cell(row=row_idx, column=3, value=s['appuis_c6_manq']).border = thin_border
                ws_recap.cell(row=row_idx, column=4, value=s['appuis_qgis_manq']).border = thin_border
                ws_recap.cell(row=row_idx, column=5, value=s['bpe_ok']).border = thin_border
                ws_recap.cell(row=row_idx, column=6, value=s['bpe_anom']).border = thin_border
                ws_recap.cell(row=row_idx, column=7, value=s['cables_ok']).border = thin_border
                ws_recap.cell(row=row_idx, column=8, value=s['cables_c6_manq']).border = thin_border
                ws_recap.cell(row=row_idx, column=9, value=s.get('rac_exclus', 0)).border = thin_border
                
                # Coloration ligne
                has_anomaly = s['appuis_c6_manq'] > 0 or s['bpe_anom'] > 0 or s['cables_c6_manq'] > 0
                fill = warn_fill if has_anomaly else ok_fill
                for col in range(1, 10):
                    ws_recap.cell(row=row_idx, column=col).fill = fill
            
            # Ligne TOTAL
            total_row = len(stats_globales) + 2
            ws_recap.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws_recap.cell(row=total_row, column=2, value=totaux['appuis_ok'])
            ws_recap.cell(row=total_row, column=3, value=totaux['c6_manq'])
            ws_recap.cell(row=total_row, column=4, value=totaux['qg_manq'])
            ws_recap.cell(row=total_row, column=5, value=totaux['bpe_ok'])
            ws_recap.cell(row=total_row, column=6, value=totaux['bpe_anom'])
            ws_recap.cell(row=total_row, column=7, value=totaux['cables_ok'])
            ws_recap.cell(row=total_row, column=8, value=totaux['cables_c6_manq'])
            ws_recap.cell(row=total_row, column=9, value=totaux['rac_exclus'])
            for col in range(1, 10):
                cell = ws_recap.cell(row=total_row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            
            # Ajuster largeurs colonnes
            for col in range(1, 10):
                ws_recap.column_dimensions[get_column_letter(col)].width = 15
            ws_recap.column_dimensions['A'].width = 35
            
            # ===== FEUILLE 2: APPUIS ABSENTS QGIS =====
            ws_appuis_qgis = wb.create_sheet("Appuis Abs. QGIS")
            ws_appuis_qgis.append(["Etude", "Numéro Appui C6"])
            for col in range(1, 3):
                ws_appuis_qgis.cell(row=1, column=col).font = header_font
                ws_appuis_qgis.cell(row=1, column=col).fill = header_fill
            
            for s in stats_globales:
                for appui in s.get('detail_appuis_c6_manq', []):
                    ws_appuis_qgis.append([s['etude'], str(appui)])
            
            # ===== FEUILLE 3: APPUIS ABSENTS C6 =====
            ws_appuis_c6 = wb.create_sheet("Appuis Abs. C6")
            ws_appuis_c6.append(["Etude", "ID Appui QGIS"])
            for col in range(1, 3):
                ws_appuis_c6.cell(row=1, column=col).font = header_font
                ws_appuis_c6.cell(row=1, column=col).fill = header_fill
            
            for s in stats_globales:
                for appui in s.get('detail_appuis_qgis_manq', []):
                    ws_appuis_c6.append([s['etude'], str(appui)])
            
            # ===== FEUILLE 4: BPE ANOMALIES =====
            ws_bpe = wb.create_sheet("BPE Anomalies")
            ws_bpe.append(["Etude", "Code Etude", "ID Poteau QGIS", "Num Appui", "Commune Appui", "Type BPE", "Problème", "Explication"])
            for col in range(1, 9):
                ws_bpe.cell(row=1, column=col).font = header_font
                ws_bpe.cell(row=1, column=col).fill = header_fill
            
            for s in stats_globales:
                # Extraire code commune de l'étude (ex: FTTH-NGE-ETUDE-B1L-63041-1 -> 63041)
                etude_name = s['etude']
                etude_commune = ''
                parts = etude_name.split('-')
                for p in parts:
                    if p.isdigit() and len(p) == 5:
                        etude_commune = p
                        break
                
                # BPE sans appui = BPE isolé, pas sur un poteau
                for bpe_id in s.get('detail_bpe_non_appui', []):
                    ws_bpe.append([
                        etude_name,
                        etude_commune,
                        '',  # Pas d'ID poteau
                        '',  # Pas de num appui
                        '',  # Pas de commune
                        f'ID BPE: {bpe_id}',
                        "BPE ISOLÉ",
                        f"Le BPE (ID={bpe_id}) n'est positionné sur AUCUN appui dans QGIS (distance > 0.5m de tout poteau). "
                        f"ACTION: Corriger le positionnement du BPE dans QGIS pour qu'il soit sur un appui, ou vérifier si l'appui manque dans la couche infra_pt_pot."
                    ])
                
                # BPE sur appui non trouvé dans C6
                for bpe in s.get('detail_bpe_appui_inconnu', []):
                    # bpe = [id_poteau, inf_num, type_bpe]
                    id_pot = bpe[0] if len(bpe) > 0 else ''
                    inf_num = str(bpe[1]) if len(bpe) > 1 else ''
                    type_bpe = bpe[2] if len(bpe) > 2 else ''
                    
                    # Parser inf_num pour extraire commune (format: 1016613/63094)
                    appui_commune = ''
                    appui_num = inf_num
                    if '/' in inf_num:
                        parts_num = inf_num.split('/')
                        appui_num = parts_num[0]
                        appui_commune = parts_num[1] if len(parts_num) > 1 else ''
                    
                    # Déterminer le problème et l'explication avec action recommandée
                    if appui_commune and etude_commune and appui_commune != etude_commune:
                        probleme = "COMMUNE DIFFÉRENTE"
                        explication = (
                            f"Le BPE ({type_bpe}) est posé sur l'appui {appui_num} qui appartient à la commune {appui_commune}. "
                            f"Or, cette étude concerne la commune {etude_commune}. "
                            f"ACTION: Ce BPE est en frontière entre 2 études. Vérifier dans l'Annexe C6 de la commune {appui_commune} si cet appui y est déclaré."
                        )
                    else:
                        probleme = "APPUI ABSENT DU C6"
                        explication = (
                            f"Le BPE ({type_bpe}) est posé sur l'appui {appui_num} dans QGIS. "
                            f"Cet appui n'apparaît PAS dans l'Annexe C6 de l'étude {etude_commune}. "
                            f"ACTION: Soit ajouter cet appui dans l'Annexe C6, soit corriger le positionnement du BPE dans QGIS."
                        )
                    
                    ws_bpe.append([
                        etude_name,
                        etude_commune,
                        str(id_pot),
                        appui_num,
                        appui_commune,
                        type_bpe,
                        probleme,
                        explication
                    ])
            
            # Ajuster largeurs colonnes BPE
            ws_bpe.column_dimensions['A'].width = 30
            ws_bpe.column_dimensions['H'].width = 60
            
            # ===== FEUILLE 5: CABLES C6 ABSENTS DE QGIS =====
            # SENS CORRECT: C6 → QGIS (câbles déclarés dans C6 mais non trouvés dans QGIS/GraceTHD)
            ws_cables = wb.create_sheet("Câbles C6 Absents")
            ws_cables.append(["Etude", "Ligne C6", "Appui Origine (C6)", "Appui Destination (C6)", "Capacité FO (C6)", "Type PBO", "Problème", "Explication"])
            for col in range(1, 9):
                ws_cables.cell(row=1, column=col).font = header_font
                ws_cables.cell(row=1, column=col).fill = header_fill
            
            for s in stats_globales:
                for cable in s.get('detail_cables_c6_absents', []):
                    # cable = dict avec: ligne_c6, appui_origine, capacite, appui_destination, type_pbo
                    ligne_c6 = cable.get('ligne_c6', '')
                    appui_orig = str(cable.get('appui_origine', ''))
                    capa = cable.get('capacite', '')
                    appui_dest = str(cable.get('appui_destination', ''))
                    type_pbo = str(cable.get('type_pbo', ''))
                    
                    # Analyser le problème - SENS C6 → QGIS
                    probleme = "CÂBLE C6 ABSENT DE QGIS"
                    explication = (
                        f"L'Annexe C6 (ligne {ligne_c6}) déclare un câble entre l'appui {appui_orig} et l'appui {appui_dest} "
                        f"avec une capacité de {capa} fibres. Ce câble n'existe PAS dans les données QGIS/GraceTHD. "
                        f"ACTION: (1) Vérifier si ce câble a été supprimé de QGIS par erreur, "
                        f"(2) Vérifier si les numéros d'appuis dans le C6 sont corrects, "
                        f"(3) Vérifier si la capacité déclarée correspond à celle de QGIS."
                    )
                    
                    ws_cables.append([
                        s['etude'], 
                        str(ligne_c6), 
                        appui_orig, 
                        appui_dest, 
                        str(capa), 
                        type_pbo,
                        probleme,
                        explication
                    ])
            
            # Ajuster largeurs colonnes Câbles
            ws_cables.column_dimensions['A'].width = 30
            ws_cables.column_dimensions['H'].width = 80
            
            # Sauvegarder
            wb.save(filepath)
            QgsMessageLog.logMessage(f"Export Excel créé: {filepath}", "PoleAerien", Qgis.Info)
            return filepath
            
        except Exception as e:
            QgsMessageLog.logMessage(f"Erreur export Excel: {e}", "PoleAerien", Qgis.Warning)
            self.message_received.emit(f"Erreur export Excel: {e}", "red")
            return None
