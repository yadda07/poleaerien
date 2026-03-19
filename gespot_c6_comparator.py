# -*- coding: utf-8 -*-
"""
Lecture des fichiers C6 Excel et comparaison champ par champ avec GESPOT.

Responsabilité unique : lire les colonnes utiles de l'onglet Export 1,
agréger la whitelist depuis l'onglet Bases, puis produire les statuts
de comparaison C1-C11 par appui.

Zéro dépendance QGIS — thread-safe.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import openpyxl

from .core_utils import is_plugin_output_file, normalize_appui_num
from .gespot_reader import GespotRecord, GespotLoadResult


# ===========================================================================
#  CONSTANTES — indices colonnes C6 (0-bases depuis le header ligne 8)
# ===========================================================================

_C6_HDR_ROW = 7        # header ligne 8 (0-base)
_C6_DATA_ROW = 8       # données à partir ligne 9 (0-base)
_C6_CENTRE_ROW = 2     # cellule E3 (0-base)
_C6_CENTRE_COL = 4     # colonne E (0-base)

_C6_COL_NUM = 0        # A — N° appui
_C6_COL_TYPE = 1       # B — Type d'appui
_C6_COL_ADRESSE = 2    # C — Adresse
_C6_COL_CTRL_VIS = 5   # F — Contrôle visuel
_C6_COL_VERTIC = 6     # G — Verticalité
_C6_COL_YELLOW = 12    # M — Absence étiquette jaune
_C6_COL_USABLE = 13    # N — Appui utilisable
_C6_COL_ENV = 14       # O — Milieu environnant
_C6_COL_ELEC = 15      # P — Environnement électrique
_C6_COL_STRAT = 16     # Q — Appui stratégique
_C6_COL_INACC = 17     # R — Appui inaccessible

_BASES_STRAT_COL = 13  # colonne M (1-base pour openpyxl)
_BASES_STRAT_ROW_START = 2  # ligne 2 (1-base)
_BASES_STRAT_EXCLUDED = {'non', 'appui stratégique', ''}

_COMPARISONS = [
    'adresse', 'centre', 'type', 'strategie', 'env',
    'elec', 'inacc', 'vertic', 'yellow', 'usable', 'ctrl_vis',
]


# ===========================================================================
#  DATACLASSES
# ===========================================================================

@dataclass
class C6Record:
    """Un appui C6 extrait de l'onglet Export 1."""
    num: str
    adresse: str
    centre: str
    type_c6: str
    ctrl_vis: str
    vertic: str
    yellow: str
    usable: str
    env: str
    elec: str
    strat: str
    inacc: str
    source_file: str


@dataclass
class ComparisonResult:
    """Résultat de comparaison pour un appui."""
    num: str

    voie_gespot: str
    num_voie_gespot: str
    adresse_c6: str
    statut_adresse: str
    detail_adresse: str

    centre_gespot: str
    centre_c6: str
    statut_centre: str
    detail_centre: str

    type_gespot: str
    type_c6: str
    statut_type: str
    detail_type: str

    strat_gespot: str
    strat_c6: str
    statut_strat: str
    detail_strat: str

    env_gespot: str
    env_c6: str
    statut_env: str
    detail_env: str

    elec_gespot: str
    elec_c6: str
    statut_elec: str
    detail_elec: str

    inacc_gespot: str
    inacc_c6: str
    statut_inacc: str
    detail_inacc: str

    recalage_gespot: str
    vertic_c6: str
    statut_vertic: str
    detail_vertic: str

    yellow_gespot: str
    yellow_c6: str
    statut_yellow: str
    detail_yellow: str

    usable_gespot: str
    usable_c6: str
    statut_usable: str
    detail_usable: str

    ctrl_vis_gespot: str
    ctrl_vis_c6: str
    statut_ctrl_vis: str
    detail_ctrl_vis: str

    dist_elec_gespot: str

    nb_ecarts: int
    statut_global: str

    source_gespot: str
    source_c6: str


@dataclass
class C6LoadResult:
    """Résultat de chargement d'un répertoire C6."""
    records: Dict[str, C6Record] = field(default_factory=dict)
    whitelist: set = field(default_factory=set)
    anomalies: List[dict] = field(default_factory=list)


@dataclass
class GespotC6Result:
    """Résultat global de la comparaison GESPOT vs C6."""
    comparisons: List[ComparisonResult] = field(default_factory=list)
    absent_c6: List[GespotRecord] = field(default_factory=list)
    absent_gespot: List[C6Record] = field(default_factory=list)
    anomalies: List[dict] = field(default_factory=list)
    output_path: str = ''


# ===========================================================================
#  LECTURE C6
# ===========================================================================

def _cell_str(ws, row: int, col: int) -> str:
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ''


def _safe_cell(row_tuple, idx: int) -> str:
    """Extrait une cellule d'un tuple openpyxl en toute securite."""
    if idx < len(row_tuple) and row_tuple[idx] is not None:
        return str(row_tuple[idx]).strip()
    return ''


def _read_whitelist(wb: openpyxl.Workbook) -> set:
    whitelist = set()
    if 'Bases' not in wb.sheetnames:
        return whitelist
    ws = wb['Bases']
    for row_idx in range(_BASES_STRAT_ROW_START, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=_BASES_STRAT_COL).value
        if val is None:
            continue
        s = str(val).strip()
        if s.lower() not in _BASES_STRAT_EXCLUDED:
            whitelist.add(s)
    return whitelist


def _read_one_c6(filepath: str,
                 fname: str) -> Tuple[Optional[dict], set, List[dict]]:
    """Lit un fichier C6 Excel. Retourne (records_dict, whitelist, anomalies)."""
    anomalies = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        anomalies.append({
            'source': 'C6', 'fichier': fname, 'num': '',
            'type': 'FICHIER_IGNORE', 'detail': str(e),
            'action': 'Verifier que le fichier est un .xlsx valide',
        })
        return None, set(), anomalies

    if 'Export 1' not in wb.sheetnames:
        wb.close()
        return None, set(), anomalies

    ws = wb['Export 1']
    whitelist = _read_whitelist(wb)
    centre_raw = _cell_str(ws, _C6_CENTRE_ROW + 1, _C6_CENTRE_COL + 1)
    if not centre_raw:
        anomalies.append({
            'source': 'C6', 'fichier': fname, 'num': '',
            'type': 'CENTRE_C6_ABSENT',
            'detail': 'Cellule E3 vide dans Export 1',
            'action': 'Renseigner le code centre dans la cellule E3',
        })

    records = {}
    for row in ws.iter_rows(min_row=_C6_DATA_ROW + 1, values_only=True):
        if not row or row[_C6_COL_NUM] is None:
            continue
        num_raw = str(row[_C6_COL_NUM]).strip()
        num_key = normalize_appui_num(num_raw)
        if not num_key:
            continue

        records[num_key] = C6Record(
            num=num_key,
            adresse=_safe_cell(row, _C6_COL_ADRESSE),
            centre=centre_raw,
            type_c6=_safe_cell(row, _C6_COL_TYPE),
            ctrl_vis=_safe_cell(row, _C6_COL_CTRL_VIS),
            vertic=_safe_cell(row, _C6_COL_VERTIC),
            yellow=_safe_cell(row, _C6_COL_YELLOW),
            usable=_safe_cell(row, _C6_COL_USABLE),
            env=_safe_cell(row, _C6_COL_ENV),
            elec=_safe_cell(row, _C6_COL_ELEC),
            strat=_safe_cell(row, _C6_COL_STRAT),
            inacc=_safe_cell(row, _C6_COL_INACC),
            source_file=fname,
        )
    wb.close()
    return records, whitelist, anomalies


def load_c6_dir(c6_dir: str) -> C6LoadResult:
    """Charge tous les fichiers C6 d'un répertoire (récursif)."""
    result = C6LoadResult()
    seen_nums: Dict[str, dict] = {}

    for subdir, _, files in os.walk(c6_dir):
        for fname in sorted(files):
            if not fname.lower().endswith('.xlsx') or '~$' in fname:
                continue
            if is_plugin_output_file(fname):
                continue
            path = os.path.join(subdir, fname)
            records, wl, anoms = _read_one_c6(path, fname)
            result.anomalies.extend(anoms)
            result.whitelist |= wl
            if records is None:
                continue
            for num_key, rec in records.items():
                if num_key in seen_nums:
                    _handle_c6_dup(num_key, rec, seen_nums[num_key],
                                   result.anomalies, result.records)
                else:
                    result.records[num_key] = rec
                    seen_nums[num_key] = {'rec': rec, 'payload': str(rec)}
    return result


def _handle_c6_dup(num: str, new_rec: C6Record,
                   existing: dict, anomalies: List[dict],
                   records: Dict[str, C6Record]) -> None:
    if str(new_rec) == existing['payload']:
        anomalies.append({
            'source': 'C6', 'fichier': new_rec.source_file, 'num': num,
            'type': 'DOUBLON_IDENTIQUE',
            'detail': 'Meme appui dans plusieurs fichiers C6 (identique)',
            'action': 'Aucune (consolide automatiquement)',
        })
    else:
        records.pop(num, None)
        anomalies.append({
            'source': 'C6',
            'fichier': f"{existing['rec'].source_file} / {new_rec.source_file}",
            'num': num,
            'type': 'DOUBLON_CONFLICTUEL',
            'detail': 'Valeurs differentes dans plusieurs fichiers C6',
            'action': 'Corriger la source C6 avant de relancer',
        })


# ===========================================================================
#  COMPARAISON
# ===========================================================================

def _cmp(a: str, b: str, label_gespot: str = 'GESPOT',
         label_c6: str = 'C6') -> Tuple[str, str]:
    """Compare deux valeurs, retourne (statut OK/KO, detail explicatif)."""
    va, vb = a.strip(), b.strip()
    na, nb = va.lower(), vb.lower()
    if na == nb:
        return 'OK', ''
    if va and not vb:
        return 'KO', f"{label_gespot}='{va}' mais {label_c6} est vide"
    if not va and vb:
        return 'KO', f"{label_gespot} est vide mais {label_c6}='{vb}'"
    return 'KO', f"{label_gespot}='{va}' vs {label_c6}='{vb}'"


def _recalage_to_vertic(recalage_raw: str) -> str:
    """BR-06 : transforme RECALAGE brut en valeur attendue C6 col G."""
    return 'Non' if recalage_raw.strip().upper() == 'O' else 'Oui'


def _compare_one(gespot: GespotRecord, c6: C6Record) -> ComparisonResult:
    vertic_gespot = _recalage_to_vertic(gespot.recalage_raw)

    s_adr, d_adr = _cmp(gespot.voie, c6.adresse, 'Voie GESPOT', 'Adresse C6')
    s_ctr, d_ctr = _cmp(gespot.centre, c6.centre, 'Centre GESPOT', 'Centre C6')
    s_typ, d_typ = _cmp(gespot.type_calc, c6.type_c6, 'Type GESPOT', 'Type C6')
    s_str, d_str = _cmp(gespot.strategie_calc, c6.strat, 'Strategie GESPOT', 'Strategie C6')
    s_env, d_env = _cmp(gespot.milieu_calc, c6.env, 'Milieu GESPOT', 'Milieu C6')
    s_elc, d_elc = _cmp(gespot.pres_elect_calc, c6.elec, 'Risque elec GESPOT', 'Risque elec C6')
    s_ina, d_ina = _cmp(gespot.inacc_calc, c6.inacc, 'Inacc GESPOT', 'Inacc C6')
    s_vtc, d_vtc = _cmp(vertic_gespot, c6.vertic, 'Verticalite GESPOT', 'Verticalite C6')
    s_ylw, d_ylw = _cmp(gespot.etat_no_yellow, c6.yellow, 'Etiquette jaune GESPOT', 'Etiquette jaune C6')
    s_usb, d_usb = _cmp(gespot.etat_usable, c6.usable, 'Utilisable GESPOT', 'Utilisable C6')
    s_ctv, d_ctv = _cmp(gespot.ctrl_visuel, c6.ctrl_vis, 'Ctrl visuel GESPOT', 'Ctrl visuel C6')

    all_statuts = [s_adr, s_ctr, s_typ, s_str, s_env, s_elc, s_ina, s_vtc, s_ylw, s_usb, s_ctv]
    nb_ecarts = sum(1 for s in all_statuts if s == 'KO')

    return ComparisonResult(
        num=gespot.num,
        voie_gespot=gespot.voie, num_voie_gespot=gespot.num_voie,
        adresse_c6=c6.adresse, statut_adresse=s_adr, detail_adresse=d_adr,
        centre_gespot=gespot.centre, centre_c6=c6.centre, statut_centre=s_ctr, detail_centre=d_ctr,
        type_gespot=gespot.type_calc, type_c6=c6.type_c6, statut_type=s_typ, detail_type=d_typ,
        strat_gespot=gespot.strategie_calc, strat_c6=c6.strat, statut_strat=s_str, detail_strat=d_str,
        env_gespot=gespot.milieu_calc, env_c6=c6.env, statut_env=s_env, detail_env=d_env,
        elec_gespot=gespot.pres_elect_calc, elec_c6=c6.elec, statut_elec=s_elc, detail_elec=d_elc,
        inacc_gespot=gespot.inacc_calc, inacc_c6=c6.inacc, statut_inacc=s_ina, detail_inacc=d_ina,
        recalage_gespot=gespot.recalage_raw,
        vertic_c6=c6.vertic, statut_vertic=s_vtc, detail_vertic=d_vtc,
        yellow_gespot=gespot.etat_no_yellow, yellow_c6=c6.yellow, statut_yellow=s_ylw, detail_yellow=d_ylw,
        usable_gespot=gespot.etat_usable, usable_c6=c6.usable, statut_usable=s_usb, detail_usable=d_usb,
        ctrl_vis_gespot=gespot.ctrl_visuel, ctrl_vis_c6=c6.ctrl_vis, statut_ctrl_vis=s_ctv, detail_ctrl_vis=d_ctv,
        dist_elec_gespot=gespot.dist_elec,
        nb_ecarts=nb_ecarts,
        statut_global='OK' if nb_ecarts == 0 else 'KO',
        source_gespot=gespot.source_file,
        source_c6=c6.source_file,
    )


def compare(gespot_result: GespotLoadResult,
            c6_result: C6LoadResult) -> GespotC6Result:
    """Effectue la jointure et la comparaison champ par champ."""
    out = GespotC6Result()
    out.anomalies.extend(gespot_result.anomalies)
    out.anomalies.extend(c6_result.anomalies)

    gespot_keys = set(gespot_result.records)
    c6_keys = set(c6_result.records)

    for num in sorted(gespot_keys & c6_keys):
        out.comparisons.append(
            _compare_one(gespot_result.records[num], c6_result.records[num])
        )

    for num in sorted(gespot_keys - c6_keys):
        out.absent_c6.append(gespot_result.records[num])

    for num in sorted(c6_keys - gespot_keys):
        out.absent_gespot.append(c6_result.records[num])

    return out


# ===========================================================================
#  EXPORT EXCEL
# ===========================================================================

def export_to_excel(result: GespotC6Result, export_dir: str) -> str:
    """Génère le rapport GESPOT_C6_ANALYSE_{timestamp}.xlsx."""
    if not os.path.isdir(export_dir):
        os.makedirs(export_dir, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filepath = os.path.join(export_dir, f'GESPOT_C6_ANALYSE_{ts}.xlsx')
    wb = openpyxl.Workbook()

    _write_analyse(wb, result.comparisons)
    _write_absent_c6(wb, result.absent_c6)
    _write_absent_gespot(wb, result.absent_gespot)
    _write_anomalies(wb, result.anomalies)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    try:
        wb.save(filepath)
    except (PermissionError, OSError) as exc:
        raise RuntimeError(
            f"Impossible d'ecrire le rapport: {filepath} "
            f"(fichier ouvert dans Excel ?) — {exc}"
        ) from exc
    finally:
        wb.close()
    result.output_path = filepath
    return filepath


def _fills():
    from openpyxl.styles import PatternFill
    return {
        'OK': PatternFill('solid', fgColor='C6EFCE'),
        'KO': PatternFill('solid', fgColor='FFC7CE'),
    }


def _hdr_fill():
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor='D9D9D9')


def _write_analyse(wb: openpyxl.Workbook,
                   comparisons: List[ComparisonResult]) -> None:
    ws = wb.create_sheet('ANALYSE')
    headers = [
        'NUM', 'VOIE', 'NUM_VOIE', 'ADRESSE_C6', 'STATUT_ADRESSE', 'DETAIL_ADRESSE',
        'CENTRE_GESPOT', 'CENTRE_C6', 'STATUT_CENTRE', 'DETAIL_CENTRE',
        'TYPE_GESPOT', 'TYPE_C6', 'STATUT_TYPE', 'DETAIL_TYPE',
        'STRAT_GESPOT', 'STRAT_C6', 'STATUT_STRAT', 'DETAIL_STRAT',
        'ENV_GESPOT', 'ENV_C6', 'STATUT_ENV', 'DETAIL_ENV',
        'ELEC_GESPOT', 'ELEC_C6', 'STATUT_ELEC', 'DETAIL_ELEC',
        'INACC_GESPOT', 'INACC_C6', 'STATUT_INACC', 'DETAIL_INACC',
        'RECALAGE', 'VERTIC_C6', 'STATUT_VERTIC', 'DETAIL_VERTIC',
        'YELLOW_GESPOT', 'YELLOW_C6', 'STATUT_YELLOW', 'DETAIL_YELLOW',
        'USABLE_GESPOT', 'USABLE_C6', 'STATUT_USABLE', 'DETAIL_USABLE',
        'CTRL_VIS_GESPOT', 'CTRL_VIS_C6', 'STATUT_CTRL_VIS', 'DETAIL_CTRL_VIS',
        'DIST_ELEC', 'NB_KO', 'STATUT_GLOBAL', 'SOURCE_GESPOT', 'SOURCE_C6',
    ]
    statut_cols = {i + 1 for i, h in enumerate(headers) if h.startswith('STATUT_')}

    from openpyxl.styles import Font, PatternFill as _PF
    hfill = _hdr_fill()
    hfont = Font(bold=True)
    nb_ko_col = headers.index('NB_KO') + 1
    statut_global_col = headers.index('STATUT_GLOBAL') + 1
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hfill
        cell.font = hfont

    fills = _fills()
    zebra = _PF('solid', fgColor='F2F2F2')

    for row_idx, cmp in enumerate(comparisons, 2):
        values = [
            cmp.num, cmp.voie_gespot, cmp.num_voie_gespot,
            cmp.adresse_c6, cmp.statut_adresse, cmp.detail_adresse,
            cmp.centre_gespot, cmp.centre_c6, cmp.statut_centre, cmp.detail_centre,
            cmp.type_gespot, cmp.type_c6, cmp.statut_type, cmp.detail_type,
            cmp.strat_gespot, cmp.strat_c6, cmp.statut_strat, cmp.detail_strat,
            cmp.env_gespot, cmp.env_c6, cmp.statut_env, cmp.detail_env,
            cmp.elec_gespot, cmp.elec_c6, cmp.statut_elec, cmp.detail_elec,
            cmp.inacc_gespot, cmp.inacc_c6, cmp.statut_inacc, cmp.detail_inacc,
            cmp.recalage_gespot, cmp.vertic_c6, cmp.statut_vertic, cmp.detail_vertic,
            cmp.yellow_gespot, cmp.yellow_c6, cmp.statut_yellow, cmp.detail_yellow,
            cmp.usable_gespot, cmp.usable_c6, cmp.statut_usable, cmp.detail_usable,
            cmp.ctrl_vis_gespot, cmp.ctrl_vis_c6, cmp.statut_ctrl_vis, cmp.detail_ctrl_vis,
            cmp.dist_elec_gespot, cmp.nb_ecarts, cmp.statut_global,
            cmp.source_gespot, cmp.source_c6,
        ]
        bg = zebra if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if bg and col_idx not in statut_cols and col_idx != nb_ko_col and col_idx != statut_global_col:
                cell.fill = bg
        for col_idx in statut_cols | {statut_global_col}:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fills.get(str(cell.value), fills['KO'])
        nb_cell = ws.cell(row=row_idx, column=nb_ko_col)
        nb_cell.fill = fills['KO'] if cmp.nb_ecarts > 0 else fills['OK']

    nb_ok = sum(1 for c in comparisons if c.statut_global == 'OK')
    nb_ko = len(comparisons) - nb_ok
    summary_row = len(comparisons) + 3
    ws.cell(row=summary_row, column=1, value=f"Total: {len(comparisons)} appuis | {nb_ok} OK | {nb_ko} KO")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    ws.freeze_panes = 'A2'
    _auto_widths(ws)


def _write_absent_c6(wb: openpyxl.Workbook,
                     records: List[GespotRecord]) -> None:
    ws = wb.create_sheet('ABSENT_C6')
    headers = ['NUM', 'VOIE', 'NUM_VOIE', 'CENTRE', 'TYPE_GESPOT',
               'STRAT_GESPOT', 'ENV_GESPOT', 'ELEC_GESPOT',
               'INACC_GESPOT', 'RECALAGE', 'DIST_ELEC', 'SOURCE_GESPOT']
    _write_sheet_header(ws, headers)
    for r in records:
        ws.append([r.num, r.voie, r.num_voie, r.centre, r.type_calc,
                   r.strategie_calc, r.milieu_calc, r.pres_elect_calc,
                   r.inacc_calc, r.recalage_raw, r.dist_elec, r.source_file])
    ws.freeze_panes = 'A2'
    _auto_widths(ws)


def _write_absent_gespot(wb: openpyxl.Workbook,
                         records: List[C6Record]) -> None:
    ws = wb.create_sheet('ABSENT_GESPOT')
    headers = ['NUM', 'ADRESSE_C6', 'CENTRE_C6', 'TYPE_C6', 'STRAT_C6',
               'ENV_C6', 'ELEC_C6', 'INACC_C6', 'VERTIC_C6', 'SOURCE_C6']
    _write_sheet_header(ws, headers)
    for r in records:
        ws.append([r.num, r.adresse, r.centre, r.type_c6, r.strat,
                   r.env, r.elec, r.inacc, r.vertic, r.source_file])
    ws.freeze_panes = 'A2'
    _auto_widths(ws)


def _write_anomalies(wb: openpyxl.Workbook,
                     anomalies: List[dict]) -> None:
    ws = wb.create_sheet('ANOMALIES_SOURCE')
    headers = ['SOURCE', 'FICHIER', 'NUM', 'TYPE_ANOMALIE', 'DETAIL',
               'ACTION_REQUISE']
    _write_sheet_header(ws, headers)
    for a in anomalies:
        ws.append([a.get('source', ''), a.get('fichier', ''),
                   a.get('num', ''), a.get('type', ''),
                   a.get('detail', ''), a.get('action', '')])
    ws.freeze_panes = 'A2'
    _auto_widths(ws)


def _write_sheet_header(ws, headers: List[str]) -> None:
    from openpyxl.styles import Font, PatternFill
    hfill = _hdr_fill()
    hfont = Font(bold=True)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hfill
        cell.font = hfont


def _auto_widths(ws, max_width: int = 50) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        width = max(
            (len(str(cell.value or '')) for cell in col),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            width + 2, max_width
        )


# ===========================================================================
#  POINT D'ENTRÉE PRINCIPAL
# ===========================================================================

def run_comparison(gespot_dir: str, c6_dir: str,
                   export_dir: str,
                   progress_cb=None) -> GespotC6Result:
    """Orchestre lecture → comparaison → export.

    Args:
        gespot_dir: Chemin du dossier GESPOT/GSPOT.
        c6_dir:     Chemin du dossier C6 (= CAP FT).
        export_dir: Dossier de sortie.
        progress_cb: callable(pct, msg) optionnel.

    Returns:
        GespotC6Result avec output_path renseigné.
    """
    def _prog(pct, msg=''):
        if progress_cb:
            progress_cb(pct, msg)

    _prog(5, 'Lecture C6...')
    c6_result = load_c6_dir(c6_dir)

    _prog(30, 'Lecture GESPOT...')
    from .gespot_reader import load_gespot_dir
    gespot_result = load_gespot_dir(gespot_dir, whitelist=c6_result.whitelist)

    _prog(60, 'Comparaison...')
    result = compare(gespot_result, c6_result)

    _prog(100, 'Termine.')
    return result
