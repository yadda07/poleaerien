# -*- coding: utf-8 -*-
"""
Unified Excel Report - Professional diagnostic workbook.

Generates a single standardized Excel file with:
  1. TABLEAU DE BORD: executive summary with KPIs per module
  2. Module sheets: consistent formatting, borders, filters, freeze panes

Style: Calibri 10pt, dark blue headers (#1F4E79), thin grey borders,
       status fills (green=OK, amber=warning, red=error).
"""

import json
import os
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ======================================================================
#  UNIFIED STYLE SYSTEM
# ======================================================================

_TAB = {
    'maj': 'F59E0B', 'capft': '34D399', 'gespot_c6': '0EA5E9',
    'comac': 'EA580C', 'c6bd': 'EC4899', 'police_c6': 'A855F7', 'c6c3a': '3B82F6',
}
_NAMES = {
    'maj': '0. MAJ BD', 'capft': '1. VERIF CAP_FT', 'gespot_c6': '2. GESPOT vs C6',
    'comac': '3. VERIF COMAC', 'c6bd': '4. C6 vs BD',
    'police_c6': '5. POLICE C6', 'c6c3a': '6. C6-C3A-BD',
}

_F_TITLE = Font(name='Calibri', size=14, bold=True, color='1F4E79')
_F_HEAD = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
_F_DATA = Font(name='Calibri', size=10)
_F_BOLD = Font(name='Calibri', size=10, bold=True)
_F_PCT = Font(name='Calibri', size=11, bold=True)

_P_HEAD = PatternFill('solid', fgColor='1F4E79')
_P_OK = PatternFill('solid', fgColor='C6EFCE')
_P_WARN = PatternFill('solid', fgColor='FFEB9C')
_P_ERR = PatternFill('solid', fgColor='FFC7CE')
_P_CRIT = PatternFill('solid', fgColor='FF6B6B')
_P_INFO = PatternFill('solid', fgColor='D6E4F0')

_BRD = Border(
    left=Side('thin', color='D9D9D9'), right=Side('thin', color='D9D9D9'),
    top=Side('thin', color='D9D9D9'), bottom=Side('thin', color='D9D9D9'),
)
_AL_C = Alignment(horizontal='center', vertical='center')
_AL_L = Alignment(horizontal='left', vertical='center')
_AL_W = Alignment(horizontal='left', vertical='center', wrap_text=True)
_AL_CV = Alignment(horizontal='center', vertical='center', wrap_text=True)

_F_MOD = Font(name='Calibri', size=10, bold=True, color='1F4E79')

_DRAWING_COLUMNS = 13
_DRAWING_IMAGE_WIDTH = 980
_DRAWING_IMAGE_HEIGHT = 715
_DRAWING_SLOT_ROWS = (4, 36)
_DRAWING_IMAGE_ANCHORS = ('A5', 'A37')
_DRAWING_ROW_HEIGHT = 18
_DRAWING_SLOTS_PER_PAGE = 2
_DRAWING_SHEET_NAME_MAX = 31


# ======================================================================
#  HELPERS
# ======================================================================

def _init_sheet(ws, key, headers, widths):
    """Apply standard header, tab color, freeze panes, auto-filter."""
    ws.sheet_properties.tabColor = _TAB.get(key, '808080')
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _F_HEAD
        cell.fill = _P_HEAD
        cell.alignment = _AL_C
        cell.border = _BRD
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'


def _row(ws, r, vals, font=None, fill=None, align=None):
    """Write one data row with consistent styling."""
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or _F_DATA
        cell.alignment = align or _AL_L
        cell.border = _BRD
        if fill:
            cell.fill = fill


def _fill_row(ws, r, fill, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = _BRD


def _df_sheet(wb, name, df, key):
    """Write DataFrame to a new sheet with standard formatting."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = _TAB.get(key, '808080')
    ncols = len(df.columns)
    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=str(col_name))
        cell.font = _F_HEAD
        cell.fill = _P_HEAD
        cell.alignment = _AL_C
        cell.border = _BRD
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(str(col_name)) + 4)
    for ri, row_data in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            safe_val = val if pd.notna(val) else ''
            cell = ws.cell(row=ri, column=ci, value=safe_val)
            cell.font = _F_DATA
            cell.border = _BRD
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(ncols)}1'
    return ws


def _write_comac_drawing_sheets(wb, result, report_options):
    etudes_pcm = result.get('etudes_pcm') or {}
    if not etudes_pcm:
        return True
    try:
        try:
            from .pcm_drawing import PcmDrawingRenderer
        except ImportError:
            from pcm_drawing import PcmDrawingRenderer
        stop_fn = lambda: _report_cancelled(report_options)
        renderer = PcmDrawingRenderer()
        entries = renderer.build_support_entries(etudes_pcm, stop_fn)
    except Exception as exc:
        _report_message(report_options, f"[REPORT] Erreur rendu dessins COMAC: {exc}", 'orange')
        return True
    if entries is None or _report_cancelled(report_options):
        return False
    if not entries:
        return True
    entries_by_etude = {}
    for entry in entries:
        entries_by_etude.setdefault(entry['etude'], []).append(entry)
    error_count = len(result.get('erreurs_pcm') or {})
    total_sheets = sum(
        (len(grp) + _DRAWING_SLOTS_PER_PAGE - 1) // _DRAWING_SLOTS_PER_PAGE
        for grp in entries_by_etude.values()
    )
    sheet_counter = 0
    for etude_name, etude_entries in entries_by_etude.items():
        if _report_cancelled(report_options):
            return False
        pages = (len(etude_entries) + _DRAWING_SLOTS_PER_PAGE - 1) // _DRAWING_SLOTS_PER_PAGE
        diagrams = renderer.render_entries(etude_entries, stop_fn)
        for page_idx in range(pages):
            if _report_cancelled(report_options):
                return False
            sheet_name = _drawing_sheet_name(etude_name, page_idx if pages > 1 else -1)
            ws = wb.create_sheet(sheet_name)
            _init_comac_drawing_sheet(
                ws, etude_name, page_idx + 1, pages,
                len(etude_entries), len(entries), error_count,
            )
            for slot_idx in range(_DRAWING_SLOTS_PER_PAGE):
                diagram = next(diagrams, None)
                if diagram is None:
                    break
                _place_comac_drawing(ws, slot_idx, diagram)
            sheet_counter += 1
            _report_drawings_progress(report_options, sheet_counter, total_sheets)
    return True


def _drawing_sheet_name(etude_name, page_idx):
    """Build Excel sheet name (max 31 chars) for a drawing page."""
    safe = (etude_name or 'COMAC').replace('/', '_').replace('\\', '_').replace(':', '_')
    prefix = 'DESSIN_'
    if page_idx >= 0:
        suffix = f'_{page_idx + 1:02d}'
    else:
        suffix = ''
    max_name = _DRAWING_SHEET_NAME_MAX - len(prefix) - len(suffix)
    if len(safe) > max_name:
        safe = safe[:max_name]
    return f"{prefix}{safe}{suffix}"


def _init_comac_drawing_sheet(
    ws, etude_name, page_idx, total_pages,
    etude_count, global_count, error_count,
):
    ws.sheet_properties.tabColor = _TAB.get('comac', 'EA580C')
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    for col_idx in range(1, _DRAWING_COLUMNS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    for row_idx in range(1, 69):
        ws.row_dimensions[row_idx].height = _DRAWING_ROW_HEIGHT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_DRAWING_COLUMNS)
    title_text = f"COMAC - {etude_name}"
    if total_pages > 1:
        title_text += f" ({page_idx}/{total_pages})"
    title = ws.cell(row=1, column=1, value=title_text)
    title.font = _F_TITLE
    title.alignment = _AL_C
    subtitle = f"{etude_count} appuis zone | {global_count} total"
    if error_count:
        subtitle += f" | {error_count} erreur(s) PCM"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_DRAWING_COLUMNS)
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = _F_MOD
    sub.alignment = _AL_C


def _place_comac_drawing(ws, slot_idx, diagram):
    top_row = _DRAWING_SLOT_ROWS[slot_idx]
    ws.merge_cells(start_row=top_row, start_column=1, end_row=top_row, end_column=_DRAWING_COLUMNS)
    header = ws.cell(
        row=top_row,
        column=1,
        value=f"{diagram.get('support', '')} | {diagram.get('connections', 0)} liaison(s)"
    )
    header.font = _F_BOLD
    header.fill = _P_INFO
    header.alignment = _AL_C
    header.border = _BRD
    stream = BytesIO(diagram.get('image_bytes', b''))
    image = ExcelImage(stream)
    image.width = _DRAWING_IMAGE_WIDTH
    image.height = _DRAWING_IMAGE_HEIGHT
    image._poleaerien_stream = stream
    ws.add_image(image, _DRAWING_IMAGE_ANCHORS[slot_idx])


# ======================================================================
#  KPI EXTRACTION (for dashboard)
# ======================================================================

def _kpi_maj(r):
    ft = r.get('liste_ft', [0, None, 0])
    bt = r.get('liste_bt', [0, None, 0])
    ok = (ft[2] if len(ft) > 2 else 0) + (bt[2] if len(bt) > 2 else 0)
    nok = (ft[0] if ft else 0) + (bt[0] if bt else 0)
    return ok, nok, f"FT:{ft[2] if len(ft)>2 else 0} id/{ft[0] if ft else 0} introuv | BT:{bt[2] if len(bt)>2 else 0} id/{bt[0] if bt else 0} introuv"

def _kpi_capft(r):
    res = r.get('resultats')
    if not res: return 0, 0, ''
    ne = sum(len(v) for v in res[0].values())
    nq = sum(len(v) for v in res[1].values())
    ok = len(res[2])
    return ok, ne + nq, f"{ok} OK | {ne} absents QGIS | {nq} absents Excel"

def _kpi_comac(r):
    res = r.get('resultats')
    if not res:
        return 0, 0, ''
    ne = sum(len(v) for v in res[0].values())
    nq = sum(len(v) for v in res[1].values())
    ok_name = len(res[2])
    ok_spatial = len(res[4]) if len(res) > 4 else 0
    ok = ok_name + ok_spatial
    hp = sum(len(v) for v in res[3].values()) if len(res) > 3 else 0
    nr = sum(len(v) for v in res[5].values()) if len(res) > 5 else 0
    base_nok = ne + nq + nr
    base_msg = f"{ok} OK ({ok_name} nom, {ok_spatial} spatial) | {ne} absents SRO | {nr} non resolus commune/code | {hp} hors perimetre | {nq} absents Excel"
    # Ajouter câbles + boîtiers au bilan
    verif_cables = r.get('verif_cables')
    cable_nok = 0
    if verif_cables:
        cable_nok = sum(1 for e in verif_cables if e.get('statut') == 'ECART' or e.get('statut', '').startswith('ABSENT'))
    verif_boitiers = r.get('verif_boitiers')
    boitier_nok = 0
    if verif_boitiers:
        boitier_nok = sum(1 for v in verif_boitiers.values() if v.get('statut') in ('ERREUR', 'ECART'))
    verif_portees = r.get('verif_portees')
    portee_nok = 0
    if verif_portees:
        portee_nok = sum(1 for e in verif_portees if e.get('statut') in ('ECART', 'ABSENT_REF'))
    total_nok = base_nok + cable_nok + boitier_nok + portee_nok
    parts = [base_msg] if base_msg else []
    if cable_nok:
        parts.append(f"{cable_nok} anomalies cables")
    if boitier_nok:
        parts.append(f"{boitier_nok} err boitiers")
    if portee_nok:
        parts.append(f"{portee_nok} ecarts portees")
    return ok, total_nok, ' | '.join(parts)

def _kpi_c6bd(r):
    df = r.get('final_df')
    if df is None or df.empty: return 0, 0, ''
    nok = int(df['Statut'].astype(str).str.contains('ABSENT').sum()) if 'Statut' in df.columns else 0
    ok = len(df) - nok
    return ok, nok, f"{ok} trouves | {nok} absents | {len(r.get('df_poteaux_out', pd.DataFrame()))} hors perimetre"

def _kpi_police(r):
    st = r.get('stats', [])
    ok = sum(s.get('nb_ok', 0) for s in st)
    nok = sum(s.get('nb_ecart', 0) + s.get('nb_absent', 0) + s.get('nb_boitier_err', 0) for s in st)
    return ok, nok, f"{ok} OK | {nok} anomalies sur {len(st)} etudes"

def _kpi_c6c3a(r):
    df = r.get('df_final')
    if df is None or df.empty: return 0, 0, ''
    nok = sum(int((df[c] == 'ABSENT').sum()) for c in df.columns if ('inf_num' in c or 'Excel' in c))
    return max(0, len(df) - nok), nok, f"{len(df)} lignes | {nok} ABSENT"

def _kpi_gespot_c6(r):
    nb_ok = r.get('nb_ok', 0)
    nb_ecart = r.get('nb_ecart', 0)
    nb_absent_c6 = r.get('nb_absent_c6', 0)
    nb_absent_gespot = r.get('nb_absent_gespot', 0)
    nb_anomalies = r.get('nb_anomalies', 0)
    return nb_ok, nb_ecart, (
        f"{nb_ok} OK | {nb_ecart} KO | "
        f"{nb_absent_c6} absents C6 | {nb_absent_gespot} absents GESPOT | "
        f"{nb_anomalies} anomalies"
    )

_KPI = {'maj': _kpi_maj, 'capft': _kpi_capft, 'gespot_c6': _kpi_gespot_c6,
        'comac': _kpi_comac, 'c6bd': _kpi_c6bd, 'police_c6': _kpi_police,
        'c6c3a': _kpi_c6c3a}


# ======================================================================
#  SUB-CHECKS (detailed verifications per module for dashboard)
# ======================================================================

def _checks_maj(r):
    ft = r.get('liste_ft', [0, None, 0])
    bt = r.get('liste_bt', [0, None, 0])
    ok_ft = ft[2] if len(ft) > 2 else 0
    nok_ft = ft[0] if ft else 0
    ok_bt = bt[2] if len(bt) > 2 else 0
    nok_bt = bt[0] if bt else 0
    return [
        ("Correspondance FT (Excel/QGIS)", ok_ft, nok_ft,
         f"{ok_ft} identifies, {nok_ft} introuvables"),
        ("Correspondance BT (Excel/QGIS)", ok_bt, nok_bt,
         f"{ok_bt} identifies, {nok_bt} introuvables"),
    ]


def _checks_capft(r):
    res = r.get('resultats')
    if not res:
        return [("Analyse appuis", 0, 0, "Pas de resultats")]
    ne = sum(len(v) for v in res[0].values())
    nq = sum(len(v) for v in res[1].values())
    ok = len(res[2])
    checks = [
        ("Appuis presents dans QGIS", ok, ne,
         f"{ok} trouves, {ne} absents QGIS"),
    ]
    if ok == 0 and nq > 0:
        checks.append(("Appuis presents dans fiches", nq, 0,
                        f"0 fiches trouvees, {nq} poteaux sans comparaison"))
    else:
        checks.append(("Appuis presents dans fiches", ok, nq,
                        f"{ok} trouves, {nq} absents fiches appuis"))
    return checks


def _checks_comac(r):
    res = r.get('resultats')
    checks = []
    if res:
        ne = sum(len(v) for v in res[0].values())
        nq = sum(len(v) for v in res[1].values())
        ok_name = len(res[2])
        ok_spatial = len(res[4]) if len(res) > 4 else 0
        ok = ok_name + ok_spatial
        hp = sum(len(v) for v in res[3].values()) if len(res) > 3 else 0
        nr = sum(len(v) for v in res[5].values()) if len(res) > 5 else 0
        detail_match = f"{ok} trouves ({ok_name} nom, {ok_spatial} spatial)" if ok_spatial else f"{ok} trouves"
        checks.append(("Correspondance appuis QGIS/Excel", ok, ne,
                        f"{detail_match}, {ne} absents couche SRO"))
        if nr:
            checks.append(("Appuis non resolus commune/code", 0, nr,
                            f"{nr} presents dans la couche SRO mais ambigu entre plusieurs communes ou non leves par le spatial"))
        if hp:
            checks.append(("Appuis hors perimetre etude", hp, 0,
                            f"{hp} existent dans couche SRO mais hors zones etude COMAC"))
        checks.append(("Appuis presents dans Excel", ok, nq,
                        f"{detail_match}, {nq} absents Excel"))
    dico_secu = r.get('dico_verif_secu')
    if dico_secu:
        ok_p, nok_p, ok_h, nok_h = 0, 0, 0, 0
        for liste_v in dico_secu.values():
            for v in liste_v:
                portee = v.get('portee', 0)
                hauteur = v.get('hauteur_sol', 0)
                if portee > 0:
                    vp = v.get('verif_portee')
                    if vp:
                        if vp.get('valide', True):
                            ok_p += 1
                        else:
                            nok_p += 1
                if hauteur > 0:
                    vh = v.get('verif_hauteur_sol')
                    if vh:
                        if vh.get('valide', True):
                            ok_h += 1
                        else:
                            nok_h += 1
        if ok_p + nok_p > 0:
            checks.append(("Verif portees NFC 11201", ok_p, nok_p,
                            f"{ok_p} OK, {nok_p} depassements"))
        if ok_h + nok_h > 0:
            checks.append(("Verif hauteurs sol", ok_h, nok_h,
                            f"{ok_h} OK, {nok_h} insuffisants"))
    verif_cables = r.get('verif_cables')
    if verif_cables:
        cables_only = [e for e in verif_cables if e.get('statut')]
        if cables_only:
            ok_c = sum(1 for e in cables_only if e.get('statut') == 'OK')
            ecart_c = sum(1 for e in cables_only if e.get('statut') == 'ECART')
            absent_c = sum(1 for e in cables_only if e.get('statut', '').startswith('ABSENT') and not e.get('statut', '').endswith('_NON_RESOLU') and not e.get('statut', '').endswith('_HORS_PERIMETRE'))
            nr_c = sum(1 for e in cables_only if e.get('statut', '').endswith('_NON_RESOLU'))
            hp_c = sum(1 for e in cables_only if e.get('statut', '').endswith('_HORS_PERIMETRE'))
            nok_c = ecart_c + absent_c + nr_c + hp_c
            checks.append(("Verif cables COMAC vs BDD", ok_c, nok_c,
                            f"{ok_c} OK, {ecart_c} ecarts, {absent_c} absents, {nr_c} non resolus, {hp_c} hors perimetre"))
    dico_boitier = r.get('dico_boitier_comac', {})
    verif_boitiers = r.get('verif_boitiers')
    if dico_boitier:
        nb_oui = sum(1 for v in dico_boitier.values() if str(v).lower() == 'oui')
        nb_non = sum(1 for v in dico_boitier.values() if str(v).lower() == 'non')
        if verif_boitiers:
            ok_b = sum(1 for v in verif_boitiers.values() if v.get('statut') == 'OK')
            ecart_b = sum(1 for v in verif_boitiers.values() if v.get('statut') == 'ECART')
            err_loc_b = sum(1 for v in verif_boitiers.values() if v.get('statut') == 'ERREUR' and v.get('bpe_noe_type') == 'appui non localisé')
            err_abs_b = sum(1 for v in verif_boitiers.values() if v.get('statut') == 'ERREUR' and v.get('bpe_noe_type') != 'appui non localisé')
            nok_b = ecart_b + err_loc_b + err_abs_b
            checks.append(("Verif boitiers COMAC vs BPE", ok_b, nok_b,
                            f"{ok_b} OK, {ecart_b} ecarts type, {err_abs_b} absents BPE, {err_loc_b} appuis non resolus ({nb_non} Non)"))
        else:
            checks.append(("Verif boitiers COMAC vs BPE", nb_non, 0,
                            f"{nb_oui} Oui, {nb_non} Non, aucune verif BPE necessaire"))
    verif_portees = r.get('verif_portees')
    if verif_portees:
        ok_p = sum(1 for e in verif_portees if e.get('statut') == 'OK')
        ecart_p = sum(1 for e in verif_portees if e.get('statut') == 'ECART')
        absent_p = sum(1 for e in verif_portees if e.get('statut') == 'ABSENT_REF')
        nok_p = ecart_p + absent_p
        checks.append(("Verif portees PCM vs reference", ok_p, nok_p,
                        f"{ok_p} OK, {ecart_p} ecarts, {absent_p} absents"))
    if not checks:
        checks = [("Analyse COMAC", 0, 0, "Pas de resultats")]
    return checks


def _checks_c6bd(r):
    checks = []
    df = r.get('final_df')
    if df is not None and not df.empty:
        nok = int(df['Statut'].astype(str).str.contains('ABSENT').sum()) \
            if 'Statut' in df.columns else 0
        ok = len(df) - nok
        checks.append(("Correspondance appuis C6/QGIS", ok, nok,
                        f"{ok} trouves, {nok} absents"))
    out_df = r.get('df_poteaux_out')
    n_out = len(out_df) if out_df is not None and not out_df.empty else 0
    if n_out > 0 and r.get('be_type') == 'axione':
        checks.append(("Poteaux hors perimetre", n_out, 0,
                        f"{n_out} hors zones etude (normal GraceTHD - zone elargie)"))
    else:
        checks.append(("Poteaux hors perimetre", 0, n_out,
                        f"{n_out} hors perimetre" if n_out > 0 else "Aucun"))
    verif = r.get('verif_etudes')
    if verif:
        sans_c6 = len(verif.get('etudes_sans_c6', []))
        c6_sans = len(verif.get('c6_sans_etude', []))
        nok_v = sans_c6 + c6_sans
        checks.append(("Coherence etudes/fichiers C6", 0, nok_v,
                        f"{sans_c6} etudes sans C6, {c6_sans} C6 sans etude"))
    if not checks:
        checks = [("Analyse C6 vs BD", 0, 0, "Pas de resultats")]
    return checks


def _checks_police(r):
    st = r.get('stats', [])
    if not st:
        return [("Analyse Police C6", 0, 0, "Pas de resultats")]
    ok_cables = sum(s.get('nb_ok', 0) for s in st)
    ecarts = sum(s.get('nb_ecart', 0) for s in st)
    absents = sum(s.get('nb_absent', 0) for s in st)
    boitier_err = sum(s.get('nb_boitier_err', 0) for s in st)
    total_appuis = sum(s.get('appuis_c6', 0) for s in st)
    return [
        ("Comparaison cables C6/BDD", ok_cables, ecarts,
         f"{ok_cables} OK, {ecarts} ecarts sur {len(st)} etudes"),
        ("Presence appuis dans BDD", total_appuis - absents, absents,
         f"{absents} absents BDD" if absents > 0 else "Tous presents"),
        ("Verification boitiers (BPE)", total_appuis - boitier_err, boitier_err,
         f"{boitier_err} erreurs boitier" if boitier_err > 0 else "Tous conformes"),
    ]


def _checks_c6c3a(r):
    df = r.get('df_final')
    if df is None or df.empty:
        return [("Analyse C6-C3A-BD", 0, 0, "Pas de resultats")]
    checks = []
    for col_check, label in [
        ("inf_num (ETUDES_QGIS)", "Presence dans etudes QGIS"),
        ("inf_num (C3A)", "Presence dans C3A"),
        ("Excel (C6)", "Presence dans C6 Excel"),
    ]:
        if col_check in df.columns:
            nok = int((df[col_check] == 'ABSENT').sum())
            ok = len(df) - nok
            checks.append((label, ok, nok, f"{ok} trouves, {nok} absents"))
    if not checks:
        nok = sum(int((df[c] == 'ABSENT').sum())
                  for c in df.columns if ('inf_num' in c or 'Excel' in c))
        ok = max(0, len(df) - nok)
        checks = [("Correspondance globale", ok, nok,
                    f"{len(df)} lignes, {nok} absents")]
    return checks


def _checks_gespot_c6(r):
    nb_compares = r.get('nb_compares', 0)
    nb_ok = r.get('nb_ok', 0)
    nb_ecart = r.get('nb_ecart', 0)
    nb_absent_c6 = r.get('nb_absent_c6', 0)
    nb_absent_gespot = r.get('nb_absent_gespot', 0)
    nb_anomalies = r.get('nb_anomalies', 0)
    checks = [
        ("Comparaison champ par champ", nb_ok, nb_ecart,
         f"{nb_compares} compares : {nb_ok} OK, {nb_ecart} KO"),
        ("Appuis absents C6 (informatif)", nb_absent_c6, 0,
         f"{nb_absent_c6} appuis GESPOT non trouves dans les fichiers C6 (couverture partielle normale)"),
        ("Appuis absents GESPOT (informatif)", nb_absent_gespot, 0,
         f"{nb_absent_gespot} appuis C6 non trouves dans le referentiel GESPOT"),
    ]
    if nb_anomalies > 0:
        checks.append(("Anomalies source (informatif)", 0, nb_anomalies,
                        f"{nb_anomalies} anomalies (doublons, fichiers invalides, etc.)"))
    return checks

_CHECKS = {
    'maj': _checks_maj, 'capft': _checks_capft, 'gespot_c6': _checks_gespot_c6,
    'comac': _checks_comac, 'c6bd': _checks_c6bd, 'police_c6': _checks_police,
    'c6c3a': _checks_c6c3a,
}


# ======================================================================
#  TABLEAU DE BORD
# ======================================================================

def _write_dashboard(wb, batch_results):
    ws = wb.active
    ws.title = "TABLEAU DE BORD"
    ws.sheet_properties.tabColor = '1F4E79'

    ws.merge_cells('A1:H1')
    t = ws.cell(row=1, column=1, value="RAPPORT DE DIAGNOSTIC - POLE AERIEN")
    t.font = _F_TITLE
    t.alignment = _AL_C

    ws.cell(row=2, column=1, value="Date :").font = _F_BOLD
    ws.cell(row=2, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M")).font = _F_DATA
    ws.cell(row=3, column=1, value="Modules :").font = _F_BOLD
    ws.cell(row=3, column=2, value=str(len(batch_results))).font = _F_DATA

    # QP-04: Resume executif — detect be_type and propagate to all results
    sro = ''
    be_label = ''
    summary_parts = []
    for key, res in batch_results.items():
        if not sro:
            sro = res.get('fddcpi_sro', '') or ''
        if not be_label:
            be_label = res.get('be_type', '')
    if be_label:
        for res in batch_results.values():
            if isinstance(res, dict) and 'be_type' not in res:
                res['be_type'] = be_label
    if sro:
        summary_parts.append(f"SRO {sro}")
    if be_label:
        summary_parts.append(be_label.upper())
    summary_parts.append(f"{len(batch_results)} module(s)")
    for key, kpi_fn in _KPI.items():
        res = batch_results.get(key)
        if res:
            ok, nok, msg = kpi_fn(res)
            if ok + nok > 0:
                name = _NAMES.get(key, key)
                summary_parts.append(f"{name}: {ok} OK/{ok+nok}")
    ws.merge_cells('A4:H4')
    c4 = ws.cell(row=4, column=1, value=' — '.join(summary_parts))
    c4.font = Font(name='Calibri', size=9, italic=True, color='555555')
    c4.alignment = _AL_C

    hdrs = ["MODULE", "VERIFICATION", "STATUT", "TOTAL",
            "CONFORMES", "NON CONFORMES", "% CONFORMITE", "DETAIL"]
    wds = [20, 32, 16, 10, 14, 16, 16, 50]
    for c, (h, w) in enumerate(zip(hdrs, wds), 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = _F_HEAD
        cell.fill = _P_HEAD
        cell.alignment = _AL_C
        cell.border = _BRD
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 6
    t_ok, t_nok = 0, 0

    for key in ['maj', 'capft', 'gespot_c6', 'comac', 'c6bd', 'police_c6', 'c6c3a']:
        res = batch_results.get(key)
        if res is None:
            continue

        checks = _CHECKS.get(key, lambda r: [("Analyse", 0, 0, "")])(res)
        if not checks:
            continue

        kpi_ok, kpi_nok, _ = _KPI.get(key, lambda r: (0, 0, ''))(res)
        t_ok += kpi_ok
        t_nok += kpi_nok

        start_row = row
        for i, (check_name, ok, nok, detail) in enumerate(checks):
            total = ok + nok
            pct = (ok / total * 100) if total > 0 else 100
            status = "CONFORME" if nok == 0 else "NON CONFORME"

            mod_name = _NAMES.get(key, key) if i == 0 else ""

            _row(ws, row, [mod_name, check_name, status, total, ok, nok,
                           f"{pct:.0f}%", detail])

            sf = _P_OK if nok == 0 else (_P_WARN if pct >= 80 else _P_ERR)
            ws.cell(row=row, column=3).fill = sf
            ws.cell(row=row, column=3).alignment = _AL_C

            pf = _P_OK if pct == 100 else (_P_WARN if pct >= 80 else _P_ERR)
            ws.cell(row=row, column=7).fill = pf
            ws.cell(row=row, column=7).alignment = _AL_C
            ws.cell(row=row, column=7).font = _F_PCT

            for c in [4, 5, 6]:
                ws.cell(row=row, column=c).alignment = _AL_C

            if i == 0:
                ws.cell(row=row, column=1).font = _F_MOD

            row += 1

        if len(checks) > 1:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=row - 1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = _AL_CV
            ws.cell(row=start_row, column=1).font = _F_MOD

    gt = t_ok + t_nok
    gp = (t_ok / gt * 100) if gt > 0 else 100
    _row(ws, row, ["TOTAL", "", "", gt, t_ok, t_nok, f"{gp:.0f}%", ""],
         font=_F_BOLD)
    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = _P_INFO
        ws.cell(row=row, column=c).border = Border(
            top=Side('medium', color='1F4E79'), bottom=Side('medium', color='1F4E79'),
            left=_BRD.left, right=_BRD.right)
    ws.cell(row=row, column=7).alignment = _AL_C
    ws.cell(row=row, column=7).font = _F_PCT

    lr = row + 2
    ws.cell(row=lr, column=1, value="LEGENDE :").font = _F_BOLD
    for i, (f, lbl) in enumerate([(_P_OK, "Conforme"), (_P_WARN, "Avertissement"),
                                   (_P_ERR, "Non conforme"), (_P_CRIT, "Critique")]):
        ws.cell(row=lr + 1 + i, column=1).fill = f
        ws.cell(row=lr + 1 + i, column=1).border = _BRD
        ws.cell(row=lr + 1 + i, column=2, value=lbl).font = _F_DATA

    ws.freeze_panes = 'A6'


# ======================================================================
#  MAJ BD
# ======================================================================

def write_maj(wb, result):
    """Write MAJ BD sheets: resume, FT, BT, introuvables."""
    ft = result.get('liste_ft', [0, pd.DataFrame(), 0, pd.DataFrame()])
    bt = result.get('liste_bt', [0, pd.DataFrame(), 0, pd.DataFrame()])
    nok_ft = ft[0] if ft else 0
    df_nok_ft = ft[1] if len(ft) > 1 else pd.DataFrame()
    ok_ft = ft[2] if len(ft) > 2 else 0
    df_ok_ft = ft[3] if len(ft) > 3 else pd.DataFrame()
    nok_bt = bt[0] if bt else 0
    df_nok_bt = bt[1] if len(bt) > 1 else pd.DataFrame()
    ok_bt = bt[2] if len(bt) > 2 else 0
    df_ok_bt = bt[3] if len(bt) > 3 else pd.DataFrame()
    key = 'maj'

    ws = wb.create_sheet("MAJ_RESUME")
    _init_sheet(ws, key, ["ELEMENT", "NOMBRE", "STATUT"], [40, 12, 18])
    data = [
        ("FT identifies", ok_ft, "OK" if ok_ft > 0 else ""),
        ("FT introuvables (Excel sans QGIS)", nok_ft,
         "ANOMALIE" if nok_ft > 0 else "OK"),
        ("BT identifies", ok_bt, "OK" if ok_bt > 0 else ""),
        ("BT introuvables (Excel sans QGIS)", nok_bt,
         "ANOMALIE" if nok_bt > 0 else "OK"),
    ]
    for r, (label, val, st) in enumerate(data, 2):
        _row(ws, r, [label, val, st])
        ws.cell(row=r, column=2).alignment = _AL_C
        ws.cell(row=r, column=3).alignment = _AL_C
        if st == "ANOMALIE":
            ws.cell(row=r, column=3).fill = _P_ERR
        elif st == "OK":
            ws.cell(row=r, column=3).fill = _P_OK

    if isinstance(df_ok_ft, pd.DataFrame) and not df_ok_ft.empty:
        _df_sheet(wb, "MAJ_FT", df_ok_ft.reset_index(), key)
    if isinstance(df_ok_bt, pd.DataFrame) and not df_ok_bt.empty:
        _df_sheet(wb, "MAJ_BT", df_ok_bt.reset_index(), key)
    if isinstance(df_nok_ft, pd.DataFrame) and not df_nok_ft.empty:
        _df_sheet(wb, "MAJ_FT_INTROUV", df_nok_ft, key)
    if isinstance(df_nok_bt, pd.DataFrame) and not df_nok_bt.empty:
        _df_sheet(wb, "MAJ_BT_INTROUV", df_nok_bt, key)


# ======================================================================
#  CAP_FT
# ======================================================================

def _write_analyse_sheet(wb, sheet_name, key, resultats, label_nok_qgis):
    """Shared writer for CAP_FT and COMAC analyse sheets."""
    introuvables_excel = resultats[0]
    introuvables_qgis  = resultats[1]
    existants          = resultats[2]

    ws = wb.create_sheet(sheet_name)
    hdrs = ["INF_NUM QGIS", "ETUDE QGIS", "INF_NUM EXCEL", "NOM FICHIER", "STATUT"]
    _init_sheet(ws, key, hdrs, [22, 30, 22, 45, 35])

    r = 2
    for fichier, appuis in introuvables_excel.items():
        for inf_num in appuis:
            _row(ws, r, ["", "", inf_num, fichier, "ABSENT QGIS"], fill=_P_ERR)
            r += 1
    for etude, appuis in introuvables_qgis.items():
        for inf_num in appuis:
            _row(ws, r, [inf_num, etude, "", "", label_nok_qgis], fill=_P_WARN)
            r += 1
    for _, values in existants.items():
        vals = list(values) + ["OK"]
        _row(ws, r, vals, fill=_P_OK)
        r += 1
    return ws


def write_capft(wb, result):
    """Write CAP_FT analysis sheet."""
    resultats = result.get('resultats')
    if not resultats:
        return
    ws_analyse = _write_analyse_sheet(wb, "CAPFT_ANALYSE", 'capft', resultats,
                         "ABSENT FICHES APPUIS")

    # Ajouter section HORS PERIMETRE (poteaux existant dans zone SRO mais hors zones etude)
    hors_perimetre = resultats[3] if len(resultats) > 3 else result.get('dico_hors_perimetre', {})
    if hors_perimetre and ws_analyse:
        sro_val = result.get('fddcpi_sro', '')
        hp_label = f"HORS PERIMETRE - existe dans la zone {sro_val} mais hors zone etude CAP_FT" if sro_val else "HORS PERIMETRE - existe dans la zone SRO mais hors zone etude CAP_FT"
        r = ws_analyse.max_row + 1
        for fichier, appuis in hors_perimetre.items():
            for inf_num in appuis:
                _row(ws_analyse, r, ["", "", inf_num, fichier, hp_label],
                     fill=_P_INFO)
                r += 1


# ======================================================================
#  COMAC
# ======================================================================

def write_comac(wb, result, report_options):
    """Write COMAC analysis sheets."""
    report_options = _resolve_report_options(report_options)
    resultats = result.get('resultats')
    if not resultats:
        return True

    key = 'comac'

    introuvables_excel = resultats[0]
    introuvables_qgis  = resultats[1]
    existants          = resultats[2]
    hors_perimetre     = resultats[3] if len(resultats) > 3 else result.get('dico_hors_perimetre', {})
    spatial_match      = resultats[4] if len(resultats) > 4 else result.get('dico_spatial_match', {})
    non_resolu         = resultats[5] if len(resultats) > 5 else result.get('dico_non_resolu', {})

    ws_analyse = wb.create_sheet("COMAC_ANALYSE")
    hdrs = ["INF_NUM QGIS", "ETUDE QGIS", "INF_NUM EXCEL", "NOM FICHIER", "STATUT", "EXPLICATION"]
    _init_sheet(ws_analyse, key, hdrs, [22, 30, 22, 45, 22, 65])

    def _row6(r, vals6, fill):
        _row(ws_analyse, r, vals6, fill=fill)
        ws_analyse.cell(row=r, column=6).alignment = _AL_W

    r = 2
    for _, values in existants.items():
        v = list(values)[:4]
        _row6(r, v + [
            "OK (nom)",
            "Correspondance directe par numero de poteau."
        ], fill=_P_OK)
        r += 1

    for _, m in spatial_match.items():
        dist = round(m.get('distance_m', 0), 2)
        _row6(r, [
            m.get('inf_num_qgis', ''), '',
            m.get('inf_num_excel', ''), m.get('fichier', ''),
            f"OK (spatial {dist}m)",
            f"Ambiguite commune levee par proximite geographique ({dist}m, tolerance 7.5m)."
        ], fill=_P_OK)
        r += 1

    for fichier, appuis in non_resolu.items():
        for inf_num in appuis:
            _row6(r, ["", "", inf_num, fichier,
                "AMBIGU COMMUNE",
                "Le code poteau est present dans plusieurs communes de la couche QGIS. "
                "Le matching spatial n'a pas leve l'ambiguite (coordonnees manquantes ou distance > 7.5m). "
                "Verifier la commune du poteau dans infra_pt_pot."
            ], fill=_P_WARN)
            r += 1

    for fichier, appuis in introuvables_excel.items():
        for inf_num in appuis:
            _row6(r, ["", "", inf_num, fichier,
                "ABSENT SRO",
                "Ce poteau est present dans le fichier COMAC mais absent de la couche infra_pt_pot du SRO. "
                "Verifier si le poteau a bien ete cree en base de donnees."
            ], fill=_P_ERR)
            r += 1

    for etude, appuis in introuvables_qgis.items():
        for inf_num in appuis:
            _row6(r, [inf_num, etude, "", "",
                "NON COUVERT",
                "Ce poteau BT est present dans la couche QGIS mais n'est reference dans aucun fichier COMAC. "
                "Verifier si une etude couvrant ce poteau est manquante ou si ce poteau est hors perimetre d'intervention."
            ], fill=_P_WARN)
            r += 1

    for fichier, appuis in hors_perimetre.items():
        for inf_num in appuis:
            _row6(r, ["", "", inf_num, fichier,
                "HORS ZONE ETUDE",
                "Ce poteau existe dans la couche SRO mais se situe hors des polygones etude_comac. "
                "Normal si le perimetre d'etude ne couvre pas ce secteur geographique."
            ], fill=_P_INFO)
            r += 1

    dico_verif_secu = result.get('dico_verif_secu')
    if dico_verif_secu:
        ws = wb.create_sheet("COMAC_SECURITE")
        hdrs = ["FICHIER", "POTEAU", "PORTEE (m)", "CAPACITE FO", "TYPE LIGNE",
                "PORTEE MAX (m)", "DEPASSEMENT (m)", "HAUTEUR SOL (m)",
                "VERIF PORTEE", "VERIF HAUTEUR"]
        _init_sheet(ws, key, hdrs, [40, 20, 14, 14, 15, 14, 16, 16, 28, 28])
        r = 2
        for fichier, liste_v in dico_verif_secu.items():
            for v in liste_v:
                portee = v.get('portee', 0)
                capacite = v.get('capacite_fo', 0)
                hauteur = v.get('hauteur_sol', 0)
                if portee == 0 and capacite == 0 and hauteur == 0:
                    continue
                vp = v.get('verif_portee')
                p_max = vp.get('portee_max', 0) if vp else 0
                dep = vp.get('depassement', 0) if vp else 0
                p_ok = vp.get('valide', True) if vp else True
                vh = v.get('verif_hauteur_sol')
                h_ok = vh.get('valide', True) if vh else True
                st_p = "OK" if p_ok else f"DEPASSEMENT (+{dep:.1f}m)"
                st_h = ""
                if hauteur > 0:
                    st_h = "OK" if h_ok else f"INSUFFISANT ({hauteur:.1f}m)"
                vals = [fichier, v.get('poteau', ''), portee, capacite,
                        v.get('type_ligne_fo', ''), p_max, dep,
                        hauteur if hauteur > 0 else '',
                        st_p if portee > 0 else '', st_h]
                _row(ws, r, vals)
                if portee > 0:
                    ws.cell(row=r, column=9).fill = _P_OK if p_ok else _P_CRIT
                if hauteur > 0:
                    ws.cell(row=r, column=10).fill = _P_OK if h_ok else _P_CRIT
                r += 1

    verif_cables = result.get('verif_cables')
    if verif_cables:
        ws = wb.create_sheet("COMAC_CABLES")
        has_boitier = any(e.get('boitier_comac') for e in verif_cables)
        hdrs = ["APPUI", "NB COMAC", "REFS COMAC", "CAPAS COMAC",
                "NB BDD", "CAPAS BDD", "STATUT", "MESSAGE"]
        widths = [20, 12, 40, 25, 12, 25, 14, 55]
        if has_boitier:
            hdrs.extend(["BOITIER COMAC", "BPE TYPE", "BOITIER STATUT"])
            widths.extend([15, 15, 15])
        _init_sheet(ws, key, hdrs, widths)
        r = 2
        for e in verif_cables:
            st = e.get('statut', '')
            vals = [
                e.get('num_appui', ''), e.get('nb_cables_comac', 0),
                e.get('cables_comac', ''),
                '+'.join(e.get('capas_comac', [])),
                e.get('nb_cables_bdd', 0),
                '+'.join(str(c) for c in e.get('capas_bdd', [])),
                st, e.get('message', '')]
            if has_boitier:
                vals.extend([
                    e.get('boitier_comac', ''),
                    e.get('bpe_noe_type', ''),
                    e.get('boitier_statut', '')])
            fill = _P_OK if st == 'OK' else (_P_CRIT if st == 'ECART' else (
                _P_WARN if st.startswith('ABSENT') else None))
            _row(ws, r, vals, fill=fill)
            if has_boitier:
                bst = e.get('boitier_statut', '')
                if bst == 'ERREUR':
                    for c in range(9, 12):
                        ws.cell(row=r, column=c).fill = _P_CRIT
                elif bst == 'OK':
                    for c in range(9, 12):
                        ws.cell(row=r, column=c).fill = _P_OK
            r += 1

    verif_portees = result.get('verif_portees')
    if verif_portees:
        ws = wb.create_sheet("COMAC_PORTEES")
        has_gracethd = any(e.get('source_ref') == 'GraceTHD' for e in verif_portees)
        hdrs = ["ETUDE", "CABLE", "CAPA FO",
                "DEPART PCM", "ARRIVEE PCM", "PORTEE PCM (m)",
                "DEPART REF", "ARRIVEE REF", "PORTEE REF (m)",
                "ECART (m)", "ECART (%)", "STATUT", "MESSAGE"]
        widths = [25, 18, 10, 18, 18, 14, 18, 18, 14, 12, 12, 14, 55]
        if has_gracethd:
            hdrs.insert(-2, "CONFIANCE")
            widths.insert(-2, 10)
        _init_sheet(ws, key, hdrs, widths)
        r = 2
        for e in verif_portees:
            st = e.get('statut', '')
            portee_pcm = e.get('portee_pcm', 0)
            portee_ref = e.get('portee_ref', 0)
            ecart_m = e.get('ecart_m', 0)
            ecart_pct = e.get('ecart_pct', 0)
            vals = [
                e.get('etude', ''), e.get('cable', ''),
                e.get('capacite_fo', 0),
                e.get('support_depart_pcm', ''),
                e.get('support_arrivee_pcm', ''),
                portee_pcm if portee_pcm else '',
                e.get('support_depart_ref', ''),
                e.get('support_arrivee_ref', ''),
                portee_ref if portee_ref else '',
                ecart_m if ecart_m else '',
                ecart_pct if ecart_pct else '',
            ]
            if has_gracethd:
                conf = e.get('confiance_ref', 0)
                vals.append(conf if conf else '')
            vals.extend([st, e.get('message', '')])
            fill = (_P_OK if st == 'OK'
                    else _P_CRIT if st == 'ECART'
                    else _P_WARN if st.startswith('ABSENT')
                    else None)
            _row(ws, r, vals, fill=fill)
            r += 1

    # --- Feuille PCM_VS_BDD ---
    _write_pcm_vs_bdd_sheet(wb, result)
    return True


def _write_pcm_vs_bdd_sheet(wb, result):
    """Write PCM vs BDD comparison sheet if data available."""
    pcm_data = result.get('pcm_vs_bdd')
    mecanique = result.get('pcm_vs_bdd_mecanique')
    if not pcm_data and not mecanique:
        return
    key = 'comac'

    # Feuille SUPPORTS
    if pcm_data:
        ws = wb.create_sheet("PCM_SUPPORTS")
        hdrs = ["ETUDE", "NOM PCM", "INF_NUM BDD", "NOE_CODEXT BDD",
                "METHODE", "TYPE PCM", "TYPE BDD", "TYPE OK",
                "ECART COORD (m)", "ETAT PCM", "ETAT BDD", "STATUT"]
        widths = [25, 18, 22, 22, 14, 10, 10, 10, 16, 16, 16, 18]
        _init_sheet(ws, key, hdrs, widths)
        r = 2
        for comp in pcm_data.etudes:
            for m in comp.supports_ok:
                _row(ws, r, [comp.num_etude, m.nom_pcm, m.inf_num_bdd, m.noe_codext_bdd,
                             m.match_method, m.type_pcm, m.type_bdd, 'OK',
                             round(m.ecart_coord_m, 1) if m.ecart_coord_m >= 0 else '',
                             m.etat_pcm, m.etat_bdd, 'OK'], fill=_P_OK)
                r += 1
            for m in comp.supports_type_ko:
                _row(ws, r, [comp.num_etude, m.nom_pcm, m.inf_num_bdd, m.noe_codext_bdd,
                             m.match_method, m.type_pcm, m.type_bdd, 'KO',
                             round(m.ecart_coord_m, 1) if m.ecart_coord_m >= 0 else '',
                             m.etat_pcm, m.etat_bdd, 'TYPE_KO'], fill=_P_CRIT)
                r += 1
            for m in comp.supports_coord_ko:
                _row(ws, r, [comp.num_etude, m.nom_pcm, m.inf_num_bdd, m.noe_codext_bdd,
                             m.match_method, m.type_pcm, m.type_bdd,
                             'OK' if m.type_coherent else 'KO',
                             round(m.ecart_coord_m, 1) if m.ecart_coord_m >= 0 else '',
                             m.etat_pcm, m.etat_bdd, 'COORD_KO'], fill=_P_WARN)
                r += 1
            for m in comp.supports_absents_bdd:
                _row(ws, r, [comp.num_etude, m.nom_pcm, '', '',
                             '', m.type_pcm, '', '',
                             '', m.etat_pcm, '', 'ABSENT_BDD'], fill=_P_ERR)
                r += 1

    # Feuille CABLES
    has_cables = bool(pcm_data) and any(
        comp.cables_ok or comp.cables_absents_bdd
        or comp.cables_capacite_ko or comp.cables_portee_ko
        for comp in pcm_data.etudes
    )
    if has_cables:
        ws_c = wb.create_sheet("PCM_CABLES")
        hdrs_c = ["ETUDE", "CABLE PCM", "CAPA PCM", "CAPA BDD", "CAPA OK",
                   "NB SEG PCM", "NB SEG BDD", "ECARTS PORTEES", "STATUT"]
        widths_c = [25, 18, 10, 10, 10, 12, 12, 45, 18]
        _init_sheet(ws_c, key, hdrs_c, widths_c)
        r_c = 2
        for comp in pcm_data.etudes:
            for cm in comp.cables_ok:
                _row(ws_c, r_c, [comp.num_etude, cm.cable_pcm, cm.capacite_pcm,
                                  cm.capacite_bdd, 'OK', cm.nb_segments_pcm,
                                  cm.nb_segments_bdd, '', 'OK'], fill=_P_OK)
                r_c += 1
            for cm in comp.cables_capacite_ko:
                _row(ws_c, r_c, [comp.num_etude, cm.cable_pcm, cm.capacite_pcm,
                                  cm.capacite_bdd, 'KO', cm.nb_segments_pcm,
                                  cm.nb_segments_bdd, '', 'CAPA_KO'], fill=_P_CRIT)
                r_c += 1
            for cm in comp.cables_portee_ko:
                ecarts_str = '; '.join(
                    f"{e['portee_pcm']}m vs {e['portee_bdd']}m ({e['ecart_pct']}%)"
                    for e in cm.ecarts_portees
                )
                _row(ws_c, r_c, [comp.num_etude, cm.cable_pcm, cm.capacite_pcm,
                                  cm.capacite_bdd,
                                  'OK' if cm.capacite_coherente else 'KO',
                                  cm.nb_segments_pcm, cm.nb_segments_bdd,
                                  ecarts_str, 'PORTEE_KO'], fill=_P_WARN)
                r_c += 1
            for cm in comp.cables_absents_bdd:
                _row(ws_c, r_c, [comp.num_etude, cm.cable_pcm, cm.capacite_pcm,
                                  '', '', cm.nb_segments_pcm, 0,
                                  '', 'ABSENT_BDD'], fill=_P_ERR)
                r_c += 1

    # Feuille MECANIQUE
    if mecanique:
        has_any = any(
            v['validation'].hypotheses_inconnues
            or v['validation'].armements_inconnus
            or v['validation'].cables_inconnus
            or v['validation'].supports_catalogue_ko
            for v in mecanique
        )
        if has_any:
            ws_m = wb.create_sheet("PCM_MECANIQUE")
            hdrs_m = ["ETUDE", "TYPE", "DETAIL", "VALEUR"]
            widths_m = [25, 20, 40, 30]
            _init_sheet(ws_m, key, hdrs_m, widths_m)
            r_m = 2
            for v in mecanique:
                etude_name = v['etude']
                val = v['validation']
                for h in val.hypotheses_inconnues:
                    _row(ws_m, r_m, [etude_name, 'HYPOTHESE INCONNUE', h, ''],
                         fill=_P_WARN)
                    r_m += 1
                for a in val.armements_inconnus:
                    _row(ws_m, r_m, [etude_name, 'ARMEMENT INCONNU',
                         a.get('nom_armement', ''),
                         f"Support: {a.get('support', '')}, Conducteur: {a.get('conducteur', '')}"],
                         fill=_P_WARN)
                    r_m += 1
                for c in val.cables_inconnus:
                    _row(ws_m, r_m, [etude_name, 'CABLE INCONNU', c, ''],
                         fill=_P_WARN)
                    r_m += 1
                for s in val.supports_catalogue_ko:
                    _row(ws_m, r_m, [etude_name, 'SUPPORT HORS CATALOGUE',
                         s.get('nom_pcm', ''),
                         f"H={s.get('hauteur', 0)} {s.get('classe', '')} E={s.get('effort', 0)}"],
                         fill=_P_INFO)
                    r_m += 1


# ======================================================================
#  C6 vs BD
# ======================================================================

def write_c6bd(wb, result):
    """Write C6 vs BD analysis sheets."""
    final_df = result.get('final_df')
    poteaux_out = result.get('df_poteaux_out')
    verif_etudes = result.get('verif_etudes')
    key = 'c6bd'

    if final_df is not None and not final_df.empty:
        ws = _df_sheet(wb, "C6BD_ANALYSE", final_df, key)
        for col in range(1, len(final_df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        if "Statut" in final_df.columns:
            ncols = len(final_df.columns)
            for idx, row_t in enumerate(final_df.itertuples(), start=2):
                statut = str(getattr(row_t, 'Statut', ''))
                if "ABSENT" in statut:
                    _fill_row(ws, idx, _P_WARN, ncols)

    if poteaux_out is not None and not poteaux_out.empty:
        ws_out = _df_sheet(wb, "C6BD_HORS_PERIM", poteaux_out, key)
        ncols = len(poteaux_out.columns)
        fill = _P_INFO if result.get('be_type') == 'axione' else _P_CRIT
        for r in range(2, len(poteaux_out) + 2):
            _fill_row(ws_out, r, fill, ncols)

    if verif_etudes:
        sans_c6 = verif_etudes.get('etudes_sans_c6', [])
        c6_sans = verif_etudes.get('c6_sans_etude', [])
        ml = max(len(sans_c6), len(c6_sans), 1)
        df_v = pd.DataFrame({
            'ETUDES CAP FT SANS C6': sans_c6 + [''] * (ml - len(sans_c6)),
            'FICHIERS C6 SANS ETUDE': c6_sans + [''] * (ml - len(c6_sans)),
        })
        ws_v = _df_sheet(wb, "C6BD_VERIF_ETUDES", df_v, key)
        ws_v.column_dimensions['A'].width = 40
        ws_v.column_dimensions['B'].width = 40
        for r in range(2, ml + 2):
            for c in [1, 2]:
                if ws_v.cell(row=r, column=c).value:
                    ws_v.cell(row=r, column=c).fill = _P_WARN


# ======================================================================
#  Police C6
# ======================================================================

def write_police(wb, result):
    """Write Police C6 sheets: recap + detail per appui."""
    stats = result.get('stats', [])
    if not stats:
        return

    key = 'police_c6'

    # Recap
    ws = wb.create_sheet("PLC6_RECAP")
    hdrs = ["ETUDE", "APPUIS C6", "OK", "ECARTS", "ABSENTS BDD",
            "BOITIER ERR", "% CONFORMITE"]
    _init_sheet(ws, key, hdrs, [35, 14, 10, 10, 14, 14, 16])
    for r, s in enumerate(stats, 2):
        total = s.get('appuis_c6', 0)
        ok = s.get('nb_ok', 0)
        nok = s.get('nb_ecart', 0) + s.get('nb_absent', 0) + s.get('nb_boitier_err', 0)
        pct = (ok / total * 100) if total > 0 else 100
        _row(ws, r, [s['etude'], total, ok, s.get('nb_ecart', 0),
                      s.get('nb_absent', 0), s.get('nb_boitier_err', 0),
                      f"{pct:.0f}%"])
        for c in [2, 3, 4, 5, 6, 7]:
            ws.cell(row=r, column=c).alignment = _AL_C
        fill = _P_OK if nok == 0 else (_P_WARN if pct >= 80 else _P_ERR)
        ws.cell(row=r, column=7).fill = fill
        if nok > 0:
            _fill_row(ws, r, _P_ERR, 7)
            ws.cell(row=r, column=7).fill = fill

    # Detail
    ws_d = wb.create_sheet("PLC6_DETAIL")
    hdrs_d = ["ETUDE", "N APPUI", "CABLES C6", "NB C6", "NB BDD",
              "CAPA C6", "CAPA BDD", "STATUT", "DETAIL",
              "BOITIER C6", "BPE TYPE", "BOITIER STATUT"]
    _init_sheet(ws_d, key, hdrs_d,
                [35, 15, 40, 10, 10, 18, 18, 14, 50, 12, 15, 15])

    dr = 2
    for s in stats:
        etude = s['etude']
        for a in s.get('detail', []):
            capas_c6 = a.get('capas_c6', [])
            capas_bdd = a.get('capas_bdd', [])
            vals = [etude, a['num_appui'], a.get('cables_c6', ''),
                    a['nb_cables_c6'], a['nb_cables_bdd'],
                    '+'.join(str(c) for c in capas_c6) if capas_c6 else '',
                    '+'.join(str(c) for c in capas_bdd) if capas_bdd else '',
                    a['statut'], a.get('message', ''),
                    a.get('boitier_c6', ''), a.get('bpe_noe_type', ''),
                    a.get('boitier_statut', '')]
            st = a['statut']
            fill = _P_OK if st == 'OK' else (_P_WARN if st.startswith('ABSENT') else _P_ERR)
            _row(ws_d, dr, vals, fill=fill)

            bst = a.get('boitier_statut', '')
            if bst == 'ERREUR':
                for c in range(10, 13):
                    ws_d.cell(row=dr, column=c).fill = _P_CRIT
            elif bst == 'OK':
                for c in range(10, 13):
                    ws_d.cell(row=dr, column=c).fill = _P_OK
            dr += 1


# ======================================================================
#  C6-C3A
# ======================================================================

def _c6c3a_highlight(ws, df, check_cols):
    """Highlight rows containing ABSENT in any of the check columns."""
    ncols = len(df.columns)
    for idx in range(len(df)):
        row_data = df.iloc[idx]
        if any(col in row_data and row_data[col] == "ABSENT" for col in check_cols):
            _fill_row(ws, idx + 2, _P_WARN, ncols)


def write_c6c3a(wb, result):
    """Write C6-C3A analysis sheets."""
    key = 'c6c3a'

    df_final = result.get('df_final')
    if df_final is not None and not df_final.empty:
        ws = _df_sheet(wb, "C6C3A_ANALYSE", df_final, key)
        for c in range(1, len(df_final.columns) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 25
        _c6c3a_highlight(ws, df_final,
                         ["inf_num (ETUDES_QGIS)", "inf_num (C3A)", "Excel (C6)"])

    df_rempl = result.get('df_final_rempl')
    if df_rempl is not None and not df_rempl.empty:
        ws2 = _df_sheet(wb, "C6C3A_REMPL", df_rempl, key)
        for c in range(1, len(df_rempl.columns) + 1):
            ws2.column_dimensions[get_column_letter(c)].width = 25
        _c6c3a_highlight(ws2, df_rempl,
                         ["inf_num (ETUDES_QGIS)", "inf_num (C3A)",
                          "Excel (C6)", "Fichier (C7) Excel"])


# ======================================================================
#  Main entry point
# ======================================================================

def write_gespot_c6(wb: openpyxl.Workbook, result: dict) -> None:
    """Écrit les feuilles GESPOT vs C6 dans le rapport unifié."""
    comparisons = result.get('comparisons', [])
    absent_c6 = result.get('absent_c6', [])
    absent_gespot = result.get('absent_gespot', [])
    anomalies = result.get('anomalies', [])

    _write_gespot_analyse(wb, comparisons)
    _write_gespot_absent_c6(wb, absent_c6)
    _write_gespot_absent_gespot(wb, absent_gespot)
    _write_gespot_anomalies(wb, anomalies)


def _write_gespot_analyse(wb, comparisons):
    ws = wb.create_sheet('GESPOT_ANALYSE')
    headers = [
        'NUM', 'VOIE', 'ADRESSE_C6', 'STATUT_ADRESSE', 'DETAIL_ADRESSE',
        'CENTRE_GESPOT', 'CENTRE_C6', 'STATUT_CENTRE', 'DETAIL_CENTRE',
        'TYPE_GESPOT', 'TYPE_C6', 'STATUT_TYPE', 'DETAIL_TYPE',
        'STRAT_GESPOT', 'STRAT_C6', 'STATUT_STRAT', 'DETAIL_STRAT',
        'ENV_GESPOT', 'ENV_C6', 'STATUT_ENV', 'DETAIL_ENV',
        'ELEC_GESPOT', 'ELEC_C6', 'STATUT_ELEC', 'DETAIL_ELEC',
        'INACC_GESPOT', 'INACC_C6', 'STATUT_INACC', 'DETAIL_INACC',
        'RECALAGE', 'VERTIC_C6', 'STATUT_VERTIC', 'DETAIL_VERTIC',
        'YELLOW_GESPOT', 'YELLOW_C6', 'STATUT_YELLOW', 'DETAIL_YELLOW',
        'USABLE_GESPOT', 'USABLE_C6', 'STATUT_USABLE', 'DETAIL_USABLE',
        'CTRL_VIS_GESPOT', 'CTRL_VIS_C6', 'STATUT_CTRL_VIS', 'DETAIL_CTRL_VIS',
        'NB_KO', 'STATUT_GLOBAL', 'SOURCE_GESPOT', 'SOURCE_C6',
    ]
    statut_cols = {i + 1 for i, h in enumerate(headers) if h.startswith('STATUT_')}
    fills = {'OK': _P_OK, 'KO': _P_ERR}

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = _P_HEAD
        c.font = _F_HEAD

    nb_ko_col = headers.index('NB_KO') + 1

    for row_idx, cmp in enumerate(comparisons, 2):
        values = [
            cmp.num, cmp.voie_gespot, cmp.adresse_c6, cmp.statut_adresse, cmp.detail_adresse,
            cmp.centre_gespot, cmp.centre_c6, cmp.statut_centre, cmp.detail_centre,
            cmp.type_gespot, cmp.type_c6, cmp.statut_type, cmp.detail_type,
            cmp.strat_gespot, cmp.strat_c6, cmp.statut_strat, cmp.detail_strat,
            cmp.env_gespot, cmp.env_c6, cmp.statut_env, cmp.detail_env,
            cmp.elec_gespot, cmp.elec_c6, cmp.statut_elec, cmp.detail_elec,
            cmp.inacc_gespot, cmp.inacc_c6, cmp.statut_inacc, cmp.detail_inacc,
            cmp.recalage_gespot, cmp.vertic_c6, cmp.statut_vertic, cmp.detail_vertic,
            cmp.yellow_gespot, cmp.yellow_c6, cmp.statut_yellow, cmp.detail_yellow,
            cmp.usable_gespot, cmp.usable_c6, cmp.statut_usable, cmp.detail_usable,
            cmp.ctrl_vis_gespot, cmp.ctrl_vis_c6, cmp.statut_ctrl_vis, cmp.detail_ctrl_vis,
            cmp.nb_ecarts, cmp.statut_global, cmp.source_gespot, cmp.source_c6,
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = _F_DATA
            if col_idx in statut_cols:
                c.fill = fills.get(str(val), _P_ERR)
        nb_cell = ws.cell(row=row_idx, column=nb_ko_col)
        nb_cell.fill = _P_ERR if cmp.nb_ecarts > 0 else _P_OK
    ws.freeze_panes = 'A2'


def _write_gespot_absent_c6(wb, records):
    ws = wb.create_sheet('GESPOT_ABSENT_C6')
    headers = ['NUM', 'VOIE', 'CENTRE', 'TYPE_GESPOT', 'STRAT_GESPOT',
               'ENV_GESPOT', 'ELEC_GESPOT', 'INACC_GESPOT', 'SOURCE_GESPOT']
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = _P_HEAD
        c.font = _F_HEAD
    for row_idx, r in enumerate(records, 2):
        row = [r.num, r.voie, r.centre, r.type_calc, r.strategie_calc,
               r.milieu_calc, r.pres_elect_calc, r.inacc_calc, r.source_file]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _F_DATA
    ws.freeze_panes = 'A2'


def _write_gespot_absent_gespot(wb, records):
    ws = wb.create_sheet('GESPOT_ABSENT_GESPOT')
    headers = ['NUM', 'ADRESSE_C6', 'CENTRE_C6', 'TYPE_C6', 'STRAT_C6',
               'ENV_C6', 'ELEC_C6', 'INACC_C6', 'SOURCE_C6']
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = _P_HEAD
        c.font = _F_HEAD
    for row_idx, r in enumerate(records, 2):
        row = [r.num, r.adresse, r.centre, r.type_c6, r.strat,
               r.env, r.elec, r.inacc, r.source_file]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _F_DATA
    ws.freeze_panes = 'A2'


def _write_gespot_anomalies(wb, anomalies):
    ws = wb.create_sheet('GESPOT_ANOMALIES')
    headers = ['SOURCE', 'FICHIER', 'NUM', 'TYPE_ANOMALIE', 'DETAIL', 'ACTION']
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = _P_HEAD
        c.font = _F_HEAD
    for row_idx, a in enumerate(anomalies, 2):
        row = [a.get('source', ''), a.get('fichier', ''), a.get('num', ''),
               a.get('type', ''), a.get('detail', ''), a.get('action', '')]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).font = _F_DATA
    ws.freeze_panes = 'A2'


_WRITERS = {
    'maj': write_maj,
    'capft': write_capft,
    'gespot_c6': write_gespot_c6,
    'comac': write_comac,
    'c6bd': write_c6bd,
    'police_c6': write_police,
    'c6c3a': write_c6c3a,
}


def _load_dictionary():
    """Load data_dictionary.json from the plugin directory."""
    dict_path = os.path.join(os.path.dirname(__file__), 'data_dictionary.json')
    with open(dict_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _write_data_dictionary(wb):
    """Write GLOSSAIRE + DICTIONNAIRE sheets from data_dictionary.json."""
    data = _load_dictionary()
    _write_glossary_sheet(wb, data)
    _write_dictionary_sheet(wb, data)


def _write_glossary_sheet(wb, data):
    """Feuille GLOSSAIRE: termes metier + codes couleur."""
    ws = wb.create_sheet("GLOSSAIRE")
    ws.sheet_properties.tabColor = '555555'

    ws.merge_cells('A1:B1')
    t = ws.cell(row=1, column=1, value="GLOSSAIRE — POLE AERIEN")
    t.font = _F_TITLE
    t.alignment = _AL_C

    hdrs = ["TERME", "DEFINITION"]
    for c, (h, w) in enumerate(zip(hdrs, [22, 90]), 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = _F_HEAD
        cell.fill = _P_HEAD
        cell.alignment = _AL_C
        cell.border = _BRD
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 4
    for entry in data.get('glossary', []):
        ws.cell(row=row, column=1, value=entry['term']).font = _F_BOLD
        ws.cell(row=row, column=1).border = _BRD
        ws.cell(row=row, column=1).alignment = _AL_L
        c2 = ws.cell(row=row, column=2, value=entry['definition'])
        c2.font = _F_DATA
        c2.border = _BRD
        c2.alignment = _AL_W
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="CODES COULEUR").font = _F_BOLD
    row += 1
    for cc in data.get('color_codes', []):
        c1 = ws.cell(row=row, column=1, value=cc.get('usage', ''))
        c1.font = _F_DATA
        c1.border = _BRD
        c1.fill = PatternFill('solid', fgColor=cc['hex'].lstrip('#'))
        c2 = ws.cell(row=row, column=2, value=cc.get('meaning', ''))
        c2.font = _F_DATA
        c2.border = _BRD
        c2.alignment = _AL_W
        row += 1

    ws.freeze_panes = 'A4'


def _write_dictionary_sheet(wb, data):
    """Feuille DICTIONNAIRE: colonnes de chaque feuille, types, valeurs possibles."""
    ws = wb.create_sheet("DICTIONNAIRE")
    ws.sheet_properties.tabColor = '555555'

    ws.merge_cells('A1:G1')
    t = ws.cell(row=1, column=1, value="DICTIONNAIRE DE DONNEES — ISO/IEC 11179")
    t.font = _F_TITLE
    t.alignment = _AL_C

    meta = data.get('metadata', {})
    ws.cell(row=2, column=1, value=f"Version {meta.get('version', '')} | {meta.get('organization', '')}").font = Font(
        name='Calibri', size=9, italic=True, color='555555')

    hdrs = ["MODULE", "FEUILLE", "COLONNE", "TYPE", "DEFINITION", "VALEURS POSSIBLES", "COULEUR"]
    widths = [18, 22, 22, 10, 55, 55, 12]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = _F_HEAD
        cell.fill = _P_HEAD
        cell.alignment = _AL_C
        cell.border = _BRD
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 5
    for module in data.get('modules', []):
        mod_name = module.get('module_name', '')
        for sheet in module.get('sheets', []):
            sheet_name = sheet.get('sheet_name', '')
            for elem in sheet.get('data_elements', []):
                vals_text = ''
                color_text = ''
                vd = elem.get('value_domain')
                if vd and vd.get('type') == 'enumerated':
                    parts = []
                    colors = []
                    for v in vd.get('values', []):
                        val = v.get('value', '')
                        defn = v.get('definition', '')
                        parts.append(f"{val} : {defn}" if defn else val)
                        clr = v.get('color', '')
                        if clr:
                            colors.append(f"{val}={clr}")
                    vals_text = '\n'.join(parts)
                    color_text = ', '.join(colors) if colors else ''
                elif vd and vd.get('type') == 'range':
                    lo = vd.get('min', vd.get('minimum', ''))
                    hi = vd.get('max', vd.get('maximum', ''))
                    vals_text = f"{lo} a {hi}" if lo is not None and hi is not None else ''
                elif vd and vd.get('type') == 'mixed':
                    sv = vd.get('special_values', [])
                    parts = [vd.get('description', '')]
                    for v in sv:
                        parts.append(f"{v.get('value', '')} : {v.get('definition', '')}")
                    vals_text = '\n'.join(p for p in parts if p)
                elif vd and vd.get('type') == 'description':
                    rules = vd.get('rules', [])
                    parts = []
                    for r in rules:
                        parts.append(f"{r.get('condition', '')} : {r.get('meaning', '')}")
                    vals_text = '\n'.join(parts)

                unit = elem.get('unit', '')
                dtype = elem.get('data_type', '')
                if unit:
                    dtype = f"{dtype} ({unit})"

                vals = [mod_name, sheet_name, elem.get('name', ''),
                        dtype, elem.get('definition', ''),
                        vals_text, color_text]
                _row(ws, row, vals)
                ws.cell(row=row, column=5).alignment = _AL_W
                ws.cell(row=row, column=6).alignment = _AL_W
                mod_name = ''
                sheet_name = ''
                row += 1

    ws.freeze_panes = 'A5'
    if row > 5:
        ws.auto_filter.ref = f'A4:G{row - 1}'


def _resolve_report_options(report_options):
    defaults = {
        'include_comac_drawings': False,
        'include_data_dictionary': False,
        'progress': None,
        'message': None,
        'is_cancelled': None,
        'drawings_progress_start': None,
        'drawings_progress_end': None,
    }
    if isinstance(report_options, dict):
        merged = dict(defaults)
        merged.update(report_options)
        merged['include_comac_drawings'] = bool(merged.get('include_comac_drawings'))
        return merged
    if report_options is None:
        return defaults
    merged = dict(defaults)
    merged['include_comac_drawings'] = bool(report_options)
    return merged


def _report_message(report_options, message, color='grey'):
    callback = report_options.get('message') if report_options else None
    if callable(callback):
        callback(message, color)
        return
    try:
        from qgis.core import QgsMessageLog, Qgis
        level = Qgis.Warning if color == 'orange' else Qgis.Info
        QgsMessageLog.logMessage(message, "PoleAerien", level)
    except Exception:
        pass


def _report_progress(report_options, value, message=None, color='grey'):
    callback = report_options.get('progress') if report_options else None
    if callable(callback):
        callback(max(0, min(100, int(value))))
    if message:
        _report_message(report_options, message, color)


def _report_cancelled(report_options):
    callback = report_options.get('is_cancelled') if report_options else None
    return bool(callable(callback) and callback())


def _report_drawings_progress(report_options, current_page, total_pages):
    start = report_options.get('drawings_progress_start')
    end = report_options.get('drawings_progress_end')
    if start is None or end is None or total_pages <= 0:
        return
    value = start + ((end - start) * current_page / total_pages)
    _report_progress(
        report_options,
        value,
        f"Rendu schémas COMAC {current_page}/{total_pages}",
    )


def _save_workbook(wb, filepath, report_options):
    try:
        wb.save(filepath)
        return filepath
    except PermissionError:
        base, ext = os.path.splitext(filepath)
        for i in range(2, 10):
            alt_path = f"{base}_{i}{ext}"
            try:
                wb.save(alt_path)
                _report_message(
                    report_options,
                    f"[REPORT] Fichier {os.path.basename(filepath)} verrouille. Sauvegarde sous {os.path.basename(alt_path)}.",
                    'orange'
                )
                return alt_path
            except OSError:
                continue
        raise


def generate_unified_report(batch_results, export_dir, report_options=None):
    """
    Generate a single Excel workbook from all module results.

    Args:
        batch_results: dict {module_key: result_dict}
        export_dir: directory to save the file

    Returns:
        str: path to the generated file
    """
    report_options = _resolve_report_options(report_options)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    sro_raw = report_options.get('sro', '') or ''
    sro_safe = sro_raw.replace('/', '_').replace('\\', '_').strip('_')
    if sro_safe:
        filename = f"rapport_{sro_safe}_{ts}.xlsx"
    else:
        filename = f"rapport_batch_{ts}.xlsx"
    filepath = os.path.join(export_dir, filename)

    _report_progress(report_options, 2, "Preparation du rapport...", 'grey')
    wb = openpyxl.Workbook()
    try:
        _write_dashboard(wb, batch_results)
        _report_progress(report_options, 8, "Tableau de bord...", 'grey')
        module_keys = [
            key for key in ['maj', 'capft', 'gespot_c6', 'comac', 'c6bd', 'police_c6', 'c6c3a']
            if batch_results.get(key) is not None
        ]
        total_modules = max(len(module_keys), 1)
        for idx, module_key in enumerate(module_keys):
            if _report_cancelled(report_options):
                return None
            result = batch_results.get(module_key)
            writer = _WRITERS.get(module_key)
            if not writer:
                continue
            range_start = 10 + (idx * 80 / total_modules)
            range_end = 10 + ((idx + 1) * 80 / total_modules)
            _report_progress(report_options, range_start, f"Feuille {module_key}...", 'grey')
            try:
                if module_key == 'comac':
                    if not writer(wb, result, report_options):
                        return None
                else:
                    writer(wb, result)
            except (KeyError, ValueError, TypeError, AttributeError, IndexError) as exc:
                _report_message(report_options, f"[REPORT] Erreur ecriture {module_key}: {exc}", 'orange')
            _report_progress(report_options, range_end)
        if _report_cancelled(report_options):
            return None
        if report_options.get('include_data_dictionary'):
            _report_progress(report_options, 91, "Dictionnaire de donnees...", 'grey')
            try:
                _write_data_dictionary(wb)
            except Exception as exc:
                _report_message(report_options, f"[REPORT] Erreur dictionnaire: {exc}", 'orange')
        if report_options.get('include_comac_drawings') and batch_results.get('comac'):
            _report_progress(report_options, 93, "Dessins COMAC...", 'grey')
            drawing_options = dict(report_options)
            drawing_options['drawings_progress_start'] = 93
            drawing_options['drawings_progress_end'] = 95
            try:
                _write_comac_drawing_sheets(wb, batch_results['comac'], drawing_options)
            except Exception as exc:
                _report_message(report_options, f"[REPORT] Erreur dessins COMAC: {exc}", 'orange')
        _report_progress(report_options, 95, "Sauvegarde du rapport...", 'grey')
        filepath = _save_workbook(wb, filepath, report_options)
        _report_progress(report_options, 100, "Rapport sauvegarde.", 'green')
        return filepath
    finally:
        try:
            wb.close()
        except Exception:
            pass
