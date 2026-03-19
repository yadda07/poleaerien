import os
import tempfile
import unittest
from types import SimpleNamespace

import openpyxl

from unified_report import generate_unified_report


def _support(name, orientation):
    return SimpleNamespace(
        nom=name,
        nature='BT',
        classe='C',
        hauteur=8.0,
        orientation=orientation,
        effort=0.0,
        a_poser=False,
        facade=False,
        portee_molle=False,
        illisible=False,
        non_calcule=False,
    )


def _etude():
    supports = {
        'A': _support('A', 0.0),
        'B': _support('B', 100.0),
    }
    lignes_bt = [
        SimpleNamespace(
            supports=['A', 'B'],
            conducteur='BT-95',
            a_poser=False,
            armements=[],
        )
    ]
    lignes_tcf = [
        SimpleNamespace(
            supports=['A', 'B'],
            cable='FO-24',
            a_poser=False,
        )
    ]
    portees_globales = [
        SimpleNamespace(support_gauche='A', support_droit='B', angle=0.0, longueur=32.0, route=False),
    ]
    return SimpleNamespace(
        supports=supports,
        lignes_bt=lignes_bt,
        lignes_tcf=lignes_tcf,
        portees_globales=portees_globales,
    )


def _batch_results():
    return {
        'comac': {
            'resultats': ({}, {}, {'A': ('A', 'ETUDE-1', 'A', 'PCM-1')}, {}, {}, {}),
            'etudes_pcm': {'ETUDE-1': _etude()},
            'erreurs_pcm': {},
        }
    }


class TestUnifiedReport(unittest.TestCase):
    def test_generate_unified_report_reports_progress_and_creates_drawings(self):
        progress = []
        messages = []
        with tempfile.TemporaryDirectory() as temp_dir:
            filepath = generate_unified_report(
                _batch_results(),
                temp_dir,
                {
                    'include_comac_drawings': True,
                    'progress': progress.append,
                    'message': lambda message, color: messages.append((message, color)),
                }
            )

            self.assertTrue(os.path.isfile(filepath))
            self.assertGreaterEqual(progress[-1], 100)
            workbook = openpyxl.load_workbook(filepath)
            try:
                self.assertIn('COMAC_ANALYSE', workbook.sheetnames)
                drawing_sheets = [s for s in workbook.sheetnames if s.startswith('DESSIN_')]
                self.assertTrue(drawing_sheets, f"No drawing sheet found in {workbook.sheetnames}")
                self.assertIn('DESSIN_ETUDE-1', workbook.sheetnames)
            finally:
                workbook.close()

    def test_generate_unified_report_stops_before_save_when_cancelled(self):
        progress = []
        state = {'cancelled': False}

        def on_progress(value):
            progress.append(value)
            if value >= 10:
                state['cancelled'] = True

        with tempfile.TemporaryDirectory() as temp_dir:
            filepath = generate_unified_report(
                _batch_results(),
                temp_dir,
                {
                    'include_comac_drawings': True,
                    'progress': on_progress,
                    'is_cancelled': lambda: state['cancelled'],
                }
            )

            self.assertIsNone(filepath)
            self.assertEqual([], os.listdir(temp_dir))
            self.assertTrue(progress)


if __name__ == '__main__':
    unittest.main()
