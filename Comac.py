#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import openpyxl
from .qgis_utils import (
    verifications_donnees_etude,
    liste_poteaux_par_etude,
    normalize_appui_num_bt
)
from .security_rules import (
    get_capacite_fo_from_code,
    verifier_portee,
    verifier_distance_sol,
    est_terrain_prive,
    EXCEL_COL_NUM_APPUI,
    EXCEL_COL_HAUTEUR_HORS_SOL,
    EXCEL_COL_CONDUCTEUR,
    EXCEL_COL_LONGUEUR_FACTURER,
    EXCEL_COL_FO_TYPE_LIGNE
)
from .pcm_parser import (
    parse_pcm_file,
    parse_repertoire_pcm,
    verifier_securite_etude,
    get_anomalies_securite,
    get_supports_portee_molle
)


class Comac:
    """ Second pour contenir quelques fonctions qui seront utilisées dans le fichier principale CheckOptyce.py """

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def verificationsDonneesComac(self, table_poteau, table_etude_comac, colonne_comac):
        """Vérifie doublons études + poteaux BT hors étude."""
        return verifications_donnees_etude(
            table_poteau, table_etude_comac, colonne_comac,
            'POT-BT', 'COMAC'
        )

    def liste_poteau_comac(self, table_poteau, table_etude_comac, colonne_comac):
        """Liste poteaux BT par étude avec détection terrains privés."""
        return liste_poteaux_par_etude(
            table_poteau, table_etude_comac, colonne_comac,
            'POT-BT', 'COMAC'
        )

    def LectureFichiersExcelsComac(self, repertoire, zone_climatique='ZVN'):
        """
        Fonction pour parcourir les fichiers Excel (.xlsx) COMAC
        et extraire les numéros de poteaux + données sécurité.
        
        Détecte automatiquement les fichiers Excel COMAC par:
        - Pattern nom: EXPORTCOMAC, Export_Comac, ou nom d'étude (NGE-*, PA-*)
        - Structure: colonne A avec numéros poteaux (format Exxxxxx ou BT-xxx)
        
        Args:
            repertoire: Chemin du répertoire contenant les fichiers Excel
            zone_climatique: 'ZVN' (vent normal) ou 'ZVF' (vent fort)
        
        Returns:
            tuple: (doublons, erreurs, dict_poteaux, dict_verif_secu)
        """
        from qgis.core import QgsMessageLog, Qgis
        
        # Validation zone climatique
        if zone_climatique not in ('ZVN', 'ZVF'):
            zone_climatique = 'ZVN'
        dicoPoteauBt_SousTraitant = {}
        dicoVerifSecu = {}  # Résultats vérifications sécurité
        fichiersComacExistants = []
        fichiersComacEnDoublons = []
        impossibiliteDelireFichier = {}
        
        # Patterns de noms de fichiers COMAC acceptés
        PATTERNS_COMAC = ['EXPORTCOMAC', 'EXPORT_COMAC', 'COMAC', 'NGE-', 'PA-']
        # Fichiers à exclure
        PATTERNS_EXCLUS = ['ANALYSE_', 'RAPPORT', 'SYNTHESE', 'RESUME', 'C6', 'C7', 'FICHEAPPUI']
        
        fichiers_trouves = 0
        fichiers_valides = 0

        for subdir, _, files in os.walk(repertoire):
            for name in files:
                if name.endswith('.xlsx') and "~$" not in name:
                    name_upper = name.upper()
                    
                    # Exclure fichiers non-COMAC
                    if any(excl in name_upper for excl in PATTERNS_EXCLUS):
                        continue
                    
                    # Vérifier si le nom correspond à un pattern COMAC
                    is_comac_pattern = any(pat in name_upper for pat in PATTERNS_COMAC)
                    
                    # Fallback: accepter tous les Excel dans dossiers d'études (NGE-*, PA-*)
                    parent_folder = os.path.basename(subdir).upper()
                    is_in_etude_folder = any(pat in parent_folder for pat in ['NGE-', 'PA-', 'B1L-', 'B1I-'])
                    
                    if is_comac_pattern or is_in_etude_folder:
                        filepath = os.path.join(subdir, name)
                        try:
                            document = openpyxl.load_workbook(filepath, data_only=True)
                        except Exception as e:
                            impossibiliteDelireFichier[filepath] = str(e)
                            continue

                        try:
                            feuille_1 = document.worksheets[0]
                        except Exception as e:
                            impossibiliteDelireFichier[filepath] = f"Feuille illisible : {e}"
                            continue

                        listePoteauBt = []
                        listeVerifSecu = []  # Vérifications par ligne

                        # Lecture à partir de la ligne 4
                        # Col A (idx 0): N° poteau, Col G (idx 6): distance cable/BT
                        # Col AO (idx 40): Type ligne FO, Col AU (idx 46): Longueur à facturer
                        for row in feuille_1.iter_rows(min_row=4, min_col=1, max_col=50, values_only=True):
                            # Validation NULL stricte (pattern Maj_Ft_Bt.py)
                            if not row or len(row) == 0:
                                continue
                            
                            numPotBt = row[0]  # Col A
                            if not numPotBt or numPotBt == '' or str(numPotBt).strip() == '':
                                continue
                            
                            nompot = str(numPotBt).replace("BT ", "BT-")
                            listePoteauBt.append(nompot)
                            
                            # Extraction données sécurité avec validation NULL explicite
                            hauteur_hors_sol_raw = row[EXCEL_COL_HAUTEUR_HORS_SOL] if len(row) > EXCEL_COL_HAUTEUR_HORS_SOL and row[EXCEL_COL_HAUTEUR_HORS_SOL] else None
                            conducteur_raw = row[EXCEL_COL_CONDUCTEUR] if len(row) > EXCEL_COL_CONDUCTEUR and row[EXCEL_COL_CONDUCTEUR] else None
                            type_ligne_fo = row[EXCEL_COL_FO_TYPE_LIGNE] if len(row) > EXCEL_COL_FO_TYPE_LIGNE and row[EXCEL_COL_FO_TYPE_LIGNE] else None
                            longueur_raw = row[EXCEL_COL_LONGUEUR_FACTURER] if len(row) > EXCEL_COL_LONGUEUR_FACTURER and row[EXCEL_COL_LONGUEUR_FACTURER] else None
                            
                            # Parse longueur (portée)
                            portee = 0.0
                            if longueur_raw:
                                try:
                                    portee = float(str(longueur_raw).replace(',', '.').replace('m', '').strip())
                                except (ValueError, AttributeError):
                                    portee = 0.0
                            
                            # Parse hauteur hors sol (distance câble/sol)
                            hauteur_sol = 0.0
                            if hauteur_hors_sol_raw:
                                try:
                                    hauteur_sol = float(str(hauteur_hors_sol_raw).replace(',', '.').replace('m', '').strip())
                                except (ValueError, AttributeError):
                                    hauteur_sol = 0.0
                            
                            # Capacité FO depuis code câble
                            capacite_fo = get_capacite_fo_from_code(type_ligne_fo) if type_ligne_fo else 0
                            
                            # Vérification portée
                            verif_portee = None
                            if portee > 0 and capacite_fo > 0:
                                verif_portee = verifier_portee(portee, capacite_fo, zone_climatique)
                            
                            # Vérification distance câble/sol (>= 4m)
                            verif_hauteur_sol = None
                            if hauteur_sol > 0:
                                verif_hauteur_sol = verifier_distance_sol(hauteur_sol)
                            
                            # Stockage résultat
                            listeVerifSecu.append({
                                'poteau': nompot,
                                'portee': portee,
                                'capacite_fo': capacite_fo,
                                'type_ligne_fo': type_ligne_fo,
                                'hauteur_sol': hauteur_sol,
                                'conducteur': conducteur_raw,
                                'verif_portee': verif_portee,
                                'verif_hauteur_sol': verif_hauteur_sol
                            })

                        fichiers_trouves += 1
                        
                        if listePoteauBt:
                            # Utiliser chemin relatif comme clé pour éviter conflits de noms
                            rel_path = os.path.relpath(filepath, repertoire)
                            etude_name = os.path.basename(subdir)  # Nom du dossier parent = nom étude
                            key = etude_name if etude_name not in dicoPoteauBt_SousTraitant else rel_path
                            
                            dicoPoteauBt_SousTraitant[key] = listePoteauBt
                            dicoVerifSecu[key] = listeVerifSecu
                            fichiers_valides += 1

                            if name in fichiersComacExistants:
                                fichiersComacEnDoublons.append(name)

                            fichiersComacExistants.append(name)
                            QgsMessageLog.logMessage(
                                f"[COMAC] Fichier détecté: {rel_path} ({len(listePoteauBt)} poteaux)",
                                "PoleAerien", Qgis.Info
                            )
        
        QgsMessageLog.logMessage(
            f"[COMAC] Lecture terminée: {fichiers_valides}/{fichiers_trouves} fichiers valides, "
            f"{sum(len(v) for v in dicoPoteauBt_SousTraitant.values())} poteaux total",
            "PoleAerien", Qgis.Info
        )

        return fichiersComacEnDoublons, impossibiliteDelireFichier, dicoPoteauBt_SousTraitant, dicoVerifSecu

    def LectureFichiersPCM(self, repertoire, zone_climatique='ZVN'):
        """
        Lecture des fichiers .pcm COMAC (XML) pour extraction complète.
        Remplace/complète la lecture Excel avec données plus fiables.
        
        Args:
            repertoire: Chemin du répertoire contenant les .pcm
            zone_climatique: 'ZVN' ou 'ZVF'
        
        Returns:
            tuple: (etudes_dict, erreurs_dict, anomalies_list, supports_pm_list)
        """
        # DEBUG COMAC PCM
        print(f"[COMAC_PCM] Début lecture répertoire: {repertoire}")
        print(f"[COMAC_PCM] Zone climatique: {zone_climatique}")
        
        etudes, erreurs = parse_repertoire_pcm(repertoire, zone_climatique)
        
        print(f"[COMAC_PCM] Nombre études parsées: {len(etudes)}")
        print(f"[COMAC_PCM] Nombre erreurs: {len(erreurs)}")
        if erreurs:
            for fpath, err in erreurs.items():
                print(f"[COMAC_PCM] ERREUR {fpath}: {err}")
        
        # Résumé capacités FO par étude
        for nom_etude, etude in etudes.items():
            lignes_fo = [l for l in etude.lignes_tcf if l.capacite_fo > 0]
            print(f"[COMAC_PCM] Etude '{nom_etude}': {len(etude.lignes_tcf)} lignes TCF, {len(lignes_fo)} avec capacité FO")
            for l in lignes_fo:
                print(f"[COMAC_PCM]   -> cable='{l.cable}' capacite_fo={l.capacite_fo} a_poser={l.a_poser} nb_portees={len(l.portees)}")
        
        anomalies = get_anomalies_securite(etudes)
        supports_pm = get_supports_portee_molle(etudes)
        
        print(f"[COMAC_PCM] Anomalies sécurité: {len(anomalies)}")
        print(f"[COMAC_PCM] Supports portée molle: {len(supports_pm)}")
        
        # Extraction poteaux par étude (compatible avec workflow existant)
        dico_poteaux = {}
        for nom_etude, etude in etudes.items():
            poteaux = list(etude.supports.keys())
            if poteaux:
                dico_poteaux[nom_etude] = poteaux
        
        print(f"[COMAC_PCM] Fin lecture - {len(dico_poteaux)} études avec poteaux")
        return etudes, erreurs, anomalies, supports_pm, dico_poteaux

    def traitementResultatFinaux(self, dicoEtudeComacPoteauQgis, dicoPoteauBt_SousTraitant):
        """Traite les résultats finaux des deux dictionnaires.
        
        Returns:
            tuple: (introuvables_excel, introuvables_qgis, existants)
        """
        dicoPotBt_Excel_Introuvable = {}
        dicoPotBtExistants = {}
        comptabilite = 0

        # Index rapide QGIS: clé normalisée -> liste (etude, inf_num_complet)
        index_qgis = {}
        for etude_comac, listePoteauxQgis in dicoEtudeComacPoteauQgis.items():
            for inf_num_bt in listePoteauxQgis:
                cle = normalize_appui_num_bt(inf_num_bt, strip_e_prefix=True)
                index_qgis.setdefault(cle, []).append((etude_comac, inf_num_bt))

        # Traitement Excel
        for excel, listePoteau in dicoPoteauBt_SousTraitant.items():
            PotBtintrouvableSt = []
            
            for poteauSt in listePoteau:
                cle_excel = normalize_appui_num_bt(poteauSt, strip_e_prefix=True)
                
                if cle_excel in index_qgis and index_qgis[cle_excel]:
                    # Correspondance trouvée - pop première occurrence
                    etude_comac, inf_num_bt = index_qgis[cle_excel].pop(0)
                    comptabilite += 1
                    dicoPotBtExistants[comptabilite] = [inf_num_bt, etude_comac, poteauSt, excel]
                    
                    # Màj dict QGIS original (copie pour éviter modif pendant itération)
                    if etude_comac in dicoEtudeComacPoteauQgis:
                        new_list = [x for x in dicoEtudeComacPoteauQgis[etude_comac] if x != inf_num_bt]
                        if new_list:
                            dicoEtudeComacPoteauQgis[etude_comac] = new_list
                        else:
                            del dicoEtudeComacPoteauQgis[etude_comac]
                    
                    # Nettoyer index si vide
                    if not index_qgis[cle_excel]:
                        del index_qgis[cle_excel]
                else:
                    PotBtintrouvableSt.append(poteauSt)

            if PotBtintrouvableSt:
                dicoPotBt_Excel_Introuvable[excel] = PotBtintrouvableSt

        return dicoPotBt_Excel_Introuvable, dicoEtudeComacPoteauQgis, dicoPotBtExistants

    def ecrireResultatsAnalyseExcels(self, resultatsFinaux, nom, dico_verif_secu=None):
        """Fonction qui permet d'écrire dans le fichier Excel à partir d'un fichier modèle.
        
        Args:
            resultatsFinaux: tuple (introuvables_excel, introuvables_qgis, existants)
            nom: Chemin du fichier Excel à créer
            dico_verif_secu: Dictionnaire des vérifications sécurité par fichier
        """

        dicoPotBt_Excel_Introuvable = resultatsFinaux[0]
        dicoEtudeComacPotQgisIntrouvable = resultatsFinaux[1]
        dicoPotBtExistants = resultatsFinaux[2]

        fichierXlsx = openpyxl.workbook.Workbook()

        ############################## FEUILLE 1 : VALIDATION_COMPLET ################################################

        my_ligne = 1  # Ligne à partir de laquelle on commencera à écrire dans le fichier Annexe
        # Parcourir fichier C3A, pour comparer ses contenus aux fichiers C6
        feuille = fichierXlsx.create_sheet("ANALYSE COMAC", 0)

        alignement = openpyxl.styles.Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False,
                                               shrink_to_fit=True, indent=0)

        # Pour mettre du GRAS à partir de la colonne 18 et 23 pour les câbles.
        feuille.cell(row=1, column=1, value="INF_NUM QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=2, value="ETUDE QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=3, value="INF_NUM EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=4, value="NOM FICHIER EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=5, value="REMARQUES").font = openpyxl.styles.Font(bold=True)

        feuille.cell(row=1, column=1).alignment = alignement
        feuille.cell(row=1, column=2).alignment = alignement
        feuille.cell(row=1, column=3).alignment = alignement
        feuille.cell(row=1, column=4).alignment = alignement
        feuille.cell(row=1, column=5).alignment = alignement

        feuille.column_dimensions["A"].width = 25
        feuille.column_dimensions["B"].width = 30
        feuille.column_dimensions["C"].width = 25
        feuille.column_dimensions["D"].width = 50
        feuille.column_dimensions["E"].width = 40

        red_color = openpyxl.styles.colors.Color(rgb='fc4e2a')
        orange_color = openpyxl.styles.colors.Color(rgb='fd8d3c')

        ####################### SOUS-TRAITS INTROUVABLE #######################################")
        for fichierExcels, listesDesAppuisSous_traitant in dicoPotBt_Excel_Introuvable.items():
            for inf_num_excel in listesDesAppuisSous_traitant:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=3, value=inf_num_excel).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=4, value=fichierExcels).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans QGIS").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)

        ####################### QGIS INTROUVABLE #######################################
        for etudeQgis, listesDesAppuisQGis in dicoEtudeComacPotQgisIntrouvable.items():
            for inf_num in listesDesAppuisQGis:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=1, value=inf_num).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=2, value=etudeQgis).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans les Excels").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)

        my_colonne = 1
        my_ligne += 1

        ####################### POTEAUX TROUVEES #######################################")
        for _, listesDesAppuisBT in dicoPotBtExistants.items():
            for valeur in listesDesAppuisBT:
                # Si tirage câble complètement validé
                feuille.cell(row=my_ligne, column=my_colonne, value=valeur)
                my_colonne += 1

            feuille.cell(row=my_ligne, column=5, value="correspondance trouvée")

            # On passe à la ligne suivante
            # Si la première liste est terminé, on passe à la ligne suivante en récommençant par au toujours au niveau de la colonne A
            my_ligne += 1
            my_colonne = 1

        ############################## FEUILLE 2 : VERIF_SECURITE ################################################
        if dico_verif_secu:
            feuille_secu = fichierXlsx.create_sheet("VERIF_SECURITE", 1)
            
            # En-têtes
            headers_secu = ["FICHIER", "POTEAU", "PORTEE (m)", "CAPACITE FO", "TYPE LIGNE", 
                           "PORTEE MAX", "DEPASSEMENT", "HAUTEUR SOL (m)", "VERIF_PORTEE", "VERIF_HAUTEUR"]
            for col_idx, header in enumerate(headers_secu, 1):
                feuille_secu.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
                feuille_secu.cell(row=1, column=col_idx).alignment = alignement
            
            # Largeurs colonnes
            feuille_secu.column_dimensions["A"].width = 40
            feuille_secu.column_dimensions["B"].width = 20
            feuille_secu.column_dimensions["C"].width = 12
            feuille_secu.column_dimensions["D"].width = 12
            feuille_secu.column_dimensions["E"].width = 15
            feuille_secu.column_dimensions["F"].width = 12
            feuille_secu.column_dimensions["G"].width = 12
            feuille_secu.column_dimensions["H"].width = 15
            feuille_secu.column_dimensions["I"].width = 30
            feuille_secu.column_dimensions["J"].width = 30
            
            green_color = openpyxl.styles.colors.Color(rgb='90EE90')
            ligne_secu = 2
            
            for fichier, liste_verif in dico_verif_secu.items():
                for verif in liste_verif:
                    portee = verif.get('portee', 0)
                    capacite = verif.get('capacite_fo', 0)
                    hauteur_sol = verif.get('hauteur_sol', 0)
                    
                    # Ignorer lignes sans données
                    if portee == 0 and capacite == 0 and hauteur_sol == 0:
                        continue
                    
                    # Vérification portée
                    verif_portee = verif.get('verif_portee')
                    portee_max = verif_portee.get('portee_max', 0) if verif_portee else 0
                    depassement = verif_portee.get('depassement', 0) if verif_portee else 0
                    portee_valide = verif_portee.get('valide', True) if verif_portee else True
                    
                    # Vérification hauteur sol (>= 4m)
                    verif_haut = verif.get('verif_hauteur_sol')
                    hauteur_valide = verif_haut.get('valide', True) if verif_haut else True
                    
                    # Statuts vérification
                    statut_portee = "OK" if portee_valide else f"PORTEE MOLLE (+{depassement:.1f}m)"
                    statut_hauteur = ""
                    if hauteur_sol > 0:
                        statut_hauteur = "OK" if hauteur_valide else f"HAUTEUR < 4m ({hauteur_sol:.1f}m)"
                    
                    feuille_secu.cell(row=ligne_secu, column=1, value=fichier)
                    feuille_secu.cell(row=ligne_secu, column=2, value=verif.get('poteau', ''))
                    feuille_secu.cell(row=ligne_secu, column=3, value=portee)
                    feuille_secu.cell(row=ligne_secu, column=4, value=capacite)
                    feuille_secu.cell(row=ligne_secu, column=5, value=verif.get('type_ligne_fo', ''))
                    feuille_secu.cell(row=ligne_secu, column=6, value=portee_max)
                    feuille_secu.cell(row=ligne_secu, column=7, value=depassement)
                    feuille_secu.cell(row=ligne_secu, column=8, value=hauteur_sol if hauteur_sol > 0 else '')
                    feuille_secu.cell(row=ligne_secu, column=9, value=statut_portee if portee > 0 else '')
                    feuille_secu.cell(row=ligne_secu, column=10, value=statut_hauteur)
                    
                    # Coloration cellule portée
                    if portee > 0:
                        color_portee = green_color if portee_valide else red_color
                        feuille_secu.cell(row=ligne_secu, column=9).fill = openpyxl.styles.PatternFill(
                            fill_type='solid', fgColor=color_portee)
                    
                    # Coloration cellule hauteur
                    if hauteur_sol > 0:
                        color_hauteur = green_color if hauteur_valide else red_color
                        feuille_secu.cell(row=ligne_secu, column=10).fill = openpyxl.styles.PatternFill(
                            fill_type='solid', fgColor=color_hauteur)
                    
                    ligne_secu += 1

        if 'Sheet' in fichierXlsx.sheetnames:
            del fichierXlsx['Sheet']

        fichierXlsx.save(filename=nom)

    def ecrireResultatsPCM(self, etudes, anomalies, supports_pm, nom_fichier):
        """
        Écrit les résultats d'analyse des fichiers .pcm dans un Excel.
        
        Args:
            etudes: Dict des études PCM parsées
            anomalies: Liste des anomalies de sécurité
            supports_pm: Liste des supports en portée molle
            nom_fichier: Chemin du fichier Excel à créer
        """
        fichierXlsx = openpyxl.workbook.Workbook()
        
        alignement = openpyxl.styles.Alignment(horizontal='center', vertical='bottom', 
                                                shrink_to_fit=True)
        red_color = openpyxl.styles.colors.Color(rgb='fc4e2a')
        orange_color = openpyxl.styles.colors.Color(rgb='fd8d3c')
        green_color = openpyxl.styles.colors.Color(rgb='90EE90')
        
        # === FEUILLE 1: RESUME ETUDES ===
        feuille_resume = fichierXlsx.create_sheet("RESUME_ETUDES", 0)
        headers_resume = ["ETUDE", "COMMUNE", "NB_SUPPORTS", "NB_LIGNES_FO", 
                         "NB_ANOMALIES", "HYPOTHESES"]
        for col, h in enumerate(headers_resume, 1):
            feuille_resume.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
        
        ligne = 2
        for nom_etude, etude in etudes.items():
            nb_anomalies = len([a for a in anomalies if a.get('etude') == nom_etude])
            nb_lignes_fo = len([l for l in etude.lignes_tcf if l.capacite_fo > 0])
            
            feuille_resume.cell(row=ligne, column=1, value=nom_etude)
            feuille_resume.cell(row=ligne, column=2, value=etude.commune)
            feuille_resume.cell(row=ligne, column=3, value=len(etude.supports))
            feuille_resume.cell(row=ligne, column=4, value=nb_lignes_fo)
            feuille_resume.cell(row=ligne, column=5, value=nb_anomalies)
            feuille_resume.cell(row=ligne, column=6, value=', '.join(etude.hypotheses))
            
            if nb_anomalies > 0:
                feuille_resume.cell(row=ligne, column=5).fill = openpyxl.styles.PatternFill(
                    fill_type='solid', fgColor=red_color)
            ligne += 1
        
        # Largeurs colonnes
        feuille_resume.column_dimensions["A"].width = 30
        feuille_resume.column_dimensions["B"].width = 20
        feuille_resume.column_dimensions["F"].width = 15
        
        # === FEUILLE 2: ANOMALIES SECURITE ===
        feuille_anom = fichierXlsx.create_sheet("ANOMALIES_SECURITE", 1)
        headers_anom = ["ETUDE", "SUPPORT DEPART", "SUPPORT ARRIVEE", "CABLE", 
                       "CAPACITE FO", "PORTEE (m)", "PORTEE MAX (m)", "DEPASSEMENT (m)",
                       "HAUTEUR TRAVERSE (m)", "PORTEE KO", "HAUTEUR KO", "A POSER"]
        for col, h in enumerate(headers_anom, 1):
            feuille_anom.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
        
        ligne = 2
        for anom in anomalies:
            feuille_anom.cell(row=ligne, column=1, value=anom.get('etude', ''))
            feuille_anom.cell(row=ligne, column=2, value=anom.get('support_depart', ''))
            feuille_anom.cell(row=ligne, column=3, value=anom.get('support_arrivee', ''))
            feuille_anom.cell(row=ligne, column=4, value=anom.get('cable', ''))
            feuille_anom.cell(row=ligne, column=5, value=anom.get('capacite_fo', 0))
            feuille_anom.cell(row=ligne, column=6, value=anom.get('portee', 0))
            feuille_anom.cell(row=ligne, column=7, value=anom.get('portee_max', 0))
            feuille_anom.cell(row=ligne, column=8, value=anom.get('depassement', 0))
            feuille_anom.cell(row=ligne, column=9, value=anom.get('hauteur_traverse', 0))
            
            portee_ko = anom.get('portee_ko', False)
            hauteur_ko = anom.get('hauteur_ko', False)
            feuille_anom.cell(row=ligne, column=10, value="OUI" if portee_ko else "")
            feuille_anom.cell(row=ligne, column=11, value="OUI" if hauteur_ko else "")
            feuille_anom.cell(row=ligne, column=12, value="OUI" if anom.get('a_poser') else "")
            
            if portee_ko:
                feuille_anom.cell(row=ligne, column=10).fill = openpyxl.styles.PatternFill(
                    fill_type='solid', fgColor=red_color)
            if hauteur_ko:
                feuille_anom.cell(row=ligne, column=11).fill = openpyxl.styles.PatternFill(
                    fill_type='solid', fgColor=orange_color)
            ligne += 1
        
        # Largeurs colonnes
        for col in ["A", "B", "C", "D"]:
            feuille_anom.column_dimensions[col].width = 20
        
        # === FEUILLE 3: SUPPORTS PORTEE MOLLE ===
        feuille_pm = fichierXlsx.create_sheet("PORTEES_MOLLES", 2)
        headers_pm = ["ETUDE", "SUPPORT", "NATURE", "HAUTEUR (m)", "CLASSE", "ETAT"]
        for col, h in enumerate(headers_pm, 1):
            feuille_pm.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
        
        ligne = 2
        for pm in supports_pm:
            feuille_pm.cell(row=ligne, column=1, value=pm.get('etude', ''))
            feuille_pm.cell(row=ligne, column=2, value=pm.get('support', ''))
            feuille_pm.cell(row=ligne, column=3, value=pm.get('nature', ''))
            feuille_pm.cell(row=ligne, column=4, value=pm.get('hauteur', 0))
            feuille_pm.cell(row=ligne, column=5, value=pm.get('classe', ''))
            feuille_pm.cell(row=ligne, column=6, value=pm.get('etat', ''))
            
            feuille_pm.cell(row=ligne, column=2).fill = openpyxl.styles.PatternFill(
                fill_type='solid', fgColor=orange_color)
            ligne += 1
        
        feuille_pm.column_dimensions["A"].width = 30
        feuille_pm.column_dimensions["B"].width = 15
        
        # === FEUILLE 4: DETAIL LIGNES FO ===
        feuille_fo = fichierXlsx.create_sheet("LIGNES_FO", 3)
        headers_fo = ["ETUDE", "CABLE", "CAPACITE FO", "A POSER", "NB PORTEES", 
                     "PORTEE MAX", "SUPPORTS"]
        for col, h in enumerate(headers_fo, 1):
            feuille_fo.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)
        
        ligne = 2
        for nom_etude, etude in etudes.items():
            for l in etude.lignes_tcf:
                if l.capacite_fo == 0:
                    continue
                feuille_fo.cell(row=ligne, column=1, value=nom_etude)
                feuille_fo.cell(row=ligne, column=2, value=l.cable)
                feuille_fo.cell(row=ligne, column=3, value=l.capacite_fo)
                feuille_fo.cell(row=ligne, column=4, value="OUI" if l.a_poser else "")
                feuille_fo.cell(row=ligne, column=5, value=len(l.portees))
                feuille_fo.cell(row=ligne, column=6, value=max(l.portees) if l.portees else 0)
                feuille_fo.cell(row=ligne, column=7, value=' → '.join(l.supports))
                ligne += 1
        
        feuille_fo.column_dimensions["A"].width = 30
        feuille_fo.column_dimensions["G"].width = 60
        
        if 'Sheet' in fichierXlsx.sheetnames:
            del fichierXlsx['Sheet']
        
        fichierXlsx.save(filename=nom_fichier)
        return len(anomalies), len(supports_pm)
