#!/usr/bin/python
# -*- coding: utf-8 -*-

from qgis.core import QgsFeatureRequest, QgsExpression, QgsSpatialIndex, NULL, QgsMessageLog, Qgis
import os
import re
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from .qgis_utils import get_layer_safe, validate_same_crs


class C6_vs_C3A_vs_Bd:
    """ Comparaison des données C6 Par rapport à la base de données"""

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def LectureFichiersExcelsC6(self, df, fichier_c6):
        """Fonction pour parcourir les fichiers Excel pour renseigner la référence des appuis."""

        df_rempl = None
        name = os.path.basename(fichier_c6)
        try:
            with pd.ExcelFile(fichier_c6) as xls:
                # On prend les données à partir de la ligne 7 (la colonne).
                df1 = pd.read_excel(xls, "Export 1", header=7, index_col=None)  # , index_col=0,  skiprows=[15], na_values=["NA"],)
            # print(f"fichier_c6 {fichier_c6}")

            # Creation d'un Dataframe à partir des données collectées dans Excel
            # En supprimant les lignes qui ont une plus de un np.NaN
            # Les colonnes qui nous intéresse sont dans la
            resultat = df1.iloc[:, 0:32:31].dropna(thresh=1)
            resultat = resultat[pd.to_numeric(resultat['N° appui'], errors='coerce').notnull()]
            # resultat.insert(2, "Nature des travaux", np.nan, True)
            resultat.insert(2, "Excel (C6)", name, True)
            # df.loc[df['Nature des travaux'] in ("Remplacement", "Recalage"), 'Nature des travaux'] = 'True'

            # On ajoute une nouvelle colonne qui est conditionnelle
            resultat['Etat'] = resultat['Nature des travaux'].apply(lambda x: 'OUI' if x == "Remplacement" else np.NaN)

            resultat = resultat.rename(columns={"Nature des travaux": "Etat (C6)", "Etat": "Nature des travaux"}, errors="raise")
            # On convertir les numéros des appuis en appuis.
            # resultat['N° appui'] = resultat['N° appui'].astype(int)
            resultat['N° appui'] = pd.to_numeric(resultat['N° appui'], errors="coerce").astype("Int64")

            df = pd.concat([df, resultat], ignore_index=True)
            # print(f"df :\n", tabulate(df, headers="keys", tablefmt="psql"))

            df_rempl = resultat.loc[resultat["Nature des travaux"] == "OUI"]

            # On supprime la colonne qui ne nous interessse pas
            df_rempl = df_rempl.drop("Nature des travaux", axis=1)

            # print(f"df_rempl :\n", tabulate(df_rempl, headers="keys", tablefmt="psql"))

        except (AttributeError, KeyError, ValueError):
            df_rempl = pd.DataFrame()

        except Exception:
            df_rempl = pd.DataFrame()
        return df, df_rempl

    def liste_poteau_c3a_excel(self, fichier_c3a):
        """Traité le fichier Excel C3A et le résume dans un DataFrame"""

        df4 = None
        df4_rempl = None
        names = ['Numéro de prestation', 'type_a', 'num_a', 'type_b', 'num_b', 'long', 'diam_alv', 'Tubage',
                 'diam_tube', 'diam_cable', 'comment_oper', 'remp_perc_a', 'remp_perc_b',
                 'pbo', 'Unnamed: 14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26',
                 '27', '28', '29', '30', '31', '32']

        name = os.path.basename(fichier_c3a)
        try:
            with pd.ExcelFile(fichier_c3a) as xls:
                # On prend les données à partir de la ligne 7 (la colonne).
                df1 = pd.read_excel(xls, "Commandes Fermes", header=13, index_col=None, names=names)

            # On donne la liste des colonnes que l'on souhaite garder
            resultat = df1.loc[:, ["type_a", "num_a", "type_b", "num_b", "comment_oper", "remp_perc_a", "remp_perc_b"]]

            # On ne garde que les laisons ayant un appui typ_a
            # resultat = resultat.loc[(resultat["type_a"] == "A") | (resultat["type_b"] == "A")]

            # On ne garde que des appuis dans ce nouveau DataFrame typ_b
            df_type_a = resultat.loc[:, ["num_a", "remp_perc_a"]].loc[(resultat["type_a"] == "A")]

            # On renomme les colonnes
            df_type_a = df_type_a.rename(columns={"num_a": "inf_num (C3A)",  "remp_perc_a": "Etat (C3A)"}, errors="raise")

            # On ne garde que des appuis dans ce nouveau DataFrame
            df_type_b = resultat.loc[:, ["num_b", "remp_perc_b"]].loc[(resultat["type_b"] == "A")]
            # On renomme les colonnes
            df_type_b = df_type_b.rename(columns={"num_b": "inf_num (C3A)", "remp_perc_b": "Etat (C3A)"}, errors="raise")

            # Fusion des deux DataFrame avec supprimé des doulons et on trie des données
            df4 = pd.concat([df_type_a, df_type_b]).drop_duplicates().reset_index(drop=True).sort_values(by=["inf_num (C3A)", "Etat (C3A)"])
            # print(f"df4 :\n", tabulate(df4, headers="keys", tablefmt="psql"))

            # On suprimé les doublons de numéro de l"appui lorsque présence remplace appui et nan. On garde le premier.
            df4.drop_duplicates(subset=["inf_num (C3A)"], inplace=True)

            df4["Nature des travaux"] = df4["Etat (C3A)"].apply(lambda x: "OUI" if x == "oui remplacement appui" else np.NaN)
            df4.insert(3, "Fichier (C3A) Excel", name, True)

            # On ajoute une nouvelle colonne avec extraction du numéro de l'appui.
            df4["N° appui"] = df4["inf_num (C3A)"].str[6:]
            # df4['N° appui'] = df4['inf_num (C3A)'].apply(lambda x: re.findall(r'^.{4}', x)[0])

            # On vertie le type du champs en entier.
            convert_dict = {'N° appui': int}
            df4 = df4.astype(convert_dict)

            # Réorganiser les colonnes
            df4 = df4.reindex(columns=["N° appui", "inf_num (C3A)", "Nature des travaux", "Etat (C3A)"])

            df4_rempl = df4.loc[df4["Nature des travaux"] == "OUI"]
            # On supprime la colonne qui ne nous interessse pas
            df4_rempl = df4_rempl.drop("Nature des travaux", axis=1)

        except AttributeError as err:
            QgsMessageLog.logMessage(
                f"[C6_vs_C3A_vs_Bd.liste_poteau_c3a_excel] {fichier_c3a}: {err}",
                "PoleAerien",
                Qgis.Warning
            )
            df4 = pd.DataFrame({})
            df4_rempl = pd.DataFrame({})

        except Exception as err:
            QgsMessageLog.logMessage(
                f"[C6_vs_C3A_vs_Bd.liste_poteau_c3a_excel] {fichier_c3a}: {err}",
                "PoleAerien",
                Qgis.Critical
            )

        return df4, df4_rempl

    def liste_poteau_etudes(self, table_poteau, table_decoupage, champs_dcp, valeur_chp_dcp):
        """Récupérer les données liées à la base de données"""
        infra_pt_pot = get_layer_safe(table_poteau, "C6_C3A_BD")
        decoupage = get_layer_safe(table_decoupage, "C6_C3A_BD")

        validate_same_crs(infra_pt_pot, decoupage, "C6_C3A_BD")

        listePoteaux = []
        listePoteaux_rempl = []
        listePoteauxComplet = []
        listePoteauxComplet_rempl = []
        nature_travaux = []
        listeEtat = []
        listeEtat_rempl = []

        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)
        clause = QgsFeatureRequest.OrderByClause('inf_type', ascending=True)
        orderby = QgsFeatureRequest.OrderBy([clause])
        request.setOrderBy(orderby)

        # SEC-02: Echapper les quotes pour éviter injection
        valeur_safe = str(valeur_chp_dcp).replace("'", "''")
        champs_safe = str(champs_dcp).replace('"', '')
        requete = QgsExpression(f'"{champs_safe}" LIKE \'{valeur_safe}\'')
        request_cap_ft = QgsFeatureRequest(requete)
        clause_cap_ft = QgsFeatureRequest.OrderByClause(champs_dcp, ascending=True)
        orderby_cap_ft = QgsFeatureRequest.OrderBy([clause_cap_ft])
        request_cap_ft.setOrderBy(orderby_cap_ft)

        idx_pot = QgsSpatialIndex()
        poteaux = {}
        for feat_pot in infra_pt_pot.getFeatures(request):
            if feat_pot.hasGeometry():
                idx_pot.addFeature(feat_pot)
                poteaux[feat_pot.id()] = feat_pot

        for feat_dcp in decoupage.getFeatures(request_cap_ft):
            if not feat_dcp.hasGeometry():
                continue
            geom_dcp = feat_dcp.geometry()
            bbox = geom_dcp.boundingBox()
            for fid in idx_pot.intersects(bbox):
                feat_pot = poteaux[fid]
                if geom_dcp.contains(feat_pot.geometry()):
                    raw_inf_num = feat_pot["inf_num"]
                    if not raw_inf_num or raw_inf_num == NULL:
                        continue
                    inf_num = str(raw_inf_num)

                    # etat = np.NaN if feat_pot["etat"] == NULL or feat_pot["etat"] == "" else feat_pot["etat"]
                    # nature = "OUI" if etat else np.NaN
                    nature = "OUI" if feat_pot["etat"] != NULL and feat_pot["etat"] != "A RECALER" and feat_pot["etat"] != "" else np.NaN
                    # print(f"etat : {feat_pot['etat']} nature : {nature} ")
                    nature_travaux.append(nature)
                    listePoteauxComplet.append(inf_num)
                    try:
                        inf_num = int(inf_num[:7])
                    except ValueError:
                        QgsMessageLog.logMessage(
                            f"[C6_vs_C3A_vs_Bd.liste_poteau_etudes] inf_num invalide: {inf_num}",
                            "PoleAerien",
                            Qgis.Warning
                        )
                        continue

                    listePoteaux.append(str(inf_num))
                    listeEtat.append(np.NaN if feat_pot["etat"] == NULL else feat_pot["etat"])

                    if nature == "OUI":
                        listePoteauxComplet_rempl.append(inf_num)
                        listePoteaux_rempl.append(str(inf_num))
                        listeEtat_rempl.append(np.NaN if feat_pot["etat"] == NULL else feat_pot["etat"])

        df = pd.DataFrame({'N° appui': listePoteaux, "Nature des travaux": nature_travaux,
                           'inf_num (ETUDES_QGIS)': listePoteauxComplet, "Etat (ETUDES_QGIS)": listeEtat})

        df_rempl = pd.DataFrame({'N° appui': listePoteaux_rempl, 'inf_num (ETUDES_QGIS)': listePoteauxComplet_rempl,
                                 "Etat (ETUDES_QGIS)": listeEtat_rempl})

        return df, df_rempl

    def liste_poteau_c3a_qgis(self, table_cmd, table_decoupage, champs_decoupage, valeur_champs_dcp):
        """Récupérer les données liées à la base de données"""
        cmd = get_layer_safe(table_cmd, "C6_C3A_BD")
        decoupage = get_layer_safe(table_decoupage, "C6_C3A_BD")

        validate_same_crs(cmd, decoupage, "C6_C3A_BD")

        dicoPoteaux = {}
        listePoteauxComplet = set()
        nature_travaux = []

        # SEC-02: Echapper les quotes pour éviter injection
        valeur_safe = str(valeur_champs_dcp).replace("'", "''")
        champs_safe = str(champs_decoupage).replace('"', '')
        requete_dcp = QgsExpression(f'"{champs_safe}" LIKE \'{valeur_safe}\'')
        request_dcp = QgsFeatureRequest(requete_dcp)
        clause_dcp = QgsFeatureRequest.OrderByClause(champs_decoupage, ascending=True)
        orderby_dcp = QgsFeatureRequest.OrderBy([clause_dcp])
        request_dcp.setOrderBy(orderby_dcp)

        requete_cmd = QgsExpression(f"type_a LIKE 'A' OR type_b LIKE 'A'")
        request_cmd = QgsFeatureRequest(requete_cmd)
        clause_cmd = QgsFeatureRequest.OrderByClause("type_a", ascending=True)
        orderby_cmd = QgsFeatureRequest.OrderBy([clause_cmd])
        request_cmd.setOrderBy(orderby_cmd)

        # Index spatial pour cmd (O(n log n) au lieu de O(n×m))
        idx_cmd = QgsSpatialIndex()
        cmd_cache = {}
        for feat_cmd in cmd.getFeatures(request_cmd):
            if feat_cmd.hasGeometry():
                idx_cmd.addFeature(feat_cmd)
                cmd_cache[feat_cmd.id()] = feat_cmd

        for feat_dcp in decoupage.getFeatures(request_dcp):
            if not feat_dcp.hasGeometry():
                continue

            geom_dcp = feat_dcp.geometry()
            bbox = geom_dcp.boundingBox()
            candidates = idx_cmd.intersects(bbox)

            for fid in candidates:
                feat_cmd = cmd_cache.get(fid)
                if not feat_cmd:
                    continue
                if geom_dcp.contains(feat_cmd.geometry()):

                    type_a = feat_cmd["type_a"]
                    type_b = feat_cmd["type_b"]

                    if type_a == "A":
                        # print(f"type_a : ", feat_cmd["type_a"], " num_a : ", feat_cmd["num_a"])

                        num_a = feat_cmd["num_a"]
                        if not num_a or num_a == NULL:
                            continue
                        listePoteauxComplet.add(num_a)

                        etat = np.NaN if feat_cmd["remp_perc_a"] == NULL or not feat_cmd["remp_perc_a"] else feat_cmd["remp_perc_a"]
                        nature = "OUI" if etat == "oui remplacement appui" else etat

                        nature_travaux.append(nature)

                        position = num_a.find('/')
                        try:
                            num_a = int(num_a.split("/")[-1])
                        except ValueError:
                            QgsMessageLog.logMessage(
                                f"[C6_vs_C3A_vs_Bd.liste_poteau_c3a_qgis] num_a invalide: {num_a}",
                                "PoleAerien",
                                Qgis.Warning
                            )
                            continue
                        # listePoteaux.append(str(num_a))

                        if num_a not in dicoPoteaux or nature == "OUI":
                            dicoPoteaux[num_a] = (nature, feat_cmd["num_a"], feat_cmd["remp_perc_a"])
                        # listeEtat.append(etat)

                    if type_b == "A":
                        # print("type_b : ", feat_cmd["type_b"], " num_b : ", feat_cmd["num_b"])
                        num_b = feat_cmd["num_b"]
                        if not num_b or num_b == NULL:
                            continue
                        listePoteauxComplet.add(num_b)

                        etat = np.NaN if feat_cmd["remp_perc_b"] == NULL or not feat_cmd["remp_perc_b"] else feat_cmd["remp_perc_b"]
                        nature = "OUI" if etat == "oui remplacement appui" else np.NaN

                        nature_travaux.append(nature)

                        position = num_b.find('/')
                        try:
                            num_b = int(num_b.split("/")[-1])
                        except ValueError:
                            QgsMessageLog.logMessage(
                                f"[C6_vs_C3A_vs_Bd.liste_poteau_c3a_qgis] num_b invalide: {num_b}",
                                "PoleAerien",
                                Qgis.Warning
                            )
                            continue

                        if num_b not in dicoPoteaux or nature == "OUI":
                            dicoPoteaux[num_b] = (nature, feat_cmd["num_b"], feat_cmd["remp_perc_b"])

        numeroAppui = []
        nature_des_travaux = []
        num_a_b = []
        remp_perc_a_b = []

        numeroAppui_rempl = []
        num_a_b_rempl = []
        remp_perc_a_b_rempl = []

        for appui, (nature_t, inf_num, remp_perc) in dicoPoteaux.items():
            # print(f"appui : {appui} nature_t : {nature_t} inf_num : {inf_num} remp_perc {remp_perc}")
            numeroAppui.append(int(appui))
            nature_des_travaux.append(nature_t)
            num_a_b.append(inf_num)
            remp_perc_a_b.append(remp_perc)

            # Remplacement des appuis
            if nature_t == "OUI":
                numeroAppui_rempl.append(int(appui))
                num_a_b_rempl.append(inf_num)
                remp_perc_a_b_rempl.append(remp_perc)

        if numeroAppui:
            df = pd.DataFrame({'N° appui': numeroAppui, "Nature des travaux": nature_des_travaux,
                               'inf_num (C3A)': num_a_b, 'ETAT (C3A)': remp_perc_a_b})

            df_rempl = pd.DataFrame({'N° appui': numeroAppui_rempl, 'inf_num (C3A)': num_a_b_rempl,
                                     'ETAT (C3A)': remp_perc_a_b_rempl})

            return df, df_rempl

        else:
            return pd.DataFrame({}), pd.DataFrame({})

    def lectureFichierC7(self, fichier_c7):
        """Traité le fichier Excel C3A et le résume dans un DataFrame"""

        df4 = None
        names = ['N° appui', 'Appui avant travaux', 'Type de travaux',
                 'Appui après travaux', 'Métal (L)\n7 m\n50 Kg ', 'Métal (L)\n8 m\n60 Kg ',
                 'Métal (R0)\n10 m\n87 Kg', 'Métal (R1)\n7 m\n72 Kg ', 'Métal (R1)\n8 m\n87 Kg ',
                 'Composite (R0) \n7 m\n38 Kg', 'Composite (R0) \n8 m\n47 Kg', 'Composite (R0) \n10 m\n64 Kg',
                 'Composite (R1) \n7 m\n39 Kg','Composite (R1) \n8 m\n48 Kg', 'Date 1ère livraison', 'Date 2ème livraison']

        name = os.path.basename(fichier_c7)
        try:
            with pd.ExcelFile(fichier_c7) as xls:
                # On prend les données à partir de la ligne 7 (la colonne).
                df1 = pd.read_excel(xls, "Commande", header=16, index_col=None, names=names)

            # print(df1.columns)
            # On donne la liste des colonnes que l'on souhaite garder
            df4 = df1.loc[:, ["N° appui", "Type de travaux", "Appui avant travaux", "Appui après travaux"]]
            # print(resultat.head)

            # On renomme les colonnes
            # df4 = resultat.rename(columns={"num_b": "inf_num (C3A)", "remp_perc_b": "Etat (C3A)"}, errors="raise")
            df4.insert(1, "Fichier (C7) Excel", name, True)

            df4 = df4.rename(columns={"Type de travaux": "Type de travaux (C7)",
                                      "Appui avant travaux": "Appui avant travaux (C7)",
                                      "Appui après travaux": "Appui après travaux (C7)"}, errors="raise")
            # print(f"resultat :\n", tabulate(df4, headers="keys", tablefmt="psql"))

            # # On vertie le type du champs en entier.
            # convert_dict = {'N° appui': int}
            # df4 = resultat.astype(convert_dict)

        except AttributeError as err:
            QgsMessageLog.logMessage(
                f"[C6_vs_C3A_vs_Bd.lectureFichierC7] {fichier_c7}: {err}",
                "PoleAerien",
                Qgis.Warning
            )
            df4 = pd.DataFrame({})

        except Exception as e:
            QgsMessageLog.logMessage(
                f"[C6_vs_C3A_vs_Bd.lectureFichierC7] {fichier_c7}: {e}",
                "PoleAerien",
                Qgis.Critical
            )
            df4 = pd.DataFrame({})

        return df4

    def ecrictureExcel(self, final, fichier):
        """Ecriture du résultat final dans un fichier Excel"""

        # headers = ['N° appui', 'Nature des travaux', 'Études', "inf_num (QGIS)", "Excel"]
        # headers = ['N° appui', 'Nature des travaux', 'Études', 'Excel', 'inf_num (QGIS)', 'Etat']
        # print(tabulate(final.values, headers, tablefmt='psql'))

        nb_erreurs = 0
        with pd.ExcelWriter(fichier, engine="openpyxl") as writer:
            sheet_name = "ANALYSE C6, C3A & BD"
            # Export DataFrame content
            final.to_excel(writer, sheet_name=sheet_name, index=False)
            # Set backgrund colors depending on cell values
            sheet = writer.sheets[sheet_name]

            alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
            index = 1

            # Parcourir les lettres
            for alpha in alphabet:
                # Largeur des colonnes
                sheet.column_dimensions[alpha].width = 25
                index += 1

            for compte, [colA, colB, colC, colD, colE, colF, colG, colH, colI] in enumerate(
                    sheet[f'A2:I{len(final) + 1}']):
                value_num = final["inf_num (ETUDES_QGIS)"].iloc[compte]  # value is "True" or "False"
                value_c3a = final["inf_num (C3A)"].iloc[compte]  # value is "True" or "False"
                value_excel = final["Excel (C6)"].iloc[compte]  # value is "True" or "False"

                # On met toute la ligne en couleur rouge
                if "ABSENT" == value_num or "ABSENT" == value_c3a or "ABSENT" == value_excel:
                    colA.fill = PatternFill("solid", start_color="feb24c")
                    colB.fill = PatternFill("solid", start_color="feb24c")
                    colD.fill = PatternFill("solid", start_color="feb24c")
                    colC.fill = PatternFill("solid", start_color="feb24c")
                    colE.fill = PatternFill("solid", start_color="feb24c")
                    colF.fill = PatternFill("solid", start_color="feb24c")
                    colG.fill = PatternFill("solid", start_color="feb24c")
                    colH.fill = PatternFill("solid", start_color="feb24c")
                    colI.fill = PatternFill("solid", start_color="feb24c")
                    nb_erreurs += 1

        return nb_erreurs

    def ecrictureExcelC6C7C3aBd(self, final, fichier):
        """Ecriture du résultat final dans un fichier Excel"""

        # headers = ['N° appui', 'Nature des travaux', 'Études', "inf_num (QGIS)", "Excel"]
        # headers = ['N° appui', 'Nature des travaux', 'Études', 'Excel', 'inf_num (QGIS)', 'Etat']
        # print(tabulate(final.values, headers, tablefmt='psql'))

        nb_erreurs = 0
        with pd.ExcelWriter(fichier, engine="openpyxl") as writer:
            sheet_name = "ANALYSE C6, C7, C3A & BD"
            # Export DataFrame content
            final.to_excel(writer, sheet_name=sheet_name, index=False)
            # Set backgrund colors depending on cell values
            sheet = writer.sheets[sheet_name]

            alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
            index = 1

            # Parcourir les lettres
            for alpha in alphabet:
                # Largeur des colonnes
                sheet.column_dimensions[alpha].width = 25
                index += 1

            for compte, [colA, colB, colC, colD, colE, colF, colG, colH, colI, colJ, colK] in enumerate(
                    sheet[f'A2:K{len(final) + 1}']):
                value_num = final["inf_num (ETUDES_QGIS)"].iloc[compte]  # value is "True" or "False"
                value_c3a = final["inf_num (C3A)"].iloc[compte]  # value is "True" or "False"
                value_excel_c6 = final["Excel (C6)"].iloc[compte]  # value is "True" or "False"
                value_excel_c7 = final["Fichier (C7) Excel"].iloc[compte]  # value is "True" or "False"

                # On met toute la ligne en couleur rouge
                if ("ABSENT" == value_num or "ABSENT" == value_c3a or "ABSENT" == value_excel_c6 or
                        "ABSENT" == value_excel_c7):
                    colA.fill = PatternFill("solid", start_color="feb24c")
                    colB.fill = PatternFill("solid", start_color="feb24c")
                    colD.fill = PatternFill("solid", start_color="feb24c")
                    colC.fill = PatternFill("solid", start_color="feb24c")
                    colE.fill = PatternFill("solid", start_color="feb24c")
                    colF.fill = PatternFill("solid", start_color="feb24c")
                    colG.fill = PatternFill("solid", start_color="feb24c")
                    colH.fill = PatternFill("solid", start_color="feb24c")
                    colI.fill = PatternFill("solid", start_color="feb24c")
                    colJ.fill = PatternFill("solid", start_color="feb24c")
                    colK.fill = PatternFill("solid", start_color="feb24c")
                    nb_erreurs += 1

        return nb_erreurs
