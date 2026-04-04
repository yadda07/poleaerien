# -*- coding: utf-8 -*-

"""

Async Tasks Module - Non-blocking execution with smooth progress.



Architecture: QgsTask + signals for UI updates.



THREADING SAFETY:

- All QGIS API calls (layer access, feature iteration) MUST happen on main thread

- Worker threads receive pre-extracted data via qgis_data parameter

- Only pure Python operations (Excel parsing, dict comparison) run in workers

- Use signals to communicate results back to main thread

"""



from qgis.PyQt.QtCore import QObject, pyqtSignal, QTimer

from qgis.core import QgsTask, QgsApplication, QgsMessageLog, Qgis
from .compat import MSG_INFO, MSG_WARNING, MSG_CRITICAL

import traceback

import random

try:
    import sip
except ImportError:
    from PyQt6 import sip

import time as _time





def _sip_is_deleted(obj):

    try:

        return sip.isdeleted(obj)

    except RuntimeError:

        return True





def _deserialize_appuis_wkb(appuis_serialized):

    from qgis.core import QgsGeometry

    appuis_data = []

    for appui in appuis_serialized or []:

        geom = None

        wkb = appui.get('geom_wkb')

        if wkb:

            geom = QgsGeometry()

            geom.fromWkb(wkb)

        appuis_data.append({

            'num_appui': appui.get('num_appui', ''),

            'inf_num': appui.get('inf_num', ''),

            'feature_id': appui.get('feature_id'),

            'geom': geom,

            'inf_type': appui.get('inf_type', ''),

            'etat': appui.get('etat', ''),

            'noe_codext': appui.get('noe_codext', ''),

        })

    return appuis_data





def _extract_wkt_endpoints(geom_wkt):

    if not geom_wkt:

        return 0.0, 0.0, 0.0, 0.0

    from qgis.core import QgsGeometry

    geom = QgsGeometry.fromWkt(geom_wkt)

    if not geom or geom.isNull() or geom.isEmpty():

        return 0.0, 0.0, 0.0, 0.0

    line = geom.asMultiPolyline()[0] if geom.isMultipart() else geom.asPolyline()

    if len(line) < 2:

        return 0.0, 0.0, 0.0, 0.0

    return line[0].x(), line[0].y(), line[-1].x(), line[-1].y()





class TaskSignals(QObject):

    """Signals for async task communication"""

    progress = pyqtSignal(int)           # Progress value 0-100

    message = pyqtSignal(str, str)       # (message, color)

    finished = pyqtSignal(dict)          # Result dict

    error = pyqtSignal(str)              # Error message





class SmoothProgressController(QObject):

    """

    Continuous, fluid progress bar controller.

    

    Design principles:

    - NEVER freezes: always moves forward, even if very slowly

    - Exponential easing toward targets (fast start, gradual deceleration)

    - Asymptotic drift: speed decreases near ceiling but never reaches zero

    - Random micro-jitter for natural, organic feel

    - Only explicit set_target(100) can reach 100%

    - Same public API: set_target, reset, set_immediate, set_progress_bar

    """

    

    valueChanged = pyqtSignal(int)

    

    # Soft ceiling for auto-drift. Drift decelerates approaching this

    # but never fully stops. Only set_target(>=100) completes to 100.

    _DRIFT_CEILING = 94.0

    

    # Easing factor per tick toward real target.

    _EASE_FACTOR = 0.10

    

    # Base drift speed range (units per tick).

    _DRIFT_BASE_MIN = 0.05

    _DRIFT_BASE_MAX = 0.18

    

    # Absolute minimum drift per tick -- guarantees bar never freezes.

    _DRIFT_FLOOR = 0.008

    

    # Ticks between drift speed re-randomization.

    _DRIFT_CHANGE_TICKS = 12

    

    def __init__(self, progress_bar=None, interval_ms=30, step=None):

        super().__init__()

        self._progress_bar = progress_bar

        self._current = 0.0

        self._target = 0.0

        self._finished = False

        self._last_int = -1

        

        self._drift_speed = self._random_drift()

        self._drift_tick = 0

        

        self._timer = QTimer(self)

        self._timer.setInterval(interval_ms)

        self._timer.timeout.connect(self._tick)

    

    def set_progress_bar(self, progress_bar):

        """Attach to a QProgressBar"""

        self._progress_bar = progress_bar

    

    def set_target(self, value):

        """Set target value, animation will interpolate to it (never regress)"""

        new_target = max(0.0, min(100.0, float(value)))

        self._target = max(self._target, new_target)

        if new_target >= 100.0:

            self._finished = True

        if not self._timer.isActive():

            self._timer.start()

    

    def set_immediate(self, value):

        """Set value immediately without animation"""

        self._current = float(value)

        self._target = float(value)

        if value >= 100:

            self._finished = True

        self._update_bar()

    

    def reset(self):

        """Reset to 0"""

        self._timer.stop()

        self._current = 0.0

        self._target = 0.0

        self._finished = False

        self._last_int = -1

        self._drift_speed = self._random_drift()

        self._drift_tick = 0

        self._update_bar()

    

    def _tick(self):

        """Main animation tick -- ALWAYS moves forward, NEVER freezes."""

        if self._finished and self._current >= 100.0:

            self._current = 100.0

            self._update_bar()

            self._timer.stop()

            return

        

        # Phase 1: Ease toward real target (exponential deceleration)

        gap = self._target - self._current

        if gap > 0.05:

            step = max(gap * self._EASE_FACTOR, 0.08)

            self._current += step

        elif gap > 0:

            self._current = self._target

        

        # Phase 2: Continuous drift -- ALWAYS active when not finished.

        # Speed scales down as we approach the ceiling, but never zero.

        if not self._finished:

            ceiling = self._DRIFT_CEILING

            room = max(ceiling - self._current, 0.0)

            # Ratio 1.0 = far from ceiling, 0.0 = at ceiling

            ratio = min(room / 30.0, 1.0)

            # Scale drift speed by ratio, but floor guarantees movement

            effective_drift = max(

                self._drift_speed * ratio,

                self._DRIFT_FLOOR

            )

            self._current += effective_drift

            # Hard cap: never exceed ceiling via drift alone

            if self._current > ceiling and self._target < ceiling:

                self._current = ceiling

            

            # Re-randomize drift speed periodically for organic feel

            self._drift_tick += 1

            if self._drift_tick >= self._DRIFT_CHANGE_TICKS:

                self._drift_speed = self._random_drift()

                self._drift_tick = 0

        

        self._current = min(self._current, 100.0)

        self._update_bar()

    

    def _random_drift(self):

        """Generate a random drift speed for organic feel."""

        return random.uniform(self._DRIFT_BASE_MIN, self._DRIFT_BASE_MAX)

    

    def _update_bar(self):

        """Update progress bar and emit signal (only on integer change)"""

        int_val = int(self._current)

        if int_val != self._last_int:

            self._last_int = int_val

            if self._progress_bar:

                self._progress_bar.setValue(int_val)

            self.valueChanged.emit(int_val)





class AsyncTaskBase(QgsTask):

    """

    Base class for async operations.

    Subclasses implement execute() method.

    """

    

    def __init__(self, name, params=None):

        super().__init__(name, QgsTask.CanCancel)

        self.params = params or {}

        self.signals = TaskSignals()

        self.result = {}

        self.exception = None

    

    def run(self):

        """Execute in background thread"""

        try:

            return self.execute()

        except ValueError as e:

            tb = traceback.format_exc()

            self.exception = str(e)

            QgsMessageLog.logMessage(f"Erreur validation: {e}\n{tb}", "PoleAerien", MSG_WARNING)

            return False

        except Exception as e:

            tb = traceback.format_exc()

            self.exception = f"{e}\n{tb}"

            QgsMessageLog.logMessage(f"Erreur task: {e}\n{tb}", "PoleAerien", MSG_CRITICAL)

            return False

    

    def execute(self):

        """Override in subclass - main logic"""

        raise NotImplementedError

    

    def finished(self, success):

        """Callback on main thread"""

        if success:

            self.signals.finished.emit(self.result)

        else:

            self.signals.error.emit(self.exception or "Annule")

    

    def cancel(self):

        """Override cancel: appelle super() pour activer isCanceled(),

        protege contre les objets SIP deja detruits lors du unload QGIS.

        """

        try:

            super().cancel()

        except (RuntimeError, Exception):

            pass

    

    def emit_progress(self, value):

        """Helper to emit progress"""

        self.signals.progress.emit(value)

    

    def emit_message(self, msg, color="black"):

        """Helper to emit message"""

        self.signals.message.emit(msg, color)





class ExcelExportTask(AsyncTaskBase):

    """Async task for Excel export (openpyxl).



    Runs pure Python export code in background to avoid UI freezes.

    """



    def __init__(self, name, export_fn, args=None, kwargs=None, payload=None):

        super().__init__(name)

        self._export_fn = export_fn

        self._args = args or []

        self._kwargs = kwargs or {}

        self._payload = payload or {}



    def execute(self):

        if self.isCanceled():

            return False



        self.emit_progress(92)

        self.emit_message("Export Excel...", "grey")



        try:

            self._export_fn(*self._args, **self._kwargs)

        except Exception as e:

            raise RuntimeError(f"ExcelExportTask.execute: export echoue (cause: {e})") from e



        if self.isCanceled():

            return False



        self.emit_progress(99)

        self.result = dict(self._payload)

        self.result['export_done'] = True

        return True





class CapFtTask(AsyncTaskBase):

    """Async task for CAP_FT analysis.

    

    ARCH-FIX: Reçoit données pré-extraites du main thread.

    Ne fait que traitement pur (comparaison, lecture Excel).

    """

    

    def __init__(self, params, qgis_data=None):

        """

        Args:

            params: Configuration (chemins, fichier_export)

            qgis_data: Données extraites sur main thread:

                - doublons, hors_etude (validation)

                - dico_qgis (poteaux QGIS)

        """

        super().__init__("Analyse CAP_FT", params)

        self.qgis_data = qgis_data or {}

    

    def execute(self):

        if self.isCanceled():

            return False

        

        self.emit_progress(5)

        self.emit_message("Verification donnees...", "grey")

        

        # Données pré-extraites sur main thread

        doublons = self.qgis_data.get('doublons', [])

        hors_etude = self.qgis_data.get('hors_etude', [])

        

        if self.isCanceled():

            return False

        

        if doublons:

            self.result = {'error_type': 'doublons', 'data': doublons}

            return True

        

        if hors_etude:

            self.emit_message(f"ATTENTION: {len(hors_etude)} poteaux hors perimetre etude", "orange")

        

        self.emit_progress(15)

        self.emit_message("Donnees QGIS chargees...", "grey")

        

        dico_qgis = self.qgis_data.get('dico_qgis', {})

        dico_poteaux_prives = self.qgis_data.get('dico_poteaux_prives', {})

        

        if self.isCanceled():

            return False

        

        self.emit_progress(40)

        self.emit_message("Lecture fichiers Excel...", "grey")

        

        # Lecture Excel - thread-safe (pas d'appel QGIS)

        from .CapFt import CapFt

        from .perf_logger import PerfLogger as _PerfLogger

        cap = CapFt()

        _t_excel = _time.perf_counter()

        dico_excel = cap.LectureFichiersExcelsCap_ft(self.params['chemin_cap_ft'])

        _PerfLogger.record('capft', 'lecture_excel',
                           (_time.perf_counter() - _t_excel) * 1000,
                           feature_count=sum(len(v) for v in dico_excel.values()))

        total_fiches = sum(len(v) for v in dico_excel.values())
        if not dico_excel:
            self.emit_message(
                f"ATTENTION: aucun fichier FicheAppui_*.xlsx trouve dans {self.params['chemin_cap_ft']}",
                "orange"
            )
        else:
            self.emit_message(
                f"{len(dico_excel)} dossiers, {total_fiches} fiches appuis detectees",
                "blue"
            )

        if self.isCanceled():

            return False

        

        self.emit_progress(60)

        total_qgis_pot = sum(len(v) for v in dico_qgis.values())
        self.emit_message(
            f"Comparaison: {total_qgis_pot} poteaux QGIS ({len(dico_qgis)} etudes) vs {total_fiches} fiches Excel...",
            "grey"
        )

        

        # Traitement pur - thread-safe
        all_inf_nums = self.qgis_data.get('all_inf_nums', set())

        _t_calc = _time.perf_counter()

        resultats = cap.traitementResultatFinauxCapFt(dico_qgis, dico_excel, all_inf_nums)

        _PerfLogger.record('capft', 'calcul_python',
                           (_time.perf_counter() - _t_calc) * 1000,
                           feature_count=total_qgis_pot)

        # Log matching results in UI
        nb_ok = len(resultats[2]) if len(resultats) > 2 else 0
        nb_absent_qgis = sum(len(v) for v in resultats[0].values())
        nb_absent_fiches = sum(len(v) for v in resultats[1].values())
        nb_hp = sum(len(v) for v in resultats[3].values()) if len(resultats) > 3 else 0
        self.emit_message(
            f"Resultat: {nb_ok} correspondances, {nb_absent_qgis} absents QGIS, "
            f"{nb_absent_fiches} absents fiches appuis"
            + (f", {nb_hp} hors perimetre" if nb_hp else ""),
            "green" if (nb_absent_qgis == 0 and nb_absent_fiches == 0) else "orange"
        )

        if self.isCanceled():

            return False

        

        self.emit_progress(90)

        

        self.result = {

            'success': True,

            'resultats': resultats,

            'dico_excel_introuvable': resultats[0],

            'dico_qgis_introuvable': resultats[1],

            'dico_hors_perimetre': resultats[3] if len(resultats) > 3 else {},

            'fichier_export': self.params['fichier_export'],

            'dico_qgis': dico_qgis,

            'dico_poteaux_prives': dico_poteaux_prives,

            'dico_excel': dico_excel,

            'hors_etude': hors_etude,

            'pending_export': True

        }

        return True





class ComacTask(AsyncTaskBase):

    """Async task for COMAC analysis.

    

    ARCH-FIX: Reçoit données pré-extraites du main thread.

    Ne fait que traitement pur (comparaison, lecture Excel).

    """

    

    def __init__(self, params, qgis_data=None):

        """

        Args:

            params: Configuration (chemins, fichier_export)

            qgis_data: Données extraites sur main thread:

                - doublons, hors_etude (validation)

                - dico_qgis (poteaux QGIS)

        """

        super().__init__("Analyse COMAC", params)

        self.qgis_data = qgis_data or {}

    

    def execute(self):

        if self.isCanceled():

            return False

        

        self.emit_progress(5)

        self.emit_message("Verification donnees...", "grey")

        

        # Données pré-extraites sur main thread

        doublons = self.qgis_data.get('doublons', [])

        hors_etude = self.qgis_data.get('hors_etude', [])

        

        if self.isCanceled():

            return False

        

        if doublons:

            self.result = {'error_type': 'doublons', 'data': doublons}

            return True

        

        if hors_etude:

            self.emit_message(f"ATTENTION: {len(hors_etude)} poteaux hors perimetre etude", "orange")

        

        self.emit_progress(15)

        self.emit_message("Donnees QGIS chargees...", "grey")

        

        dico_qgis = self.qgis_data.get('dico_qgis', {})

        dico_poteaux_prives = self.qgis_data.get('dico_poteaux_prives', {})

        

        if self.isCanceled():

            return False

        

        self.emit_progress(40)

        self.emit_message("Lecture fichiers Excel...", "grey")

        

        # Lecture Excel - thread-safe (pas d'appel QGIS)

        from .Comac import Comac

        from .perf_logger import PerfLogger as _PerfLogger

        com = Comac()

        _t_excel_comac = _time.perf_counter()

        doublons, erreurs_lecture, dico_excel, dico_verif_secu, dico_cables_comac, dico_boitier_comac, dico_coords = com.LectureFichiersExcelsComac(

            self.params['chemin_comac'],

            self.params.get('zone_climatique', 'ZVN')

        )

        

        if self.isCanceled():

            return False

        _PerfLogger.record('comac', 'lecture_excel',

                           (_time.perf_counter() - _t_excel_comac) * 1000,

                           feature_count=sum(len(v) for v in dico_excel.values()))

        

        if doublons:

            for dup_name, path1, path2 in doublons:

                self.emit_message(
                    f"ATTENTION: fichier en doublon '{dup_name}' dans:\n  - {path1}\n  - {path2}",
                    "orange"
                )

        

        if erreurs_lecture:

            for fpath, err_msg in erreurs_lecture.items():

                self.emit_message(f"ATTENTION: lecture impossible {fpath}: {err_msg}", "orange")

        

        if self.isCanceled():

            return False

        

        self.emit_progress(55)

        total_qgis_pot = sum(len(v) for v in dico_qgis.values())
        total_excel_pot = sum(len(v) for v in dico_excel.values())
        self.emit_message(
            f"Comparaison: {total_qgis_pot} poteaux QGIS ({len(dico_qgis)} etudes) vs "
            f"{total_excel_pot} poteaux Excel ({len(dico_excel)} fichiers)...",
            "grey"
        )

        

        # Traitement pur - thread-safe
        all_inf_nums = self.qgis_data.get('all_inf_nums', set())
        coords_qgis = self.qgis_data.get('coords_qgis', {})

        spatial_tol = self.params.get('spatial_tolerance', 7.5)
        _t_calc_comac = _time.perf_counter()
        resultats = com.traitementResultatFinaux(
            dico_qgis, dico_excel, all_inf_nums,
            coords_qgis=coords_qgis, coords_excel=dico_coords,
            spatial_tolerance=spatial_tol
        )
        _PerfLogger.record('comac', 'calcul_python',
                           (_time.perf_counter() - _t_calc_comac) * 1000,
                           feature_count=sum(len(v) for v in dico_qgis.values()))

        # Log matching results in UI
        nb_name = len(resultats[2]) if len(resultats) > 2 else 0
        nb_spatial = len(resultats[4]) if len(resultats) > 4 else 0
        nb_absent_sro = sum(len(v) for v in resultats[0].values())
        nb_absent_excel = sum(len(v) for v in resultats[1].values())
        nb_hp = sum(len(v) for v in resultats[3].values()) if len(resultats) > 3 else 0
        nb_non_resolu = sum(len(v) for v in resultats[5].values()) if len(resultats) > 5 else 0
        self.emit_message(
            f"Resultat: {nb_name + nb_spatial} correspondances ({nb_name} nom, {nb_spatial} spatial), "
            f"{nb_absent_sro} absents couche SRO, {nb_absent_excel} absents Excel"
            + (f", {nb_non_resolu} non resolus commune/code" if nb_non_resolu else "")
            + (f", {nb_hp} hors perimetre" if nb_hp else ""),
            "green" if (nb_absent_sro == 0 and nb_non_resolu == 0) else "orange"
        )

        if self.isCanceled():

            return False

        

        # === Vérification câbles COMAC vs BDD + boîtiers ===

        verif_cables = []

        verif_boitiers_result = {}

        cables_all_for_cache = None

        name_map = {}

        appuis_data = []

        sro = self.params.get('sro')

        has_verif_work = sro and (dico_cables_comac or dico_boitier_comac)

        

        if has_verif_work:

            self.emit_progress(65)

            

            try:

                from .db_connection import DatabaseConnection

                from .cable_analyzer import compter_cables_par_appui, verifier_boitiers

                


                appuis_data = _deserialize_appuis_wkb(self.qgis_data.get('appuis', []))

                

                # === Mapping noms Excel -> QGIS (pour cables et boitiers) ===
                from .core_utils import normalize_appui_num
                name_map = {}  # {excel_norm -> qgis_inf_num}
                absent_sro_keys = set()
                hors_perimetre_keys = set()
                non_resolu_keys = set()
                if resultats and len(resultats) > 2:
                    for _, vals in resultats[2].items():
                        name_map[normalize_appui_num(vals[2], keep_commune=True)] = vals[0]
                if resultats and len(resultats) > 4:
                    for _, m in resultats[4].items():
                        name_map[normalize_appui_num(m['inf_num_excel'], keep_commune=True)] = m['inf_num_qgis']
                if resultats and len(resultats) > 0:
                    for _, appuis in resultats[0].items():
                        for appui in appuis:
                            norm = normalize_appui_num(appui, keep_commune=True)
                            if norm:
                                absent_sro_keys.add(norm)
                if resultats and len(resultats) > 3:
                    for _, appuis in resultats[3].items():
                        for appui in appuis:
                            norm = normalize_appui_num(appui, keep_commune=True)
                            if norm:
                                hors_perimetre_keys.add(norm)
                if resultats and len(resultats) > 5:
                    for _, appuis in resultats[5].items():
                        for appui in appuis:
                            norm = normalize_appui_num(appui, keep_commune=True)
                            if norm:
                                non_resolu_keys.add(norm)

                self.emit_message(
                    f"Mapping poteaux: {len(name_map)} correspondances Excel->QGIS "
                    f"({nb_name} nom, {nb_spatial} spatial)",
                    "blue" if name_map else "orange"
                )
                reverse_map = {v: k for k, v in name_map.items()}

                # === Routage NGE / Axione ===
                be_type = self.params.get('be_type', 'nge')
                gracethd_dir = self.params.get('gracethd_dir', '')

                if be_type == 'axione' and gracethd_dir:
                    # Axione: charger cables depuis GraceTHD
                    from .gracethd_reader import GraceTHDReader
                    self.emit_message(f"Vérif câbles: GraceTHD ({gracethd_dir})...", "grey")
                    reader = GraceTHDReader(gracethd_dir)
                    cables_all = reader.load_cables_as_segments('DI')
                    bpe_list = reader.load_bpe()
                else:
                    # NGE: pipeline fddcpi2 standard
                    cables_all = self.params.get('fddcpi_cables_cache')

                    if cables_all is not None:

                        self.emit_message(

                            f"Vérif câbles: fddcpi2({sro}) depuis cache ({len(cables_all)} segments)",

                            "grey"

                        )

                    else:

                        self.emit_message(f"Vérif câbles: connexion PostgreSQL fddcpi2({sro})...", "grey")

                        db = DatabaseConnection()

                        if db.connect():

                            _t_fddcpi = _time.perf_counter()

                            cables_all = db.execute_fddcpi2(sro)

                            _PerfLogger.record('comac', 'fddcpi2_query',

                                               (_time.perf_counter() - _t_fddcpi) * 1000,

                                               sro=sro, feature_count=len(cables_all) if cables_all else 0)

                        else:

                            cables_all = None

                            self.emit_message("Vérif câbles: connexion PostgreSQL impossible", "orange")

                

                if cables_all is not None:

                    cables_all_for_cache = cables_all

                    cables = [c for c in cables_all
                              if c.posemode in (1, 2) and c.cab_capa >= 6]

                    

                    nb_exclu = len(cables_all) - len(cables)
                    nb_cdi = sum(1 for c in cables if c.cab_type == 'CDI')
                    nb_tra = sum(1 for c in cables if c.cab_type == 'TRA')
                    nb_rac = sum(1 for c in cables if c.cab_type == 'RAC')
                    nb_other = len(cables) - nb_cdi - nb_tra - nb_rac
                    nb_cuivre = sum(1 for c in cables_all
                                     if c.posemode in (1, 2) and c.cab_capa < 6)

                    self.emit_message(

                        f"  {len(cables)} câbles FO aériens/façade "
                        f"(CDI={nb_cdi}, TRA={nb_tra}, RAC={nb_rac}"
                        f"{f', autres={nb_other}' if nb_other else ''}"
                        f", {nb_exclu} exclus dont {nb_cuivre} cuivre capa<6"
                        f" sur {len(cables_all)})",

                        "blue"

                    )

                    

                    if self.isCanceled():

                        return False

                    

                    self.emit_progress(75)

                    self.emit_message("Vérif câbles: intersection appuis...", "grey")

                    

                    # Compter câbles par appui (group_by_gid=True: câbles physiques, pas segments découpés)
                    cable_match = 'line' if (be_type == 'axione' and gracethd_dir) else 'endpoint'
                    # GraceTHD: tolerance large (cable = route complete). Endpoint: 1.5m (cables parfois decales de l'appui)
                    cable_tol = 2.0 if cable_match == 'line' else 1.5

                    cables_par_appui = compter_cables_par_appui(cables, appuis_data, tolerance=cable_tol, group_by_gid=True, match_mode=cable_match, cab_types={'CDI'})

                    # Diagnostic: combien d'appuis ont des cables matches
                    nb_with_cables = sum(1 for v in cables_par_appui.values() if v.get('count', 0) > 0)
                    src_label = "GraceTHD" if (be_type == 'axione' and gracethd_dir) else "BDD"
                    self.emit_message(
                        f"  {nb_with_cables}/{len(cables_par_appui)} appuis QGIS avec câbles {src_label} (mode={cable_match})",
                        "blue"
                    )

                    verif_cables = []

                    if dico_cables_comac:

                        self.emit_progress(80)

                        # Appliquer le mapping aux cles cables COMAC
                        # normalize_appui_num(keep_commune=True): meme normalisation que extraire_appuis_wkb
                        dico_cables_translated = {}
                        nb_translated = 0
                        for key_comac, refs in dico_cables_comac.items():
                            mapped = name_map.get(key_comac)
                            if mapped:
                                dico_cables_translated[normalize_appui_num(mapped, keep_commune=True)] = refs
                                nb_translated += 1
                            else:
                                dico_cables_translated[key_comac] = refs

                        # Comparer COMAC vs cables
                        verif_cables = com.comparer_comac_cables(dico_cables_translated, cables_par_appui, ref_label=src_label)

                        absent_prefix = f"ABSENT_{src_label.upper().replace(' ', '_')}"
                        for entry in verif_cables:
                            if not entry.get('statut', '').startswith('ABSENT'):
                                continue
                            num = entry.get('num_appui', '')
                            excel_key = reverse_map.get(num, num)
                            excel_norm = normalize_appui_num(excel_key, keep_commune=True)
                            if excel_norm in non_resolu_keys:
                                entry['statut'] = f"{absent_prefix}_NON_RESOLU"
                                entry['message'] = (
                                    "Code present dans la couche SRO, mais ambigu entre plusieurs communes "
                                    "ou non leve par le matching spatial"
                                )
                            elif excel_norm in hors_perimetre_keys:
                                entry['statut'] = f"{absent_prefix}_HORS_PERIMETRE"
                                entry['message'] = (
                                    f"Appui present dans la couche SRO {sro}, mais hors zones etude COMAC"
                                    if sro else "Appui present dans la couche SRO, mais hors zones etude COMAC"
                                )
                            elif excel_norm in absent_sro_keys:
                                entry['statut'] = f"{absent_prefix}_ABSENT_SRO"
                                entry['message'] = (
                                    f"Appui inexistant dans infra_pt_pot filtre sur le SRO {sro}"
                                    if sro else "Appui inexistant dans infra_pt_pot filtre SRO"
                                )

                        

                        nb_ok = sum(1 for v in verif_cables if v['statut'] == 'OK')

                        nb_ecart = sum(1 for v in verif_cables if v['statut'] == 'ECART')

                        nb_absent = sum(1 for v in verif_cables if v['statut'] == absent_prefix or v['statut'].endswith('_ABSENT_SRO'))

                        nb_commune = sum(1 for v in verif_cables if v['statut'].endswith('_NON_RESOLU'))

                        nb_hors_p_cables = sum(1 for v in verif_cables if v['statut'].endswith('_HORS_PERIMETRE'))

                        self.emit_message(

                            f"  Vérif câbles: {nb_ok} OK, {nb_ecart} écarts, {nb_absent} absents {src_label}"
                            + (f", {nb_commune} commune/code ambigus" if nb_commune else "")
                            + (f", {nb_hors_p_cables} hors perimetre" if nb_hors_p_cables else ""),

                            "green" if (nb_ecart == 0 and nb_absent == 0 and nb_commune == 0 and nb_hors_p_cables == 0) else "orange"

                        )

                        # Detail par appui en echec (comme Police C6)
                        for v in verif_cables:
                            nb_src = v.get('nb_cables_comac', '?')
                            if v['statut'].endswith('_NON_RESOLU'):
                                self.emit_message(
                                    f"  [COMMUNE] Appui {v['num_appui']}: COMAC={nb_src} câbles, "
                                    f"{src_label}=0 | {v.get('message', '')}",
                                    "orange"
                                )
                            elif v['statut'].endswith('_HORS_PERIMETRE'):
                                self.emit_message(
                                    f"  [HORS_PERIMETRE] Appui {v['num_appui']}: COMAC={nb_src} câbles, "
                                    f"{src_label}=0 | {v.get('message', '')}",
                                    "orange"
                                )
                            elif v['statut'].startswith('ABSENT'):
                                self.emit_message(
                                    f"  [ABSENT] Appui {v['num_appui']}: COMAC={nb_src} câbles, "
                                    f"{src_label}=0 | {v.get('message', '')}",
                                    "orange"
                                )
                            elif v['statut'] == 'ECART':
                                self.emit_message(
                                    f"  [ECART] Appui {v['num_appui']}: COMAC={nb_src}, "
                                    f"{src_label}={v.get('nb_cables_bdd', 0)} | {v.get('message', '')}",
                                    "orange"
                                )

                

                # === Toujours injecter valeur boîtier brute dans verif_cables ===

                if dico_boitier_comac:
                    # Reverse lookup: qgis_name -> excel_name (pour retrouver boitier apres traduction cables)
                    for entry in verif_cables:
                        num = entry['num_appui']
                        excel_key = reverse_map.get(num, num)
                        entry['boitier_comac'] = dico_boitier_comac.get(excel_key, '')

                

                # === Vérification boîtiers vs BPE (uniquement boîtier=oui) ===
                # Traduire cles boitier BT -> GraceTHD
                # normalize_appui_num(keep_commune=True): meme normalisation que extraire_appuis_wkb
                boitier_oui_translated = {}
                if dico_boitier_comac:
                    for k, v in dico_boitier_comac.items():
                        if str(v).lower() == 'oui':
                            raw_key = name_map.get(k, k)
                            boitier_oui_translated[normalize_appui_num(raw_key, keep_commune=True)] = v

                if boitier_oui_translated:

                    self.emit_progress(85)

                    self.emit_message(

                        f"Vérif boîtiers: {len(boitier_oui_translated)} appuis avec boîtier=oui...",

                        "grey"

                    )

                    

                    # Récupérer BPE du SRO
                    if be_type == 'axione' and gracethd_dir:
                        # BPE deja charges depuis GraceTHD (plus haut)
                        pass
                    else:
                        db_bpe = get_shared_connection()

                        if db_bpe.connect():

                            bpe_list = db_bpe.query_bpe_by_sro(sro)

                        else:

                            bpe_list = []

                            self.emit_message("  BPE: connexion PostgreSQL impossible", "orange")

                    

                    # Préparer géométries BPE pour matching spatial

                    bpe_geoms = []

                    from qgis.core import QgsGeometry

                    for bpe in bpe_list:

                        if bpe['geom_wkt']:

                            g = QgsGeometry.fromWkt(bpe['geom_wkt'])

                            if g and not g.isNull():

                                bpe_geoms.append({'geom': g, 'noe_type': bpe['noe_type'], 'gid': bpe['gid']})

                    

                    self.emit_message(

                        f"  {len(bpe_geoms)} BPE récupérés pour vérification boîtier",

                        "blue"

                    )

                    

                    boitier_source = {appui: 'oui' for appui in boitier_oui_translated}

                    verif_boitiers_result = verifier_boitiers(

                        boitier_source, appuis_data, bpe_geoms

                    )

                    

                    nb_ok_b = sum(1 for v in verif_boitiers_result.values() if v['statut'] == 'OK')

                    nb_ecart_b = sum(1 for v in verif_boitiers_result.values() if v['statut'] == 'ECART')

                    nb_err_loc = sum(
                        1 for v in verif_boitiers_result.values()
                        if v['statut'] == 'ERREUR' and v.get('bpe_noe_type') == 'coordonnees GPS absentes'
                    )

                    nb_err_b = sum(
                        1 for v in verif_boitiers_result.values()
                        if v['statut'] == 'ERREUR' and v.get('bpe_noe_type') != 'coordonnees GPS absentes'
                    )

                    self.emit_message(

                        f"  Vérif boîtiers: {nb_ok_b} OK"
                        + (f", {nb_ecart_b} écarts type" if nb_ecart_b else "")
                        + (f", {nb_err_b} absents BPE" if nb_err_b else "")
                        + (f", {nb_err_loc} appuis non resolus" if nb_err_loc else ""),

                        "green" if (nb_err_b == 0 and nb_err_loc == 0 and nb_ecart_b == 0) else "orange"

                    )

                    

                    # Injecter résultats BPE dans verif_cables existants

                    appuis_in_cables = set()

                    for entry in verif_cables:

                        num = entry['num_appui']

                        appuis_in_cables.add(num)

                        bv = verif_boitiers_result.get(num, {})

                        entry['bpe_noe_type'] = bv.get('bpe_noe_type', '')

                        entry['boitier_statut'] = bv.get('statut', '')

                    

                    # Ajouter les appuis boîtier-only (pas de câbles COMAC)

                    for num_appui, bv in verif_boitiers_result.items():

                        if num_appui not in appuis_in_cables:

                            verif_cables.append({

                                'num_appui': num_appui,

                                'nb_cables_comac': 0,

                                'cables_comac': '',

                                'capas_comac': [],

                                'nb_cables_bdd': 0,

                                'capas_bdd': [],

                                'statut': '',

                                'message': '',

                                'boitier_comac': dico_boitier_comac.get(reverse_map.get(num_appui, num_appui), ''),

                                'bpe_noe_type': bv.get('bpe_noe_type', ''),

                                'boitier_statut': bv.get('statut', ''),

                            })

                elif dico_boitier_comac:

                    nb_total = len(dico_boitier_comac)

                    nb_non = sum(1 for v in dico_boitier_comac.values() if str(v).lower() == 'non')

                    self.emit_message(

                        f"Vérif boîtiers: {nb_total} appuis lus, {nb_non} Non, 0 Oui",

                        "grey"

                    )

                    

                # === Log fichier écarts câbles ===
                if verif_cables:
                    try:
                        from .cable_analyzer import write_ecart_log
                        import os as _os
                        _export_dir = _os.path.dirname(self.params.get('fichier_export', ''))
                        _log_path = write_ecart_log(verif_cables, _export_dir, sro)
                        if _log_path:
                            self.emit_message(
                                f"Log écarts câbles: {_os.path.basename(_log_path)}", "grey"
                            )
                    except Exception as _le:
                        self.emit_message(f"[!] write_ecart_log: {_le}", "grey")

            except Exception as e:

                self.emit_message(f"Vérif câbles/boîtiers: erreur {e}", "orange")

        elif not sro:

            self.emit_message("Vérif câbles: SRO non disponible, étape ignorée", "grey")

        elif not dico_cables_comac and not dico_boitier_comac:

            self.emit_message("Vérif câbles: aucune donnée câble/boîtier, étape ignorée", "grey")

        

        etudes_pcm = {}
        erreurs_pcm = {}
        chemin_comac = self.params.get('chemin_comac', '')

        if chemin_comac:
            try:
                from .pcm_parser import parse_repertoire_pcm

                self.emit_progress(89)
                self.emit_message("Lecture fichiers PCM...", "grey")

                etudes_pcm, erreurs_pcm = parse_repertoire_pcm(
                    chemin_comac, self.params.get('zone_climatique', 'ZVN')
                )

                if etudes_pcm:
                    self.emit_message(
                        f"  {len(etudes_pcm)} etudes PCM lues pour dessin et controles",
                        "blue"
                    )
                elif erreurs_pcm:
                    self.emit_message(
                        f"Lecture PCM: {len(erreurs_pcm)} erreur(s)",
                        "orange"
                    )
                else:
                    self.emit_message("Lecture PCM: aucun fichier PCM trouve", "grey")
            except Exception as e_pcm_read:
                self.emit_message(f"Lecture PCM: erreur {e_pcm_read}", "orange")

        # === Verification portees PCM vs BDD/GraceTHD ===
        verif_portees = []

        if etudes_pcm and sro:
            try:
                from .pcm_parser import extraire_portees_par_cable
                from .cable_analyzer import reconstituer_portees_bdd, extraire_portees_gracethd, comparer_portees

                self.emit_progress(91)
                self.emit_message("Verif portees: preparation des troncons PCM...", "grey")

                if etudes_pcm:
                    portees_pcm = extraire_portees_par_cable(etudes_pcm)

                    if portees_pcm:
                        self.emit_message(
                            f"  {len(portees_pcm)} troncons PCM depuis {len(etudes_pcm)} etudes",
                            "blue"
                        )

                        # Deserialiser appuis si pas deja fait (has_verif_work=False)
                        if not appuis_data:
                            appuis_data = _deserialize_appuis_wkb(self.qgis_data.get('appuis', []))

                        self.emit_progress(93)
                        self.emit_message("Verif portees: reconstruction troncons reference...", "grey")

                        be_type = self.params.get('be_type', 'nge')
                        gracethd_dir = self.params.get('gracethd_dir', '')

                        troncons_ref = []
                        if be_type == 'axione' and gracethd_dir:
                            from .gracethd_reader import GraceTHDReader
                            reader_p = GraceTHDReader(gracethd_dir)
                            cables_nodes = reader_p.load_cables_with_nodes('DI')
                            troncons_ref = extraire_portees_gracethd(
                                cables_nodes, appuis_data, tolerance=2.0
                            )
                        elif cables_all_for_cache is not None:
                            troncons_ref = reconstituer_portees_bdd(
                                cables_all_for_cache, appuis_data, tolerance=0.5
                            )
                        else:
                            from .db_connection import get_shared_connection as _get_conn_p
                            self.emit_message("Verif portees: chargement cables fddcpi2...", "grey")
                            db_p = _get_conn_p()
                            if db_p.connect():
                                cables_p = db_p.execute_fddcpi2(sro)
                                if cables_p:
                                    cables_all_for_cache = cables_p
                                    troncons_ref = reconstituer_portees_bdd(
                                        cables_p, appuis_data, tolerance=0.5
                                    )

                        if troncons_ref:
                            self.emit_message(
                                f"  {len(troncons_ref)} troncons reference ({troncons_ref[0].source})",
                                "blue"
                            )

                            self.emit_progress(95)
                            self.emit_message("Verif portees: comparaison PCM vs reference...", "grey")

                            # Construire pcm_name_map:
                            # - cles = noms SANS commune (pour matcher PCM bruts: BT0030)
                            # - valeurs = noms SANS commune (pour matcher _translate sur ref)
                            # _translate() appelle _norm() sans keep_commune -> strip /commune
                            # Les deux cotes doivent etre dans le meme format
                            from .core_utils import normalize_appui_num
                            pcm_name_map = {}
                            for k, v in name_map.items():
                                k_sans = k.split('/')[0] if '/' in k else k
                                v_sans = normalize_appui_num(v)
                                pcm_name_map[k_sans] = v_sans

                            self.emit_message(
                                f"  Verif portees: pcm_name_map {len(pcm_name_map)} entrees"
                                f" (sample keys: {list(pcm_name_map.keys())[:5]})"
                                f" -> (sample vals: {list(pcm_name_map.values())[:5]})",
                                "grey"
                            )
                            # Log sample troncons ref pour diagnostic
                            if troncons_ref:
                                sample_ref = [(t.support_depart, t.support_arrivee, t.portee_m) for t in troncons_ref[:5]]
                                self.emit_message(
                                    f"  Troncons ref (sample): {sample_ref}",
                                    "grey"
                                )
                            # Log sample portees PCM pour diagnostic
                            if portees_pcm:
                                sample_pcm = [(p.support_depart, p.support_arrivee, p.portee_m, p.a_poser) for p in portees_pcm[:5]]
                                self.emit_message(
                                    f"  Portees PCM (sample): {sample_pcm}",
                                    "grey"
                                )

                            verif_portees = comparer_portees(
                                portees_pcm, troncons_ref,
                                tolerance_pct=15.0,
                                name_map=pcm_name_map,
                            )

                            nb_ok_p = sum(1 for v in verif_portees if v['statut'] == 'OK')
                            nb_ecart_p = sum(1 for v in verif_portees if v['statut'] == 'ECART')
                            nb_abs_ref = sum(1 for v in verif_portees if v['statut'] == 'ABSENT_REF')
                            self.emit_message(
                                f"  Verif portees: {nb_ok_p} OK, {nb_ecart_p} ecarts, "
                                f"{nb_abs_ref} absents ref",
                                "green" if (nb_ecart_p == 0 and nb_abs_ref == 0) else "orange"
                            )
                        else:
                            self.emit_message("Verif portees: aucun troncon reference reconstruit", "grey")
                    else:
                        self.emit_message("Verif portees: aucun troncon FO dans les PCM", "grey")
                elif erreurs_pcm:
                    self.emit_message(
                        f"Verif portees: {len(erreurs_pcm)} erreur(s) lecture PCM", "orange"
                    )
                else:
                    self.emit_message("Verif portees: aucun fichier PCM trouve", "grey")

            except Exception as e:
                self.emit_message(f"Verif portees: erreur {e}", "orange")

        # === Comparaison PCM vs BDD (supports POT-BT uniquement) ===
        pcm_vs_bdd_result = None
        chemin_comac_pcm = self.params.get('chemin_comac', '')

        if chemin_comac_pcm and sro:
            try:
                from .pcm_bdd_comparator import comparer_batch_pcm_vs_bdd
                from .pcm_parser import parse_repertoire_pcm as _parse_pcm_batch

                self.emit_progress(96)
                self.emit_message("Comparaison PCM vs BDD (supports POT-BT)...", "grey")

                if not etudes_pcm:
                    etudes_pcm, _ = _parse_pcm_batch(
                        chemin_comac_pcm,
                        self.params.get('zone_climatique', 'ZVN')
                    )

                if etudes_pcm:
                    if not appuis_data:
                        appuis_data = _deserialize_appuis_wkb(self.qgis_data.get('appuis', []))

                    if appuis_data:
                        poteaux_bdd_dicts = []
                        for appui in appuis_data:
                            inf_type = appui.get('inf_type', '')
                            if inf_type != 'POT-BT':
                                continue
                            geom = appui.get('geom')
                            ax, ay = 0.0, 0.0
                            if geom and not geom.isNull():
                                pt = geom.asPoint()
                                ax, ay = pt.x(), pt.y()
                            poteaux_bdd_dicts.append({
                                'inf_num': appui.get('inf_num', '') or appui.get('num_appui', ''),
                                'noe_codext': appui.get('inf_num', '') or appui.get('num_appui', ''),
                                'inf_type': inf_type,
                                'etat': appui.get('etat', ''),
                                'x': ax,
                                'y': ay,
                            })

                        pcm_vs_bdd_result = comparer_batch_pcm_vs_bdd(
                            etudes_pcm, poteaux_bdd_dicts, []
                        )

                        self.emit_message(
                            f"PCM vs BDD: {pcm_vs_bdd_result.nb_supports_ok} supports OK, "
                            f"{pcm_vs_bdd_result.nb_supports_absent} absents, "
                            f"{pcm_vs_bdd_result.nb_supports_type_ko} type KO, "
                            f"{pcm_vs_bdd_result.nb_supports_coord_ko} coord KO",
                            "green" if (
                                pcm_vs_bdd_result.nb_supports_absent == 0
                                and pcm_vs_bdd_result.nb_supports_type_ko == 0
                                and pcm_vs_bdd_result.nb_supports_coord_ko == 0
                            ) else "orange"
                        )
                    else:
                        self.emit_message(
                            "PCM vs BDD: aucun appui QGIS disponible, comparaison supports ignoree",
                            "orange"
                        )
                else:
                    self.emit_message("PCM vs BDD: aucun fichier PCM trouve", "grey")
            except Exception as e_pcm:
                self.emit_message(f"PCM vs BDD: erreur {e_pcm}", "orange")
        elif etudes_pcm and not sro:
            self.emit_message("PCM vs BDD: SRO non disponible, etape ignoree", "grey")

        self.emit_progress(97)

        self.result = {

            'success': True,

            'resultats': resultats,

            'dico_excel_introuvable': resultats[0],

            'dico_qgis_introuvable': resultats[1],

            'dico_hors_perimetre': resultats[3] if len(resultats) > 3 else {},

            'dico_spatial_match': resultats[4] if len(resultats) > 4 else {},

            'dico_non_resolu': resultats[5] if len(resultats) > 5 else {},

            'fichier_export': self.params['fichier_export'],

            'dico_qgis': dico_qgis,

            'dico_poteaux_prives': dico_poteaux_prives,

            'dico_excel': dico_excel,

            'dico_verif_secu': dico_verif_secu,

            'verif_cables': verif_cables,

            'verif_boitiers': verif_boitiers_result,

            'dico_boitier_comac': dico_boitier_comac,

            'verif_portees': verif_portees,

            'etudes_pcm': etudes_pcm,

            'erreurs_pcm': erreurs_pcm,

            'hors_etude': hors_etude,

            'pending_export': True,

            'fddcpi_cables_all': cables_all_for_cache,

            'fddcpi_sro': sro,

            'appuis_wkb': self.qgis_data.get('appuis', []),

            'be_type': self.params.get('be_type', 'nge'),

            'coords_absents': self._filter_coords_absents(resultats[0], dico_coords),

            'pcm_vs_bdd': pcm_vs_bdd_result,

        }

        return True

    @staticmethod
    def _filter_coords_absents(dico_introuvables, dico_coords):
        """Filtre les coordonnees pour ne garder que les poteaux absents QGIS.
        
        Args:
            dico_introuvables: {fichier: [noms_poteaux]} depuis traitementResultatFinaux
            dico_coords: {nompot: (x, y)} depuis LectureFichiersExcelsComac
        
        Returns:
            list[dict]: [{nom, x, y, fichier}] pour les absents avec coordonnees
        """
        if not dico_introuvables or not dico_coords:
            return []
        result = []
        for fichier, poteaux in dico_introuvables.items():
            for pot in poteaux:
                coords = dico_coords.get(pot)
                if coords:
                    result.append({
                        'nom': pot,
                        'x': coords[0],
                        'y': coords[1],
                        'fichier': fichier
                    })
        return result


class C6BdTask(AsyncTaskBase):

    """Async task for C6 vs BD comparison.

    

    ARCH-FIX: Reçoit données pré-extraites du main thread.

    Ne fait que traitement pur (lecture Excel, comparaison pandas).

    """

    

    def __init__(self, params, qgis_data=None):

        """

        Args:

            params: Configuration (chemins, fichier_export)

            qgis_data: Données extraites sur main thread:

                - df_qgis (DataFrame poteaux QGIS)

        """

        super().__init__("Comparaison C6 vs BD", params)

        self.qgis_data = qgis_data or {}

    

    def execute(self):

        if self.isCanceled():

            return False

        

        self.emit_progress(10)

        self.emit_message("Lecture fichiers C6...", "grey")

        

        # Lecture Excel - thread-safe (pas d'appel QGIS)

        import pandas as pd

        import os

        from .C6_vs_Bd import C6_vs_Bd

        from .core_utils import normalize_appui_num

        

        c6bd = C6_vs_Bd()

        df = pd.DataFrame(columns=["N° appui", "Nature des travaux", "Études", "Excel"])

        

        try:

            df1 = c6bd.LectureFichiersExcelsC6(df, self.params['repertoire_c6'])

        except Exception as e:

            raise RuntimeError(f"Lecture C6 echouee: {e}") from e

        

        if self.isCanceled():

            return False

        

        self.emit_progress(40)

        self.emit_message("Fusion donnees...", "grey")

        

        # Données pré-extraites sur main thread

        df2 = self.qgis_data.get('df_qgis')

        if df2 is None:

            raise ValueError("Donnees QGIS manquantes")

        

        # Traitement pur pandas - thread-safe

        def _uniq_join(sr):

            vals = [str(v).strip() for v in sr.tolist() 

                    if v is not None and str(v).strip() and str(v).strip() != "nan"]

            if not vals:

                return ""

            return ", ".join(sorted(set(vals)))

        

        df1 = df1.copy()

        df2 = df2.copy()

        

        df1["appui_key"] = df1["N° appui"].apply(normalize_appui_num)

        df2["appui_key"] = df2["N° appui"].apply(normalize_appui_num)

        

        df1 = df1[df1["appui_key"].astype(bool)]

        df2 = df2[df2["appui_key"].astype(bool)]

        

        if self.isCanceled():

            return False

        

        self.emit_progress(55)

        

        df1_g = (

            df1.groupby("appui_key", dropna=False)

            .agg({

                "N° appui": "first",

                "Excel": _uniq_join,

                "Études": _uniq_join,

                "Nature des travaux": _uniq_join,

            })

            .reset_index()

            .rename(columns={

                "Excel": "Fichiers (Excel)",

                "Études": "Études (Excel)",

                "Nature des travaux": "Nature (Excel)",

            })

        )

        

        df2_g = (

            df2.groupby("appui_key", dropna=False)

            .agg({

                "inf_num (QGIS)": "first",

                "Études": _uniq_join,

                "Nature des travaux": _uniq_join,

                "Etat": _uniq_join,

            })

            .reset_index()

            .rename(columns={

                "Études": "Études (QGIS)",

                "Nature des travaux": "Nature (QGIS)",

            })

        )

        

        if self.isCanceled():

            return False

        

        self.emit_progress(70)

        self.emit_message("Calcul statuts...", "grey")

        

        final_df = pd.merge(df1_g, df2_g, on=["appui_key"], how="outer")

        # Convertir en string avant fillna pour éviter erreur de type Int64 vs string

        final_df["N° appui"] = final_df["N° appui"].astype(str).replace('<NA>', '').replace('nan', '')

        final_df["N° appui"] = final_df.apply(

            lambda r: r["N° appui"] if r["N° appui"] else str(r["inf_num (QGIS)"] or ""), axis=1

        )

        

        def _statut(row):

            def _val(v):

                if v is None or (isinstance(v, float) and pd.isna(v)):

                    return ""

                s = str(v).strip()

                return "" if s.lower() == "nan" else s

            

            excel_val = _val(row.get("Fichiers (Excel)"))

            qgis_val = _val(row.get("inf_num (QGIS)"))

            has_excel = bool(excel_val)

            has_qgis = bool(qgis_val)

            

            if has_excel and has_qgis:

                et_x = set([v.strip() for v in _val(row.get("Études (Excel)")).split(",") if v.strip()])

                et_q = set([v.strip() for v in _val(row.get("Études (QGIS)")).split(",") if v.strip()])

                if et_x and et_q and et_x.isdisjoint(et_q):

                    return "A VERIFIER"

                return "OK"

            if has_excel and not has_qgis:

                return "ABSENT QGIS"

            if has_qgis and not has_excel:

                return "ABSENT EXCEL"

            return "ABSENT"

        

        final_df["Statut"] = final_df.apply(_statut, axis=1)

        final_df = final_df.sort_values(by=["Statut", "appui_key"], ascending=[True, True])

        

        cols = [

            "N° appui", "appui_key", "Statut",

            "Fichiers (Excel)", "Études (Excel)", "Nature (Excel)",

            "inf_num (QGIS)", "Études (QGIS)", "Nature (QGIS)", "Etat",

        ]

        final_df = final_df[[c for c in cols if c in final_df.columns]]

        

        if self.isCanceled():

            return False

        

        self.emit_progress(85)

        

        self.result = {

            'success': True,

            'final_df': final_df,

            'fichier_export': self.params['fichier_export'],

            'df_poteaux_out': self.qgis_data.get('df_poteaux_out'),

            'verif_etudes': self.qgis_data.get('verif_etudes'),

            'pending_export': True

        }

        return True





class PoliceC6Task(AsyncTaskBase):

    """Async task for Police C6 analysis.

    

    ARCHITECTURE:

    - Main thread: Extract QGIS data (appuis, SRO)

    - Worker thread: PostgreSQL query (fddcpi2), data comparison, Excel export

    

    Thread-safe: Ne fait aucun appel QGIS dans execute().

    """

    

    def __init__(self, params, qgis_data=None):

        """

        Args:

            params: Configuration

                - c6_files: Liste de (etude_name, fichier_c6)

                - export_path: Chemin export Excel

                - sro: Code SRO pour fddcpi2

            qgis_data: Données pré-extraites sur main thread

                - appuis: Liste de {'num_appui': str, 'geom': WKT/coords}

        """

        super().__init__("Analyse Police C6", params)

        self.qgis_data = qgis_data or {}

        self._cancelled = False

    

    def execute(self):

        if self.isCanceled():

            return False

        

        self.emit_progress(5)

        self.emit_message("Initialisation analyse Police C6...", "grey")

        

        import os

        

        # Import ici pour éviter import circulaire

        from .db_connection import DatabaseConnection

        from .cable_analyzer import compter_cables_par_appui, _parse_attaches_geoms

        

        from qgis.core import QgsGeometry

        

        c6_files = self.params.get('c6_files', [])

        sro = self.params.get('sro', '')

        export_path = self.params.get('export_path', '')

        

        # Désérialiser WKB → QgsGeometry (thread-safe)

        appuis_data = []

        for appui in self.qgis_data.get('appuis', []):

            geom = None

            wkb = appui.get('geom_wkb')

            if wkb:

                geom = QgsGeometry()

                geom.fromWkb(wkb)

            appuis_data.append({

                'num_appui': appui.get('num_appui', ''),

                'feature_id': appui.get('feature_id'),

                'geom': geom

            })

        

        if not c6_files:

            self.result = {'error': 'Aucun fichier C6 fourni'}

            return True

        

        if not sro:

            self.result = {'error': 'SRO non défini'}

            return True

        

        if self.isCanceled():

            return False

        

        # 1. Récupération câbles via cache ou DatabaseConnection

        self.emit_progress(10)

        

        try:

            # === Routage NGE / Axione ===
            be_type = self.params.get('be_type', 'nge')
            gracethd_dir = self.params.get('gracethd_dir', '')
            attaches_raw = []

            if be_type == 'axione' and gracethd_dir:
                # Axione: charger cables et BPE depuis GraceTHD
                from .gracethd_reader import GraceTHDReader
                self.emit_message(f"GraceTHD: chargement ({gracethd_dir})...", "grey")
                reader = GraceTHDReader(gracethd_dir)
                cables_all = reader.load_cables_as_segments('DI')
                bpe_list = reader.load_bpe()

                self.emit_message(
                    f"  {len(cables_all)} câbles DI charges + {len(bpe_list)} BPE depuis GraceTHD",
                    "blue"
                )

            else:
                # NGE: pipeline fddcpi2 standard
                cables_all = self.params.get('fddcpi_cables_cache')

                if cables_all is not None:

                    self.emit_message(

                        f"fddcpi2({sro}) depuis cache ({len(cables_all)} segments)",

                        "grey"

                    )

                else:

                    self.emit_message(f"Connexion PostgreSQL et appel fddcpi2({sro})...", "grey")

                    db = DatabaseConnection()

                    if not db.connect():

                        self.result = {'error': 'Connexion PostgreSQL impossible'}

                        return True

                    _t_fddcpi_c6 = _time.perf_counter()

                    cables_all = db.execute_fddcpi2(sro)

                    from .perf_logger import PerfLogger as _PerfLogger

                    _PerfLogger.record('police_c6', 'fddcpi2_query',

                                       (_time.perf_counter() - _t_fddcpi_c6) * 1000,

                                       sro=sro, feature_count=len(cables_all) if cables_all else 0)

            # Garder câbles de distribution aériens + façade (cab_type='CDI' + posemode 1 ou 2)
            cables = [c for c in cables_all if c.cab_type == 'CDI' and c.posemode in (1, 2)]

            nb_exclu = len(cables_all) - len(cables)

            

            self.emit_message(

                f"  {len(cables)} câbles CDI aériens/façade retenus ({nb_exclu} exclus sur {len(cables_all)})",

                "blue"

            )

            

            if be_type != 'axione' or not gracethd_dir:
                # NGE: utiliser le cache P-02 si disponible, sinon query PG
                bpe_list_cache = self.params.get('bpe_list_cache')
                attaches_cache = self.params.get('attaches_cache')

                if bpe_list_cache is not None:
                    bpe_list = bpe_list_cache
                    attaches_raw = attaches_cache or []
                    self.emit_message(
                        f"  {len(bpe_list)} BPE + {len(attaches_raw)} attaches depuis cache P-02",
                        "grey"
                    )
                else:
                    db_bpe = DatabaseConnection()

                    if db_bpe.connect():

                        bpe_list = db_bpe.query_bpe_by_sro(sro)

                        attaches_raw = db_bpe.query_attaches_by_sro(sro)

                    else:

                        bpe_list = []

                        self.emit_message("  BPE: connexion PostgreSQL impossible", "orange")

            

            attaches_parsed = _parse_attaches_geoms(attaches_raw)

            if attaches_parsed:

                self.emit_message(

                    f"  {len(attaches_parsed)} attaches chargees pour detection indirecte",

                    "grey"

                )

            

            # Préparer géométries BPE pour matching spatial

            bpe_geoms = []

            for bpe in bpe_list:

                if bpe['geom_wkt']:

                    g = QgsGeometry.fromWkt(bpe['geom_wkt'])

                    if g and not g.isNull():

                        bpe_geoms.append({'geom': g, 'noe_type': bpe['noe_type'], 'gid': bpe['gid']})

            

            self.emit_message(

                f"  {len(bpe_geoms)} BPE récupérés pour vérification boîtier",

                "blue"

            )

        except Exception as e:

            self.result = {'error': f'Erreur fddcpi2: {e}'}

            return True

        

        if self.isCanceled():

            return False

        

        # 2. Pre-calcul cables_par_appui (une seule fois pour toutes les etudes)

        p_match = 'line' if (be_type == 'axione' and gracethd_dir) else 'endpoint'
        p_src_label = "GraceTHD" if (be_type == 'axione' and gracethd_dir) else "BDD"

        # GraceTHD: tolerance large (cable = route complete). Endpoint: 1.5m (cables parfois decales de l'appui)
        p_tol = 2.0 if p_match == 'line' else 1.5

        cables_par_appui_cached = compter_cables_par_appui(
            cables, appuis_data, tolerance=p_tol,
            attaches_parsed=attaches_parsed,
            match_mode=p_match
        )

        nb_with = sum(1 for v in cables_par_appui_cached.values() if v.get('count', 0) > 0)
        self.emit_message(
            f"Pré-calcul câbles: {nb_with}/{len(cables_par_appui_cached)} appuis avec câbles {p_src_label} (mode={p_match})",
            "blue"
        )

        # 3. Traitement de chaque étude (P-04: parallele via ThreadPoolExecutor)

        self.emit_progress(20)

        stats_globales = []

        all_anomaly_cables = []

        total = len(c6_files)

        self.emit_message(
            f"Traitement de {total} etude(s)...", "blue"
        )

        for i, (etude_name, c6_file) in enumerate(c6_files):
            if self.isCanceled():
                break

            self.emit_message(f"[{i+1}/{total}] {etude_name}...", "blue")
            self.emit_progress(20 + int(((i + 1) / total) * 60))

            res = _run_one_study(
                etude_name, c6_file,
                cables_par_appui_cached, appuis_data,
                bpe_geoms, attaches_parsed,
                cables, p_src_label
            )

            for msg, color in res.get('messages', []):
                self.emit_message(msg, color)

            if res.get('stats'):
                stats_globales.append(res['stats'])
                all_anomaly_cables.extend(res.get('anomaly_cables', []))

        

        if self.isCanceled():

            return False

        

        # Study loop summary
        nb_etudes_ok = sum(1 for s in stats_globales if s['nb_ecart'] == 0 and s['nb_absent'] == 0 and s.get('nb_boitier_err', 0) == 0)
        nb_etudes_anom = len(stats_globales) - nb_etudes_ok
        nb_etudes_skip = total - len(stats_globales)
        summary_parts = [f"{len(stats_globales)}/{total} etudes traitees"]
        if nb_etudes_ok:
            summary_parts.append(f"{nb_etudes_ok} sans anomalie")
        if nb_etudes_anom:
            summary_parts.append(f"{nb_etudes_anom} avec anomalie(s)")
        if nb_etudes_skip:
            summary_parts.append(f"{nb_etudes_skip} ignoree(s)/erreur")
        self.emit_message(
            "Bilan: " + ", ".join(summary_parts),
            "green" if nb_etudes_anom == 0 and nb_etudes_skip == 0 else "orange"
        )

        # Bilan consolide absents: BT vs non-BT
        all_absents_bt = set()
        all_absents_other = set()
        total_ok = 0
        total_ecart = 0
        for s in stats_globales:
            total_ok += s.get('nb_ok', 0)
            total_ecart += s.get('nb_ecart', 0)
            for c in s.get('detail', []):
                if c['statut'].startswith('ABSENT'):
                    num = c['num_appui']
                    if num.startswith('BT'):
                        all_absents_bt.add(num)
                    else:
                        all_absents_other.add(num)
        self.emit_message(
            f"Câbles consolidé: {total_ok} OK, {total_ecart} écarts, "
            f"{len(all_absents_bt)} BT absents (normal), {len(all_absents_other)} non-BT absents",
            "blue"
        )
        if all_absents_other:
            self.emit_message(
                f"  Non-BT absents: {sorted(all_absents_other)[:20]}",
                "orange"
            )

        # 3. Export Excel (skip en mode batch: le rapport unifie s'en charge)

        self.emit_progress(85)

        skip_export = self.params.get('skip_individual_export', False)

        if export_path and stats_globales and not skip_export:

            self.emit_message("Export Excel...", "grey")

            try:

                excel_file = self._export_excel(stats_globales, export_path)

                self.emit_message(f"Export: {os.path.basename(excel_file)}", "green")

            except Exception as e:

                self.emit_message(f"[!] Erreur export: {e}", "orange")

        

        self.emit_progress(100)

        

        # Sérialiser câbles pour chargement couche QGIS (main thread)

        cables_for_layer = []

        for cable in cables:

            cables_for_layer.append({

                'gid_dc2': cable.gid_dc2,

                'gid_dc': cable.gid_dc,

                'cab_capa': cable.cab_capa,

                'cab_type': cable.cab_type,

                'cb_etiquet': cable.cb_etiquet,

                'posemode': cable.posemode,

                'length': cable.length,

                'geom_wkt': cable.geom_wkt

            })

        

        self.result = {

            'success': True,

            'stats': stats_globales,

            'total_etudes': len(c6_files),

            'etudes_traitees': len(stats_globales),

            'cables': cables_for_layer,

            'anomaly_cables': all_anomaly_cables,

            'sro': sro,

            'fddcpi_cables_all': cables_all,

            'fddcpi_sro': sro,

            'appuis_wkb': self.qgis_data.get('appuis', []),

            'be_type': be_type,

        }

        return True

    

    def _verifier_boitiers(self, boitier_c6, appuis_data, bpe_geoms, attaches_parsed=None):

        """Delegue a cable_analyzer.verifier_boitiers()."""

        from .cable_analyzer import verifier_boitiers

        return verifier_boitiers(boitier_c6, appuis_data, bpe_geoms, attaches_parsed=attaches_parsed)



    def _export_excel(self, stats, export_path):

        """Export des résultats vers Excel avec récap + détail par appui."""

        import os

        from datetime import datetime

        

        try:

            import openpyxl

            from openpyxl.styles import Font, PatternFill, Alignment

        except ImportError:

            raise RuntimeError("openpyxl non installé")

        

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if os.path.isdir(export_path):

            filename = os.path.join(export_path, f"PoliceC6_Export_{timestamp}.xlsx")

        else:

            filename = export_path

        

        wb = openpyxl.Workbook()

        

        header_font = Font(bold=True, color="FFFFFF")

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ecart_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        absent_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        

        # === Feuille 1: Récapitulatif par étude ===

        ws_recap = wb.active

        ws_recap.title = "Récapitulatif"

        

        recap_headers = ["Étude", "Appuis C6", "OK", "Écarts", "Absents BDD", "Boîtier absent"]

        for col, header in enumerate(recap_headers, 1):

            cell = ws_recap.cell(row=1, column=col, value=header)

            cell.font = header_font

            cell.fill = header_fill

        

        for row, stat in enumerate(stats, 2):

            ws_recap.cell(row=row, column=1, value=stat['etude'])

            ws_recap.cell(row=row, column=2, value=stat['appuis_c6'])

            ws_recap.cell(row=row, column=3, value=stat.get('nb_ok', 0))

            ws_recap.cell(row=row, column=4, value=stat.get('nb_ecart', 0))

            ws_recap.cell(row=row, column=5, value=stat.get('nb_absent', 0))

            ws_recap.cell(row=row, column=6, value=stat.get('nb_boitier_err', 0))

            

            # Coloriser si anomalies

            has_anomaly = (stat.get('nb_ecart', 0) > 0 or stat.get('nb_absent', 0) > 0

                           or stat.get('nb_boitier_err', 0) > 0)

            if has_anomaly:

                for c in range(1, 7):

                    ws_recap.cell(row=row, column=c).fill = ecart_fill

        

        # Ajuster largeurs

        ws_recap.column_dimensions['A'].width = 35

        for col_letter in ['B', 'C', 'D', 'E', 'F']:

            ws_recap.column_dimensions[col_letter].width = 15

        

        # === Feuille 2: Détail par appui ===

        ws_detail = wb.create_sheet("Détail par appui")

        

        detail_headers = [

            "Étude", "N° Appui", "Câbles C6", "Nb C6",

            "Nb BDD", "Capa C6", "Capa BDD", "Statut", "Détail",

            "Boîtier C6", "BPE Type", "Boîtier Statut"

        ]

        for col, header in enumerate(detail_headers, 1):

            cell = ws_detail.cell(row=1, column=col, value=header)

            cell.font = header_font

            cell.fill = header_fill

        

        boitier_err_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        

        detail_row = 2

        for stat in stats:

            etude = stat['etude']

            for appui in stat.get('detail', []):

                capas_c6 = appui.get('capas_c6', [])

                capas_bdd = appui.get('capas_bdd', [])

                

                ws_detail.cell(row=detail_row, column=1, value=etude)

                ws_detail.cell(row=detail_row, column=2, value=appui['num_appui'])

                ws_detail.cell(row=detail_row, column=3, value=appui.get('cables_c6', ''))

                ws_detail.cell(row=detail_row, column=4, value=appui['nb_cables_c6'])

                ws_detail.cell(row=detail_row, column=5, value=appui['nb_cables_bdd'])

                ws_detail.cell(row=detail_row, column=6, value='+'.join(str(c) for c in capas_c6) if capas_c6 else '')

                ws_detail.cell(row=detail_row, column=7, value='+'.join(str(c) for c in capas_bdd) if capas_bdd else '')

                ws_detail.cell(row=detail_row, column=8, value=appui['statut'])

                ws_detail.cell(row=detail_row, column=9, value=appui.get('message', ''))

                ws_detail.cell(row=detail_row, column=10, value=appui.get('boitier_c6', ''))

                ws_detail.cell(row=detail_row, column=11, value=appui.get('bpe_noe_type', ''))

                ws_detail.cell(row=detail_row, column=12, value=appui.get('boitier_statut', ''))

                

                # Coloriser selon statut

                if appui['statut'] == 'OK':

                    fill = ok_fill

                elif appui['statut'].startswith('ABSENT'):

                    fill = absent_fill

                else:

                    fill = ecart_fill

                

                for c in range(1, 10):

                    ws_detail.cell(row=detail_row, column=c).fill = fill

                

                # Coloriser colonnes boîtier si erreur

                if appui.get('boitier_statut') == 'ERREUR':

                    for c in range(10, 13):

                        ws_detail.cell(row=detail_row, column=c).fill = boitier_err_fill

                elif appui.get('boitier_statut') == 'OK':

                    for c in range(10, 13):

                        ws_detail.cell(row=detail_row, column=c).fill = ok_fill

                

                detail_row += 1

        

        # Ajuster largeurs

        ws_detail.column_dimensions['A'].width = 35

        ws_detail.column_dimensions['B'].width = 15

        ws_detail.column_dimensions['C'].width = 40

        ws_detail.column_dimensions['D'].width = 10

        ws_detail.column_dimensions['E'].width = 10

        ws_detail.column_dimensions['F'].width = 18

        ws_detail.column_dimensions['G'].width = 18

        ws_detail.column_dimensions['H'].width = 15

        ws_detail.column_dimensions['I'].width = 55

        ws_detail.column_dimensions['J'].width = 12

        ws_detail.column_dimensions['K'].width = 15

        ws_detail.column_dimensions['L'].width = 15

        

        try:

            wb.save(filename)

        except (PermissionError, OSError) as _e:

            QgsMessageLog.logMessage(

                f"[POLICE_C6] Impossible d'ecrire le rapport: {filename} \u2014 {_e}",

                "PoleAerien", MSG_CRITICAL

            )

            raise

        return filename





class GespotC6Task(AsyncTaskBase):
    """Async task for GESPOT vs C6 comparison.

    Zero QGIS API calls - pur Python fichier-fichier.
    Paramètres requis : gespot_dir, c6_dir, export_dir.
    """

    def __init__(self, params: dict):
        super().__init__('GESPOT vs C6', params)

    def execute(self) -> bool:
        from .gespot_c6_comparator import run_comparison

        gespot_dir = self.params['gespot_dir']
        c6_dir = self.params['c6_dir']
        export_dir = self.params['export_dir']

        def _prog(pct, msg=''):
            if self.isCanceled():
                return
            self.emit_progress(pct)
            if msg:
                self.emit_message(msg, 'grey')

        result = run_comparison(gespot_dir, c6_dir, export_dir,
                                progress_cb=_prog)

        if self.isCanceled():
            return False

        nb_ok = sum(1 for c in result.comparisons if c.statut_global == 'OK')
        nb_ecart = sum(1 for c in result.comparisons if c.statut_global == 'KO')
        self.emit_message(
            f"GESPOT vs C6 : {len(result.comparisons)} compares "
            f"({nb_ok} OK, {nb_ecart} KO), "
            f"{len(result.absent_c6)} absents C6, "
            f"{len(result.absent_gespot)} absents GESPOT, "
            f"{len(result.anomalies)} anomalies source",
            'blue',
        )

        self.result = {
            'success': True,
            'nb_compares': len(result.comparisons),
            'nb_ok': nb_ok,
            'nb_ecart': nb_ecart,
            'nb_absent_c6': len(result.absent_c6),
            'nb_absent_gespot': len(result.absent_gespot),
            'nb_anomalies': len(result.anomalies),
            'comparisons': result.comparisons,
            'absent_c6': result.absent_c6,
            'absent_gespot': result.absent_gespot,
            'anomalies': result.anomalies,
            'output_path': result.output_path,
        }
        return True


def _run_one_study(
    etude_name, c6_file,
    cables_par_appui_cached, appuis_data,
    bpe_geoms, attaches_parsed,
    cables, p_src_label
):
    """P-04: Process one Police C6 study. Thread-safe (own PoliceC6 instance).

    Returns dict with keys: stats, anomaly_cables, messages, error.
    Messages are collected (not emitted) so caller emits them in order.
    """
    from .PoliceC6 import PoliceC6
    from .cable_analyzer import collect_anomaly_cables

    messages = []
    police = PoliceC6()

    try:
        donnees_c6, _, boitier_c6 = police.lire_annexe_c6(c6_file)
        if not donnees_c6:
            messages.append((f"  [!] {etude_name}: C6 vide ou illisible", "orange"))
            return {'stats': None, 'anomaly_cables': [], 'messages': messages}

        total_cables_c6 = sum(len(v) for v in donnees_c6.values())
        nb_boitier_c6 = len(boitier_c6)
        messages.append((
            f"  {etude_name}: {len(donnees_c6)} appuis, "
            f"{total_cables_c6} cables, {nb_boitier_c6} boitiers", "grey"
        ))

        from .cable_analyzer import verifier_boitiers
        verif_boitier = verifier_boitiers(
            boitier_c6, appuis_data, bpe_geoms, attaches_parsed=attaches_parsed
        )

        comparaison = police.comparer_c6_cables(
            donnees_c6, cables_par_appui_cached, ref_label=p_src_label
        )

        for entry in comparaison:
            num = entry['num_appui']
            bv = verif_boitier.get(num, {})
            entry['boitier_c6'] = bv.get('boitier_source', '')
            entry['bpe_trouve'] = bv.get('bpe_trouve', False)
            entry['bpe_noe_type'] = bv.get('bpe_noe_type', '')
            entry['boitier_statut'] = bv.get('statut', '')

        nb_ok = sum(1 for c in comparaison if c['statut'] == 'OK')
        nb_ecart = sum(1 for c in comparaison if c['statut'] == 'ECART')
        nb_absent = sum(1 for c in comparaison if c['statut'].startswith('ABSENT'))
        nb_boitier_err = sum(1 for c in comparaison if c.get('boitier_statut') == 'ERREUR')

        messages.append((
            f"  {etude_name}: {nb_ok} OK, {nb_ecart} ecarts, "
            f"{nb_absent} absents {p_src_label}"
            + (f", {nb_boitier_err} boitier(s) absent(s)" if nb_boitier_err else ""),
            "green" if nb_ecart == 0 and nb_absent == 0 and nb_boitier_err == 0 else "orange"
        ))

        for c in comparaison:
            if c['statut'] != 'OK':
                tag = "ABSENT" if c['statut'].startswith('ABSENT') else c['statut']
                messages.append((
                    f"    [{tag}] Appui {c['num_appui']}: "
                    f"C6={c['nb_cables_c6']} cables, {p_src_label}={c['nb_cables_bdd']} | {c['message']}",
                    "red" if c['statut'].startswith('ABSENT') else "orange"
                ))
            if c.get('boitier_statut') == 'ERREUR':
                messages.append((
                    f"    [BOITIER] Appui {c['num_appui']}: "
                    f"C6={c['boitier_c6']} mais aucun BPE trouve dans rayon 1m", "red"
                ))

        stats = {
            'etude': etude_name, 'appuis_c6': len(donnees_c6),
            'nb_ok': nb_ok, 'nb_ecart': nb_ecart,
            'nb_absent': nb_absent, 'nb_boitier_err': nb_boitier_err,
            'detail': comparaison
        }

        anom = []
        if nb_ecart > 0:
            anom = collect_anomaly_cables(
                comparaison, cables_par_appui_cached, cables, etude=etude_name
            )

        return {'stats': stats, 'anomaly_cables': anom, 'messages': messages}

    except Exception as e:
        messages.append((f"  [X] {etude_name}: Erreur: {e}", "red"))
        return {'stats': None, 'anomaly_cables': [], 'messages': messages, 'error': str(e)}


def run_async_task(task):

    """Submit task to QGIS task manager"""

    QgsApplication.taskManager().addTask(task)

    return task

