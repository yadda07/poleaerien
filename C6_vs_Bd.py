#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
C6_vs_Bd.py - Comparaison des données C6 par rapport à la base de données.

Fonctionnalités:
1. Lecture des fichiers Excel C6 (détection dynamique feuille/colonne)
2. Liste des poteaux FT couverts par polygones CAP FT (IN)
3. Liste des poteaux FT NON couverts par aucun polygone (OUT)
4. Vérification que les noms d'études dans CAP FT existent dans le répertoire C6
5. Auto-détection du champ étude dans la couche CAP FT
"""

from qgis.core import Qgis, QgsProject, QgsFeatureRequest, QgsExpression, NULL, QgsMessageLog, QgsSpatialIndex
import os
import re
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from .core_utils import normalize_appui_num


# Patterns pour auto-détection du champ étude
ETUDE_FIELD_PATTERNS = [
    r'^nom[_\s]?etude[s]?$',
    r'^etude[s]?$',
    r'^name$',
    r'^nom$',
    r'^decoupage$',
    r'^zone$',
]


class C6_vs_Bd:
    """Comparaison des données C6 par rapport à la base de données."""

    def __init__(self):
        """Constructeur."""
        pass

    def detect_etude_field(self, layer):
        """
        Auto-détecte le champ étude dans une couche.
        
        Args:
            layer: QgsVectorLayer (couche CAP FT)
            
        Returns:
            str: Nom du champ détecté ou None si non trouvé
        """
        if not layer or not layer.isValid():
            return None
        
        field_names = [f.name() for f in layer.fields()]
        
        for pattern in ETUDE_FIELD_PATTERNS:
            for field_name in field_names:
                if re.match(pattern, field_name.lower()):
                    QgsMessageLog.logMessage(
                        f"[C6_vs_Bd] Champ étude détecté: {field_name}",
                        "PoleAerien", Qgis.Info
                    )
                    return field_name
        
        QgsMessageLog.logMessage(
            f"[C6_vs_Bd] Aucun champ étude détecté. Champs disponibles: {field_names}",
            "PoleAerien", Qgis.Warning
        )
        return None

    def LectureFichiersExcelsC6(self, df, repertoire):
        """
        Parcourir les fichiers Excel C6 pour extraire les numéros d'appui.
        
        Ignore automatiquement les fichiers non-C6:
        - FicheAppui_*.xlsx (fiches individuelles)
        - *_C7*.xlsx, *Annexe C7*.xlsx (fichiers C7)
        - GESPOT_*.xlsx (exports GESPOT)
        """
        # Patterns de fichiers à ignorer (non-C6)
        ignore_patterns = [
            'ficheappui',
            'gespot',
            '_c7',
            'annexe c7',
            'annexe_c7',
        ]

        for subdir, _, files in os.walk(repertoire):
            for name in files:
                if not name.endswith('.xlsx') or "~$" in name or name.startswith("ANALYSE_"):
                    continue
                
                # Ignorer fichiers non-C6
                name_lower = name.lower()
                if any(pattern in name_lower for pattern in ignore_patterns):
                    continue

                cheminComplet = os.path.join(subdir, name)

                try:
                    with pd.ExcelFile(cheminComplet) as xls:
                        sheet_names = xls.sheet_names
                        target_sheet = None
                        header_row = 7  # Par défaut ligne 8 (0-indexed = 7)
                        
                        # Essayer plusieurs noms de feuilles courants pour C6
                        for candidate in ["Export 1", "Export1", "Saisies terrain"]:
                            if candidate in sheet_names:
                                target_sheet = candidate
                                if candidate == "Saisies terrain":
                                    header_row = 24
                                break
                        
                        # Si aucune feuille connue, prendre la première
                        if target_sheet is None and sheet_names:
                            target_sheet = sheet_names[0]
                        
                        if target_sheet is None:
                            continue  # Fichier vide, ignorer silencieusement
                        
                        # Vérifier que le fichier a assez de lignes
                        df_check = pd.read_excel(xls, target_sheet, header=None, nrows=header_row + 2)
                        if len(df_check) < header_row + 1:
                            continue  # Fichier trop court, pas un C6
                        
                        df1 = pd.read_excel(xls, target_sheet, header=header_row, index_col=None)
                    
                    if df1.empty:
                        continue
                    
                    # Détecter colonne N° appui
                    col_appui = None
                    for col in df1.columns:
                        col_norm = ''.join(c for c in str(col).lower() if c.isalnum())
                        if 'appui' in col_norm and 'nappui' in col_norm or col_norm == 'nappui':
                            col_appui = col
                            break
                        if 'appui' in col_norm:
                            col_appui = col
                    
                    if col_appui is None:
                        # Pas un fichier C6, ignorer silencieusement
                        continue
                    
                    # Renommer pour uniformiser
                    df1 = df1.rename(columns={col_appui: 'N° appui'})
                    
                    # Sélectionner colonnes utiles
                    cols_to_keep = ['N° appui']
                    if 'Nature des travaux' in df1.columns:
                        cols_to_keep.append('Nature des travaux')
                    
                    resultat = df1[cols_to_keep].dropna(thresh=1)
                    resultat = resultat[pd.to_numeric(resultat['N° appui'], errors='coerce').notnull()]
                    
                    if resultat.empty:
                        continue
                    
                    resultat['N° appui'] = pd.to_numeric(resultat['N° appui'], errors="coerce").astype("Int64")

                    # Ajouter nom du fichier
                    resultat = resultat.copy()
                    resultat.insert(2, "Excel", name, True)
                    resultat.insert(3, "Études", name.replace(".xlsx", ""), True)
                    df = pd.concat([df, resultat], ignore_index=True)

                except Exception as err:
                    # Log uniquement les erreurs inattendues (pas les fichiers non-C6)
                    if "index" not in str(err) and "Colonne" not in str(err):
                        QgsMessageLog.logMessage(
                            f"[C6_vs_Bd.LectureFichiersExcelsC6] {name}: {err}",
                            "PoleAerien", Qgis.Warning
                        )
                    continue

        return df

    def liste_poteau_cap_ft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):
        """
        Récupérer les poteaux FT couverts par les polygones CAP FT.
        
        Optimisé avec index spatial pour éviter O(n*m).
        """
        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        # Index spatial des polygones CAP FT
        cap_ft_index = QgsSpatialIndex(etude_cap_ft.getFeatures())
        cap_ft_cache = {}
        for feat in etude_cap_ft.getFeatures():
            cap_ft_cache[feat.id()] = {
                'geom': feat.geometry(),
                'etude': feat[colonne_cap_ft]
            }

        listePoteaux = []
        listePoteauxComplet = []
        fichiers = []
        listeEtat = []
        nature_travaux = []

        # Filtrer poteaux FT
        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)

        for feat_pot in infra_pt_pot.getFeatures(request):
            pt_geom = feat_pot.geometry()
            if not pt_geom or pt_geom.isEmpty():
                continue

            # Chercher polygones candidats via index spatial
            candidate_ids = cap_ft_index.intersects(pt_geom.boundingBox())
            
            for cid in candidate_ids:
                cap_data = cap_ft_cache.get(cid)
                if cap_data and cap_data['geom'].contains(pt_geom):
                    inf_num = feat_pot["inf_num"]
                    etudes = cap_data['etude']

                    etat = "" if feat_pot["etat"] == NULL else feat_pot["etat"]
                    nature = "Recalage" if etat == "A RECALER" else ("Remplacement" if etat == "A REMPLACER" else etat)

                    # Normaliser le numéro d'appui
                    num_appui = normalize_appui_num(inf_num)
                    if not num_appui:
                        continue

                    listePoteaux.append(num_appui)
                    listePoteauxComplet.append(inf_num)
                    fichiers.append(etudes)
                    listeEtat.append(etat)
                    nature_travaux.append(nature)
                    break  # Un seul polygone par poteau

        df = pd.DataFrame({
            'N° appui': listePoteaux,
            "Études": fichiers,
            "Nature des travaux": nature_travaux,
            'inf_num (QGIS)': listePoteauxComplet,
            "Etat": listeEtat
        })

        return df

    def liste_poteaux_ft_out(self, table_poteau, table_etude_cap_ft):
        """
        Liste les poteaux FT NON couverts par aucun polygone CAP FT.
        
        Args:
            table_poteau: Nom de la couche poteaux (infra_pt_pot)
            table_etude_cap_ft: Nom de la couche polygones CAP FT
            
        Returns:
            pd.DataFrame: Poteaux FT hors périmètre avec colonnes [N° appui, inf_num, etat]
        """
        infra_pt_pot = QgsProject.instance().mapLayersByName(table_poteau)[0]
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        # Index spatial des polygones CAP FT pour performance
        cap_ft_index = QgsSpatialIndex(etude_cap_ft.getFeatures())
        cap_ft_geoms = {f.id(): f.geometry() for f in etude_cap_ft.getFeatures()}

        poteaux_out = []
        requete = QgsExpression("inf_type LIKE 'POT-FT'")
        request = QgsFeatureRequest(requete)

        for feat_pot in infra_pt_pot.getFeatures(request):
            pt_geom = feat_pot.geometry()
            if not pt_geom or pt_geom.isEmpty():
                continue

            # Chercher les polygones candidats via l'index spatial
            candidate_ids = cap_ft_index.intersects(pt_geom.boundingBox())
            is_covered = False

            for cid in candidate_ids:
                if cap_ft_geoms[cid].contains(pt_geom):
                    is_covered = True
                    break

            if not is_covered:
                inf_num = feat_pot["inf_num"]
                etat = "" if feat_pot["etat"] == NULL else str(feat_pot["etat"])
                num_appui = normalize_appui_num(inf_num)
                if num_appui:
                    poteaux_out.append({
                        'N° appui': num_appui,
                        'inf_num': inf_num,
                        'etat': etat
                    })

        df_out = pd.DataFrame(poteaux_out)
        QgsMessageLog.logMessage(
            f"[C6_vs_Bd] Poteaux FT hors périmètre: {len(df_out)}",
            "PoleAerien", Qgis.Info
        )
        return df_out

    def verifier_etudes_c6(self, table_etude_cap_ft, colonne_cap_ft, repertoire_c6):
        """
        Vérifie que chaque nom d'étude dans CAP FT a un fichier Excel correspondant.
        
        Args:
            table_etude_cap_ft: Nom de la couche polygones CAP FT
            colonne_cap_ft: Nom de la colonne contenant les noms d'études
            repertoire_c6: Chemin du répertoire contenant les fichiers C6
            
        Returns:
            dict: {
                'etudes_capft': set des noms d'études dans CAP FT,
                'fichiers_c6': set des noms de fichiers Excel (sans extension),
                'etudes_sans_c6': list des études sans fichier C6,
                'c6_sans_etude': list des fichiers C6 sans étude correspondante
            }
        """
        etude_cap_ft = QgsProject.instance().mapLayersByName(table_etude_cap_ft)[0]

        # Récupérer les noms d'études uniques dans CAP FT
        etudes_capft = set()
        for feat in etude_cap_ft.getFeatures():
            val = feat[colonne_cap_ft]
            if val and val != NULL:
                etudes_capft.add(str(val).strip())

        # Récupérer les noms de fichiers Excel dans le répertoire C6
        fichiers_c6 = set()
        for subdir, _, files in os.walk(repertoire_c6):
            for name in files:
                if name.endswith('.xlsx') and "~$" not in name and not name.startswith("ANALYSE_"):
                    fichiers_c6.add(name.replace(".xlsx", ""))

        # Comparaison
        etudes_sans_c6 = [e for e in etudes_capft if e not in fichiers_c6]
        c6_sans_etude = [f for f in fichiers_c6 if f not in etudes_capft]

        result = {
            'etudes_capft': etudes_capft,
            'fichiers_c6': fichiers_c6,
            'etudes_sans_c6': etudes_sans_c6,
            'c6_sans_etude': c6_sans_etude
        }

        QgsMessageLog.logMessage(
            f"[C6_vs_Bd] Études CAP FT: {len(etudes_capft)}, Fichiers C6: {len(fichiers_c6)}, "
            f"Études sans C6: {len(etudes_sans_c6)}, C6 sans étude: {len(c6_sans_etude)}",
            "PoleAerien", Qgis.Info
        )

        return result

    def ecrictureExcel(self, final, fichier, poteaux_out=None, verif_etudes=None):
        """
        Ecriture du résultat final dans un fichier Excel multi-feuilles.
        
        Args:
            final: DataFrame principal (analyse C6 vs BD)
            fichier: Chemin du fichier Excel de sortie
            poteaux_out: DataFrame des poteaux FT hors périmètre (optionnel)
            verif_etudes: dict résultat de verifier_etudes_c6 (optionnel)
            
        Returns:
            int: Nombre d'erreurs (lignes avec ABSENT)
        """
        nb_erreurs = 0
        
        with pd.ExcelWriter(fichier, engine="openpyxl") as writer:
            # Feuille 1: Analyse principale C6 BD
            sheet_name = "ANALYSE C6 BD"
            final.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet = writer.sheets[sheet_name]
            
            # Ajuster largeur colonnes
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                sheet.column_dimensions[col_letter].width = 20
            
            # Colorer les lignes avec statut ABSENT
            fill_orange = PatternFill("solid", start_color="feb24c")
            if "Statut" in final.columns:
                for idx, row in enumerate(final.itertuples(), start=2):
                    statut = str(getattr(row, 'Statut', ''))
                    if "ABSENT" in statut:
                        for cell in sheet[idx]:
                            cell.fill = fill_orange
                        nb_erreurs += 1

            # Feuille 2: Poteaux FT hors périmètre
            if poteaux_out is not None and not poteaux_out.empty:
                sheet_out = "POTEAUX HORS PERIMETRE"
                poteaux_out.to_excel(writer, sheet_name=sheet_out, index=False)
                ws_out = writer.sheets[sheet_out]
                for alpha in ["A", "B", "C"]:
                    ws_out.column_dimensions[alpha].width = 20
                fill_red = PatternFill("solid", start_color="ff6b6b")
                for row in ws_out.iter_rows(min_row=2, max_row=len(poteaux_out) + 1):
                    for cell in row:
                        cell.fill = fill_red

            # Feuille 3: Vérification études vs C6
            if verif_etudes:
                sheet_verif = "VERIF ETUDES"
                etudes_sans_c6 = verif_etudes.get('etudes_sans_c6', [])
                c6_sans_etude = verif_etudes.get('c6_sans_etude', [])
                max_len = max(len(etudes_sans_c6), len(c6_sans_etude), 1)
                
                df_verif = pd.DataFrame({
                    'Études CAP FT sans fichier C6': etudes_sans_c6 + [''] * (max_len - len(etudes_sans_c6)),
                    'Fichiers C6 sans étude CAP FT': c6_sans_etude + [''] * (max_len - len(c6_sans_etude))
                })
                
                df_verif.to_excel(writer, sheet_name=sheet_verif, index=False)
                ws_verif = writer.sheets[sheet_verif]
                ws_verif.column_dimensions["A"].width = 40
                ws_verif.column_dimensions["B"].width = 40
                
                fill_warn = PatternFill("solid", start_color="feb24c")
                for row in ws_verif.iter_rows(min_row=2, max_row=max_len + 1):
                    for cell in row:
                        if cell.value:
                            cell.fill = fill_warn

        return nb_erreurs