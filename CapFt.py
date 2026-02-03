#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
import openpyxl
from .qgis_utils import (
    verifications_donnees_etude,
    liste_poteaux_par_etude,
    normalize_appui_num_bt
)


class CapFt:
    """ Second pour contenir quelques fonctions qui seront utilisées dans le fichier principale CheckOptyce.py """

    def __init__(self):
        """Le constructeur de ma classe
        Il prend pour attribut de classe les *** """

    def verificationsDonneesCapft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):
        """Vérifie doublons études + poteaux FT hors étude."""
        return verifications_donnees_etude(
            table_poteau, table_etude_cap_ft, colonne_cap_ft,
            'POT-FT', 'CAP_FT'
        )

    def liste_poteau_cap_ft(self, table_poteau, table_etude_cap_ft, colonne_cap_ft):
        """Liste poteaux FT par étude avec détection terrains privés."""
        return liste_poteaux_par_etude(
            table_poteau, table_etude_cap_ft, colonne_cap_ft,
            'POT-FT', 'CAP_FT'
        )

    def LectureFichiersExcelsCap_ft(self, repertoire):
        """Fonction pour parcourir les fichiers Excel pour renseigner la référence des appuis."""
        dicoPoteauFt_SousTraitant = {}

        for subdir, _, files in os.walk(repertoire):
            for name in files:
                if "FicheAppui_" in name and name.endswith('.xlsx'):
                    cheminComplet = os.path.join(subdir, name)

                    # Récupère le dossier qui contient directement le fichier
                    dossier_parent = os.path.basename(os.path.dirname(cheminComplet))

                    # Ajoute le fichier au dictionnaire
                    if dossier_parent not in dicoPoteauFt_SousTraitant:
                        dicoPoteauFt_SousTraitant[dossier_parent] = []

                    dicoPoteauFt_SousTraitant[dossier_parent].append(name)

        return dicoPoteauFt_SousTraitant

    def traitementResultatFinauxCapFt(self, dicoEtudeCapFtPoteauQgis, dicoPoteauFt_SousTraitant):
        dicoPotFt_Excel_Introuvable = {}
        dicoPotFtExistants = {}
        comptabilite = 0

        # Étape 1 : index rapide
        index_qgis = {}  # clé normalisée -> liste de (etude_cap_ft, inf_num_ft)
        for etude_cap_ft, listePoteauxQgis in dicoEtudeCapFtPoteauQgis.items():
            for inf_num_ft in listePoteauxQgis:
                cle = normalize_appui_num_bt(inf_num_ft, strip_e_prefix=True)
                index_qgis.setdefault(cle, []).append((etude_cap_ft, inf_num_ft))

        # Étape 2 : traitement Excel
        for excel, listePoteau in dicoPoteauFt_SousTraitant.items():
            PotFtintrouvableSt = []

            for poteauSt in listePoteau:
                potSt = poteauSt.replace("FicheAppui_", "").replace(".xlsx", "").strip()
                potSt = normalize_appui_num_bt(potSt, strip_e_prefix=True)

                infos_list = index_qgis.get(potSt)
                if infos_list:
                    # Pop la première correspondance
                    etude_cap_ft, inf_num_ft = infos_list.pop(0)
                    comptabilite += 1
                    dicoPotFtExistants[comptabilite] = [inf_num_ft, etude_cap_ft, poteauSt, excel]

                    # Supprimer le poteau traité de dicoEtudeCapFtPoteauQgis
                    if etude_cap_ft in dicoEtudeCapFtPoteauQgis:
                        if inf_num_ft in dicoEtudeCapFtPoteauQgis[etude_cap_ft]:
                            dicoEtudeCapFtPoteauQgis[etude_cap_ft].remove(inf_num_ft)
                        # Si plus de poteaux restants pour cette étude, supprimer la clé
                        if not dicoEtudeCapFtPoteauQgis[etude_cap_ft]:
                            dicoEtudeCapFtPoteauQgis.pop(etude_cap_ft)

                    # Si plus d'infos pour cette clé dans l'index, supprimer la clé
                    if not infos_list:
                        index_qgis.pop(potSt, None)
                else:
                    PotFtintrouvableSt.append(poteauSt)

            if PotFtintrouvableSt:
                dicoPotFt_Excel_Introuvable[excel] = PotFtintrouvableSt

        return dicoPotFt_Excel_Introuvable, dicoEtudeCapFtPoteauQgis, dicoPotFtExistants

    def ecrireResultatsAnalyseExcelsCapFt(self, resultatsFinaux, nom):
        """Fonction qui permet d'écrire dans le fichier Excel à partir d'un fichier modèle'"""

        dicoPotFt_Excel_Introuvable = resultatsFinaux[0]
        dicoEtudeCapFtPotQgisIntrouvable = resultatsFinaux[1]
        dicoPotFtExistants = resultatsFinaux[2]

        # print("dicoPotFt_Excel_Introuvable:",dicoPotFt_Excel_Introuvable)
        # print("dicoEtudeCapFtPotQgisIntrouvable:",dicoEtudeCapFtPotQgisIntrouvable)
        # print("dicoPotFtExistants:",dicoPotFtExistants)

        # # Créer un set des inf_num_ft existants pour recherche rapide
        # inf_nums_existants = set(item['inf_num_ft'] for item in dicoPotFtExistants)
        #
        # # Supprimer directement dans la liste originale
        # for item in dicoEtudeCapFtPotQgisIntrouvable[:]:  # on parcourt une copie
        #     if item['inf_num_ft'] in inf_nums_existants:
        #         dicoEtudeCapFtPotQgisIntrouvable.remove(item)
        # # Supprimer directement dans la liste originale
        # for item in dicoPotFt_Excel_Introuvable[:]:  # on parcourt une copie
        #     if item['inf_num_ft'] in inf_nums_existants:
        #         dicoPotFt_Excel_Introuvable.remove(item)
        #
        fichierXlsx = openpyxl.workbook.Workbook()

        ############################## FEUILLE 1 : VALIDATION_COMPLET ################################################
        my_ligne = 1  # Ligne à partir de laquelle on commencera à écrire dans le fichier Annexe
        # Parcourir fichier C3A, pour comparer ses contenus aux fichiers C6
        feuille = fichierXlsx.create_sheet("ANALYSE CAP_FT", 0)

        alignement = openpyxl.styles.Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False,
                                               shrink_to_fit=True, indent=0)

        # Pour mettre du GRAS à partir de la colonne 18 et 23 pour les câbles.
        feuille.cell(row=1, column=1, value="INF_NUM QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=2, value="ETUDE QGIS").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=3, value="INF_NUM EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=4, value="NOM FICHIER EXCEL").font = openpyxl.styles.Font(bold=True)
        feuille.cell(row=1, column=5, value="REMARQUES").font = openpyxl.styles.Font(bold=True)

        feuille.cell(row=1, column=1).alignment = alignement
        feuille.cell(row=1, column=2).alignment = alignement
        feuille.cell(row=1, column=3).alignment = alignement
        feuille.cell(row=1, column=4).alignment = alignement
        feuille.cell(row=1, column=5).alignment = alignement

        feuille.column_dimensions["A"].width = 25
        feuille.column_dimensions["B"].width = 30
        feuille.column_dimensions["C"].width = 25
        feuille.column_dimensions["D"].width = 50
        feuille.column_dimensions["E"].width = 40

        red_color = openpyxl.styles.colors.Color(rgb='fc4e2a')
        orange_color = openpyxl.styles.colors.Color(rgb='fd8d3c')

        ####################### SOUS-TRAITS INTROUVABLE #######################################"
        for fichierExcels, listesDesAppuisSous_traitant in dicoPotFt_Excel_Introuvable.items():
            for inf_num_excel in listesDesAppuisSous_traitant:
                # Si tiragle câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=3, value=inf_num_excel).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=4, value=fichierExcels).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans QGIS").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=red_color)

        ####################### QGIS INTROUVABLE #######################################
        for etudeQgis, listesDesAppuisQGis in dicoEtudeCapFtPotQgisIntrouvable.items():
            for inf_num in listesDesAppuisQGis:
                # Si tirage câble complètement validé
                my_ligne += 1

                feuille.cell(row=my_ligne, column=1, value=inf_num).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=2, value=etudeQgis).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)
                feuille.cell(row=my_ligne, column=5, value="infra inexistant dans les Fiches Appuis").fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=orange_color)

        my_colonne = 1
        my_ligne += 1

        ####################### POTEAUX TROUVEES #######################################")
        for _, listesDesAppuisBT in dicoPotFtExistants.items():
            for valeur in listesDesAppuisBT:
                feuille.cell(row=my_ligne, column=my_colonne, value=valeur)
                my_colonne += 1

            feuille.cell(row=my_ligne, column=5, value="correspondance trouvée")

            # On passe à la ligne suivante
            # Si la première liste est terminé, on passe à la ligne suivante en récommençant par au toujours
            # au niveau de la colonne A
            my_ligne += 1
            my_colonne = 1

        if 'Sheet' in fichierXlsx.sheetnames:
            del fichierXlsx['Sheet']

        fichierXlsx.save(filename=nom)   # Enregistrement du fichier
