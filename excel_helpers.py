# -*- coding: utf-8 -*-
"""
Helpers Excel pour le plugin PoleAerien.
Centralise constantes couleurs et fonctions utilitaires openpyxl.
"""

# Constantes couleurs (référence unique)
COLORS = {
    'red': 'fc4e2a',
    'orange': 'fd8d3c',
    'orange_light': 'feb24c',
    'yellow': 'ffeb9c',
    'green': '90EE90',
    'green_light': 'c6efce',
    'green_dark': 'dcfce7',
}


def get_center_alignment():
    """Retourne alignement centré standard.
    
    Returns:
        openpyxl.styles.Alignment configuré
    """
    from openpyxl.styles import Alignment
    return Alignment(
        horizontal='center',
        vertical='bottom',
        shrink_to_fit=True
    )


def get_fill(color_name):
    """Retourne PatternFill pour une couleur.
    
    Args:
        color_name: Clé COLORS ou code hex direct
        
    Returns:
        openpyxl.styles.PatternFill
    """
    from openpyxl.styles import PatternFill
    from openpyxl.styles.colors import Color
    
    hex_color = COLORS.get(color_name, color_name)
    return PatternFill(fill_type='solid', fgColor=Color(rgb=hex_color))


def set_header_cell(sheet, row, col, value, width=None, bold=True):
    """Configure une cellule d'en-tête.
    
    Args:
        sheet: Feuille openpyxl
        row: Numéro de ligne (1-indexed)
        col: Numéro de colonne (1-indexed)
        value: Valeur de la cellule
        width: Largeur colonne (optionnel)
        bold: Police en gras (défaut True)
        
    Returns:
        Cell configurée
    """
    from openpyxl.styles import Font
    
    cell = sheet.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    cell.alignment = get_center_alignment()
    
    if width is not None:
        sheet.column_dimensions[cell.column_letter].width = width
    
    return cell


def cleanup_default_sheet(workbook):
    """Supprime la feuille 'Sheet' par défaut si présente.
    
    Args:
        workbook: Workbook openpyxl
    """
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']


def set_column_widths(sheet, widths_dict):
    """Définit largeurs de colonnes en batch.
    
    Args:
        sheet: Feuille openpyxl
        widths_dict: Dict {lettre_colonne: largeur}
    """
    for col_letter, width in widths_dict.items():
        sheet.column_dimensions[col_letter].width = width


def apply_fill_to_row(sheet, row, start_col, end_col, fill):
    """Applique un fill à une plage de cellules sur une ligne.
    
    Args:
        sheet: Feuille openpyxl
        row: Numéro de ligne
        start_col: Colonne début (1-indexed)
        end_col: Colonne fin (1-indexed, incluse)
        fill: PatternFill à appliquer
    """
    for col in range(start_col, end_col + 1):
        sheet.cell(row=row, column=col).fill = fill
