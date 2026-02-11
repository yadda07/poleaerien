#!/usr/bin/python
# -*-coding:Utf-8-*-
"""
Police C6 v2.0 - Vérification câbles découpés via fddcpi2

CHANGEMENT MAJEUR: Abandon de GraceTHD au profit de la fonction PostgreSQL fddcpi2
qui retourne des câbles correctement découpés par rapport à l'infrastructure.

Logique:
1. Charger les câbles découpés depuis PostgreSQL (fddcpi2)
2. Pour chaque appui, compter les extrémités de câbles qui le touchent
3. Récupérer les capacités depuis les câbles
4. Comparer avec l'Annexe C6 (nombre de câbles + capacités)
5. Générer les anomalies
"""

from qgis.core import (
    Qgis,
    QgsProject,
    QgsExpression,
    QgsFeatureRequest,
    QgsVectorLayer,
    QgsFeature,
    QgsGeometry,
    QgsPointXY,
    QgsCoordinateReferenceSystem,
    QgsMessageLog,
    QgsSpatialIndex,
    QgsWkbTypes,
    NULL,
    QgsRectangle
)
from qgis.PyQt.QtWidgets import QApplication

from .qgis_utils import (
    remove_group,
    layer_group_error,
    insert_layer_in_group,
    get_layer_safe,
    validate_same_crs,
    normalize_appui_num,
)
from .dataclasses_results import (
    PoliceC6Result, CableCapaciteResult, BoitierValidationResult,
    EtudeC6Result, ParcourAutoC6Result
)
from .db_connection import DatabaseConnection, CableSegment, extract_sro_from_layer
from .cable_analyzer import CableAnalyzer, AppuiChargeResult, extraire_appuis_from_layer
from .security_rules import get_capacite_fo_from_code, get_capacites_possibles

from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
import os
import re
import openpyxl
import pandas as pd


class PoliceC6Cancelled(Exception):
    """Annulation demandée par l'utilisateur."""
    pass


LYR_INFRA_PT_POT = "infra_pt_pot"
LYR_BPE = "bpe"
STYLE_DIR = os.path.join(os.path.dirname(__file__), "styles")


@dataclass
class AnalyseChargeResult:
    """Résultat de l'analyse de charge pour une étude"""
    etude: str
    sro: str = ""
    
    # Stats globales
    total_appuis: int = 0
    appuis_ok: int = 0
    appuis_anomalie: int = 0
    
    # Câbles
    total_cables_bdd: int = 0
    total_cables_c6: int = 0
    
    # Capacités
    total_capa_bdd: int = 0
    total_capa_c6: int = 0
    
    # Détails anomalies
    anomalies: List[AppuiChargeResult] = field(default_factory=list)
    
    # Erreur éventuelle
    erreur: str = ""


class PoliceC6:
    """
    Police C6 v2.0 - Vérification infrastructure via câbles découpés (fddcpi2).
    
    Utilise la fonction PostgreSQL fddcpi2 pour obtenir des câbles correctement
    découpés par rapport à l'infrastructure, puis compare avec l'Annexe C6.
    """

    def __init__(self):
        """Initialise les attributs de classe."""
        self._cancel_requested = False
        self.db_connection = DatabaseConnection()
        self.cable_analyzer = CableAnalyzer(tolerance=0.5)
        self._reset_state()

    def request_cancel(self):
        """Demande d'annulation."""
        self._cancel_requested = True

    def _check_cancel(self):
        """Vérifie si annulation demandée."""
        QApplication.processEvents()
        if self._cancel_requested:
            raise PoliceC6Cancelled()

    def _reset_state(self):
        """Réinitialise l'état interne pour une nouvelle analyse."""
        self._cancel_requested = False
        
        # Stats appuis
        self.nb_appui_corresp = 0
        self.nb_appui_absent = 0
        self.nb_appui_absentPot = 0
        self.absence = []
        self.idPotAbsent = []
        
        # Stats BPE
        self.nb_pbo_corresp = 0
        self.ebp_non_appui = []
        self.ebp_appui_inconnu = []
        
        # Stats câbles (nouvelle logique)
        self.resultats_charge: Dict[str, AppuiChargeResult] = {}
        self.anomalies_cables: List[AppuiChargeResult] = []
        
        # Données C6
        self.donnees_c6: Dict[str, Dict] = {}  # num_appui -> {cables, capacites}

    # =========================================================================
    # CONNEXION BDD
    # =========================================================================
    def connect_database(self) -> bool:
        """Établit la connexion à la base PostgreSQL."""
        return self.db_connection.connect()

    def disconnect_database(self):
        """Ferme la connexion."""
        self.db_connection.disconnect()

    # =========================================================================
    # EXTRACTION SRO
    # =========================================================================
    def extraire_sro(self, layer: QgsVectorLayer = None) -> Optional[str]:
        """
        Extrait le code SRO depuis une couche QGIS.
        Si aucune couche fournie, cherche infra_pt_pot ou cables.
        """
        if layer:
            return extract_sro_from_layer(layer)
        
        # Chercher dans les couches du projet
        for layer_name in ['infra_pt_pot', 'cables', 't_cable']:
            layers = QgsProject.instance().mapLayersByName(layer_name)
            if layers:
                sro = extract_sro_from_layer(layers[0])
                if sro:
                    return sro
        
        return None

    # =========================================================================
    # ANALYSE PRINCIPALE - NOUVELLE LOGIQUE
    # =========================================================================
    def analyser_charge_cables(
        self,
        sro: str,
        layer_appuis: QgsVectorLayer,
        donnees_c6: Dict[str, str],
        only_aerien: bool = True
    ) -> AnalyseChargeResult:
        """
        Analyse la charge de câbles par appui via fddcpi2.
        
        Args:
            sro: Code SRO (ex: '63041/B1I/PMZ/00003')
            layer_appuis: Couche des poteaux (infra_pt_pot)
            donnees_c6: Dict[num_appui, nom_cable] depuis l'Annexe C6
            only_aerien: Si True, ne compte que les câbles aériens
        
        Returns:
            AnalyseChargeResult
        """
        result = AnalyseChargeResult(etude="", sro=sro)
        
        try:
            self._check_cancel()
            
            # 1. Charger les câbles découpés depuis PostgreSQL
            QgsMessageLog.logMessage(
                f"Chargement câbles découpés pour SRO: {sro}",
                "POLICE_C6", Qgis.Info
            )
            
            if not self.db_connection.connection:
                if not self.connect_database():
                    result.erreur = "Impossible de se connecter à la base PostgreSQL"
                    return result
            
            cables_decoupes = self.db_connection.execute_fddcpi2(sro)
            
            if not cables_decoupes:
                result.erreur = f"Aucun câble trouvé pour SRO: {sro}"
                return result
            
            self._check_cancel()
            
            # 2. Extraire les appuis depuis la couche QGIS
            appuis = extraire_appuis_from_layer(layer_appuis)
            
            if not appuis:
                result.erreur = "Aucun appui trouvé dans la couche"
                return result
            
            self._check_cancel()
            
            # 3. Analyser la charge par appui
            self.cable_analyzer.analyser_charge_appuis(
                appuis=appuis,
                cables_decoupes=cables_decoupes,
                only_aerien=only_aerien
            )
            
            self._check_cancel()
            
            # 4. Enrichir avec les données C6
            self.cable_analyzer.enrichir_avec_c6(donnees_c6)
            
            # 5. Collecter les résultats
            self.resultats_charge = self.cable_analyzer.resultats
            self.anomalies_cables = self.cable_analyzer.get_anomalies()
            
            stats = self.cable_analyzer.get_stats_globales()
            result.total_appuis = stats['total_appuis']
            result.appuis_ok = stats['appuis_ok']
            result.appuis_anomalie = stats['appuis_anomalie']
            result.total_cables_bdd = stats['total_cables_bdd']
            result.total_cables_c6 = stats['total_cables_c6']
            result.total_capa_bdd = stats['total_capa_bdd']
            result.total_capa_c6 = stats['total_capa_c6']
            result.anomalies = self.anomalies_cables
            
            QgsMessageLog.logMessage(
                f"Analyse terminée: {result.appuis_ok} OK, {result.appuis_anomalie} anomalies",
                "POLICE_C6", Qgis.Info
            )
            
            return result
            
        except PoliceC6Cancelled:
            result.erreur = "Analyse annulée"
            raise
        except Exception as e:
            result.erreur = str(e)
            QgsMessageLog.logMessage(
                f"Erreur analyse charge: {e}",
                "POLICE_C6", Qgis.Warning
            )
            return result

    # =========================================================================
    # LECTURE ANNEXE C6
    # =========================================================================
    def lire_annexe_c6(self, chemin_c6: str) -> Tuple[Dict[str, List[str]], List[Dict], Dict[str, str]]:
        """
        Lit l'Annexe C6 et extrait les références câbles par appui.
        
        STRUCTURE EXCEL C6:
        Un appui peut avoir PLUSIEURS lignes Excel, une par câble.
        La colonne num_appui peut être vide sur les lignes de continuation
        (cellules fusionnées ou répétition implicite).
        
        Args:
            chemin_c6: Chemin vers le fichier Excel Annexe C6
        
        Returns:
            Tuple[donnees_par_appui, liste_brute, boitier_par_appui]
            - donnees_par_appui: Dict[num_appui, List[nom_cable]]
            - liste_brute: Liste de dicts avec toutes les colonnes
            - boitier_par_appui: Dict[num_appui, str] (PB/PEO ou vide)
        """
        donnees_par_appui = {}
        liste_brute = []
        boitier_par_appui = {}
        
        if not os.path.exists(chemin_c6):
            QgsMessageLog.logMessage(
                f"Fichier C6 introuvable: {chemin_c6}",
                "POLICE_C6", Qgis.Warning
            )
            return donnees_par_appui, liste_brute, boitier_par_appui
        
        try:
            wb = openpyxl.load_workbook(chemin_c6, data_only=True)
            
            sheet = None
            for name in ['Export 1', 'export 1', 'Appui', 'APPUI', 'appui', 'Appuis']:
                if name in wb.sheetnames:
                    sheet = wb[name]
                    break
            
            if not sheet:
                sheet = wb.active
            
            # Scanner les lignes pour trouver la VRAIE ligne d'en-têtes
            # (la ligne 1 est souvent des métadonnées: n° commande, date, etc.)
            header_row_idx = None
            headers = []
            col_num_appui = -1
            col_nom_cable = -1
            col_effort_dispo = -1
            col_pose_boitier = -1
            
            appui_candidates = [
                'n° appui', 'n°appui', 'num_appui', 'numero_appui',
                'numappui', 'pt_ad_numsu', 'num appui'
            ]
            cable_candidates = [
                'nom du câble', 'nom du cable', 'nom_cable', 'nomcable',
                'nom_du_cable', 'nom_du_câble', 'cable', 'câble'
            ]
            effort_candidates = [
                'effort disponible avant ajout câble',
                'effort disponible avant ajout cable',
                'effort disponible avant ajout',
                'effort dispo avant ajout',
                'effort disponible'
            ]
            boitier_candidates = [
                "pose d'un boitier optique", "pose d'un boîtier optique",
                "pose d'un boitier", "pose d'un boîtier",
                'pose boitier optique', 'pose boîtier optique',
                'pose boitier', 'pose boîtier',
                'pose_boitier', 'pose_boîtier'
            ]
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                if not row:
                    continue
                row_headers = [str(cell or '').strip().lower() for cell in row]
                
                test_appui = self._find_column_index(row_headers, appui_candidates)
                test_cable = self._find_column_index(row_headers, cable_candidates)
                
                # On a trouvé la ligne d'en-têtes si les 2 colonnes sont trouvées
                # ET qu'elles pointent vers des colonnes différentes
                if test_appui >= 0 and test_cable >= 0 and test_appui != test_cable:
                    header_row_idx = row_idx
                    headers = row_headers
                    col_num_appui = test_appui
                    col_nom_cable = test_cable
                    col_effort_dispo = self._find_column_index(row_headers, effort_candidates)
                    col_pose_boitier = self._find_column_index(row_headers, boitier_candidates)
                    break
            
            if header_row_idx is None or col_num_appui == -1:
                # Log toutes les premières lignes pour debug
                debug_rows = []
                for r_idx, r in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
                    if r:
                        debug_rows.append(f"  Row{r_idx}: {[str(c or '')[:30] for c in r[:5]]}")
                QgsMessageLog.logMessage(
                    f"En-têtes C6 non trouvées. Premières lignes:\n" + "\n".join(debug_rows),
                    "POLICE_C6", Qgis.Warning
                )
                wb.close()
                return donnees_par_appui, liste_brute, boitier_par_appui
            
            current_appui = None
            appuis_edf = set()
            
            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1
            ):
                if not row or all(v is None for v in row):
                    continue
                
                raw_num = str(row[col_num_appui] or '').strip() if col_num_appui < len(row) else ''
                nom_cable = str(row[col_nom_cable] or '').strip() if col_nom_cable >= 0 and col_nom_cable < len(row) else ''
                
                # Détecter appui EDF: colonne "Effort disponible avant ajout câble" renseignée
                effort_val = None
                if col_effort_dispo >= 0 and col_effort_dispo < len(row):
                    effort_val = row[col_effort_dispo]
                
                # Détecter pose boîtier (PB ou PEO)
                boitier_val = ''
                if col_pose_boitier >= 0 and col_pose_boitier < len(row):
                    bv = str(row[col_pose_boitier] or '').strip().upper()
                    if bv in ('PB', 'PEO'):
                        boitier_val = bv
                
                if raw_num:
                    current_appui = normalize_appui_num(raw_num)
                    if current_appui not in donnees_par_appui:
                        donnees_par_appui[current_appui] = []
                    # Appui EDF = effort disponible NON renseigné (pas de valeur)
                    # Appui Orange = effort disponible renseigné (a une valeur)
                    if col_effort_dispo >= 0 and (effort_val is None or str(effort_val).strip() == ''):
                        appuis_edf.add(current_appui)
                    # Enregistrer pose boîtier (première valeur non vide trouvée)
                    if boitier_val and current_appui not in boitier_par_appui:
                        boitier_par_appui[current_appui] = boitier_val
                
                if current_appui and nom_cable:
                    is_cable = bool(re.match(r'^L\d', nom_cable, re.IGNORECASE))
                    
                    if is_cable:
                        donnees_par_appui[current_appui].append(nom_cable)
                    
                    liste_brute.append({
                        'ligne': row_idx,
                        'num_appui': raw_num or current_appui,
                        'num_appui_norm': current_appui,
                        'nom_cable': nom_cable,
                        'is_cable': is_cable,
                        'is_edf': current_appui in appuis_edf
                    })
            
            # Exclure les appuis EDF (ceux SANS effort disponible)
            if appuis_edf:
                for appui_edf in appuis_edf:
                    donnees_par_appui.pop(appui_edf, None)
            
            wb.close()
            
            return donnees_par_appui, liste_brute, boitier_par_appui
            
        except Exception as e:
            QgsMessageLog.logMessage(
                f"Erreur lecture C6: {e}",
                "POLICE_C6", Qgis.Warning
            )
            return donnees_par_appui, liste_brute, boitier_par_appui

    def _find_column_index(self, headers: List[str], candidates: List[str]) -> int:
        """Trouve l'index d'une colonne parmi plusieurs noms possibles."""
        def normalize(s):
            """Supprime accents et caractères spéciaux pour comparaison."""
            import unicodedata
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return s.lower().strip()
        
        for candidate in candidates:
            c_norm = normalize(candidate)
            for idx, header in enumerate(headers):
                h_norm = normalize(header)
                if c_norm in h_norm or h_norm in c_norm:
                    return idx
        return -1

    # =========================================================================
    # ANALYSE APPUIS (conservé de l'ancienne version)
    # =========================================================================
    def analyseAppui(self, infra_pt_pot, champs, valeur, listeChampsAppui):
        """
        Analyse les appuis: compare QGIS avec Annexe C6.
        Conservé de l'ancienne version pour compatibilité.
        """
        self._check_cancel()
        
        # Expression de filtre
        if '/' in str(valeur):
            valeur_escaped = valeur.replace('/', '-')
            expr = QgsExpression(f'"{champs}" = \'{valeur}\' OR "{champs}" = \'{valeur_escaped}\'')
        else:
            expr = QgsExpression(f'"{champs}" = \'{valeur}\'')
        
        # Récupérer les appuis de l'étude
        appuis_qgis = {}
        request = QgsFeatureRequest(expr)
        
        for feat in infra_pt_pot.getFeatures(request):
            num = feat['pt_ad_numsu']
            if num and num != NULL:
                num_norm = normalize_appui_num(str(num))
                appuis_qgis[num_norm] = {
                    'feature': feat,
                    'num': str(num),
                    'geom': feat.geometry()
                }
        
        self._check_cancel()
        
        # Comparer avec C6
        nb_corresp = 0
        absents_c6 = []  # Appuis C6 absents de QGIS
        absents_qgis = []  # Appuis QGIS absents de C6
        
        # Appuis C6 présents dans QGIS
        for num_c6, data in self.donnees_c6.items():
            num_c6_norm = normalize_appui_num(num_c6)
            if num_c6_norm in appuis_qgis:
                nb_corresp += 1
            else:
                absents_c6.append(num_c6)
        
        # Appuis QGIS absents de C6
        for num_qgis in appuis_qgis.keys():
            if num_qgis not in [normalize_appui_num(k) for k in self.donnees_c6.keys()]:
                absents_qgis.append(num_qgis)
        
        self.nb_appui_corresp = nb_corresp
        self.nb_appui_absentPot = len(absents_c6)
        self.nb_appui_absent = len(absents_qgis)
        self.absence = absents_c6
        self.idPotAbsent = absents_qgis
        
        return nb_corresp, len(absents_c6), len(absents_qgis)

    # =========================================================================
    # ANALYSE BPE (conservé de l'ancienne version)
    # =========================================================================
    def analyseBPE(self, infra_pt_pot, bpe_layer, champs, valeur, listeChampsBpe):
        """
        Analyse les BPE: vérifie correspondance avec appuis.
        Conservé de l'ancienne version pour compatibilité.
        """
        self._check_cancel()
        
        if '/' in str(valeur):
            valeur_escaped = valeur.replace('/', '-')
            expr = QgsExpression(f'"{champs}" = \'{valeur}\' OR "{champs}" = \'{valeur_escaped}\'')
        else:
            expr = QgsExpression(f'"{champs}" = \'{valeur}\'')
        
        request = QgsFeatureRequest(expr)
        
        # Récupérer les appuis de l'étude
        appuis_etude = {}
        for feat in infra_pt_pot.getFeatures(request):
            num = feat['pt_ad_numsu']
            if num and num != NULL:
                appuis_etude[normalize_appui_num(str(num))] = feat.geometry()
        
        self._check_cancel()
        
        # Vérifier les BPE
        nb_pbo_ok = 0
        bpe_sans_appui = []
        bpe_appui_inconnu = []
        
        for feat in bpe_layer.getFeatures(request):
            bpe_geom = feat.geometry()
            bpe_id = feat['bp_pt_code'] if 'bp_pt_code' in [f.name() for f in bpe_layer.fields()] else feat.id()
            
            # Chercher appui le plus proche
            appui_trouve = False
            for num_appui, appui_geom in appuis_etude.items():
                if bpe_geom.distance(appui_geom) < 1.0:  # 1m de tolérance
                    # Vérifier si appui est dans C6
                    if num_appui in [normalize_appui_num(k) for k in self.donnees_c6.keys()]:
                        nb_pbo_ok += 1
                    else:
                        bpe_appui_inconnu.append({
                            'bpe_id': bpe_id,
                            'num_appui': num_appui
                        })
                    appui_trouve = True
                    break
            
            if not appui_trouve:
                bpe_sans_appui.append({
                    'bpe_id': bpe_id,
                    'geom': bpe_geom
                })
            
            self._check_cancel()
        
        self.nb_pbo_corresp = nb_pbo_ok
        self.ebp_non_appui = bpe_sans_appui
        self.ebp_appui_inconnu = bpe_appui_inconnu
        
        return nb_pbo_ok, len(bpe_sans_appui), len(bpe_appui_inconnu)

    # =========================================================================
    # UTILITAIRES
    # =========================================================================
    def removeGroup(self, name):
        """Suppression d'un groupe de couches."""
        remove_group(name)

    def layerGroupError(self, couche, nom_etude):
        """Ajoute une couche dans un groupe d'erreur."""
        layer_group_error(couche, nom_etude)

    def appliquerstyle(self, nom_style: str):
        """Applique un style QML à une couche."""
        style_path = os.path.join(STYLE_DIR, f"{nom_style}.qml")
        if os.path.exists(style_path):
            layers = QgsProject.instance().mapLayersByName(nom_style)
            for layer in layers:
                layer.loadNamedStyle(style_path)
                layer.triggerRepaint()

    def comparer_c6_cables(
        self,
        donnees_c6: Dict[str, List[str]],
        cables_par_appui: Dict[str, Dict]
    ) -> List[Dict]:
        """Compare cables C6 vs BDD. Delegue a cable_analyzer.comparer_source_cables()."""
        from .cable_analyzer import comparer_source_cables
        return comparer_source_cables(
            donnees_c6, cables_par_appui,
            source_label="C6",
            get_capacites_fn=get_capacites_possibles
        )

    @staticmethod
    def _capacites_compatibles(
        capas_possibles_c6: List[List[int]],
        capas_bdd: List[int]
    ) -> bool:
        """Delegue a cable_analyzer.capacites_compatibles()."""
        from .cable_analyzer import capacites_compatibles
        return capacites_compatibles(capas_possibles_c6, capas_bdd)
